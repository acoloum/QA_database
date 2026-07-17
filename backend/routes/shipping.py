from flask import Blueprint, jsonify, request, send_file
from ..services.shipping_service import ShippingService
from ..services.spc_report import SpcReportService
from ..utils import api_error, auth_required, require_perm, handle_db_error, validate_upload_file

shipping_bp = Blueprint('shipping', __name__)

@shipping_bp.route('/api/data', methods=['GET'])
@auth_required
def get_data():
    try:
        result = ShippingService.get_list(request.args)
        return jsonify(result)
    except Exception as e:
        return api_error(str(e), 500)

@shipping_bp.route('/api/data/<int:data_id>', methods=['GET'])
@auth_required
def get_shipping_data(data_id):
    """根據 ID 獲取單筆出貨檢驗資料"""
    try:
        result = ShippingService.get_by_id(data_id)
        if result is None:
            return api_error('資料不存在', 404)
        return jsonify(result)
    except Exception as e:
        return api_error(str(e), 500)

@shipping_bp.route('/api/stats', methods=['GET'])
@auth_required
def get_shipping_stats():
    """獲取出貨檢驗的 SPC 統計數據"""
    try:
        result = ShippingService.get_stats(request.args)
        return jsonify(result)
    except Exception as e:
        from flask import current_app
        current_app.logger.exception("Shipping error: %s", str(e))
        return api_error(str(e), 500)

@shipping_bp.route('/api/data/<int:data_id>/measurements', methods=['GET'])
@auth_required
def get_shipping_measurements(data_id):
    """取得單筆出貨記錄的量測明細（含離群標記）"""
    try:
        return jsonify(ShippingService.get_measurements(data_id))
    except Exception as e:
        return api_error(str(e), 500)


@shipping_bp.route('/api/measurements/<int:measurement_id>/exclusion', methods=['PATCH'])
@auth_required
def set_measurement_exclusion(measurement_id):
    """標示/解除量測值離群排除（AIAG-VDA SPC 2026 §6.6）"""
    try:
        body = request.get_json(silent=True) or {}
        result = ShippingService.set_measurement_exclusion(
            measurement_id,
            excluded=bool(body.get('排除統計')),
            reason=body.get('排除原因'),
        )
        return jsonify(result)
    except ValueError as e:
        return api_error(str(e), 400)
    except Exception as e:
        return api_error(str(e), 500)


@shipping_bp.route('/api/spc-report', methods=['GET'])
@auth_required
def export_spc_report():
    """匯出 SPC 統計分析報告 (Excel)，包含原始數據 + SPC 統計與圖表"""
    try:
        from openpyxl import load_workbook
        from copy import copy

        field = request.args.get('field', '外徑')
        filters = {
            'vendor': request.args.get('vendor', ''),
            'material': request.args.get('material', ''),
            'spec': request.args.get('spec', ''),
            'start_date': request.args.get('start_date', ''),
            'end_date': request.args.get('end_date', ''),
        }

        # 1. 產生原始數據工作簿
        raw_output = ShippingService.export_excel(request.args)
        wb_main = load_workbook(raw_output)
        # 將預設工作表重新命名為「原始數據」
        ws_raw = wb_main.active
        ws_raw.title = '原始數據'

        # 2. 產生 SPC 報告工作簿
        stats_data = ShippingService.get_stats(request.args)
        spc_output = SpcReportService.generate_report(stats_data, field, filters)
        wb_spc = load_workbook(spc_output)

        # 3. 將 SPC 報告的所有工作表複製到主工作簿
        for sheet_name in wb_spc.sheetnames:
            ws_src = wb_spc[sheet_name]
            ws_dst = wb_main.create_sheet(title=sheet_name)

            # 複製欄寬
            for col_letter, dim in ws_src.column_dimensions.items():
                ws_dst.column_dimensions[col_letter].width = dim.width

            # 複製列高
            for row_num, dim in ws_src.row_dimensions.items():
                ws_dst.row_dimensions[row_num].height = dim.height

            # 複製合併儲存格
            for merged_range in ws_src.merged_cells.ranges:
                ws_dst.merge_cells(str(merged_range))

            # 複製儲存格內容與樣式
            for row in ws_src.iter_rows():
                for cell in row:
                    new_cell = ws_dst.cell(row=cell.row, column=cell.column, value=cell.value)
                    if cell.has_style:
                        new_cell.font = copy(cell.font)
                        new_cell.border = copy(cell.border)
                        new_cell.fill = copy(cell.fill)
                        new_cell.number_format = cell.number_format
                        new_cell.protection = copy(cell.protection)
                        new_cell.alignment = copy(cell.alignment)

            # 複製圖表
            for chart in ws_src._charts:
                ws_dst.add_chart(chart)

        # 4. 輸出合併後的工作簿
        from io import BytesIO
        output = BytesIO()
        wb_main.save(output)
        output.seek(0)

        filename = f'出貨檢驗_SPC報告_{field}_{filters["material"] or "all"}.xlsx'
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        from flask import current_app
        current_app.logger.exception("Shipping error: %s", str(e))
        return api_error(str(e), 500)

@shipping_bp.route('/api/add', methods=['POST'])
@shipping_bp.route('/api/update', methods=['POST'])
@auth_required
@require_perm('shipping.create')
def save_data():
    try:
        is_update = request.path.endswith('update')
        ShippingService.save_data(request.json, is_update=is_update)
        return jsonify({"success": True})
    except ValueError as e:
        return api_error(str(e), 400)
    except Exception as e:
        db_err = handle_db_error(e)
        return api_error(db_err.get('message', '資料庫錯誤'), 500, detail=db_err)

@shipping_bp.route('/api/delete', methods=['POST'])
@auth_required
@require_perm('shipping.delete')
def delete_data():
    try:
        record_id = request.json.get('id')
        if not record_id:
            return api_error("缺少記錄 ID", 400)
        
        ShippingService.delete_data(record_id)
        return jsonify({"success": True})
    except Exception as e:
        return api_error(str(e), 500)

@shipping_bp.route('/api/import', methods=['POST', 'OPTIONS'])
@auth_required
@require_perm('shipping.create')
def shipping_import():
    if request.method == 'OPTIONS':
        return '', 200

    if 'file' not in request.files:
        return api_error("沒有上傳檔案", 400)

    file = request.files['file']
    if file.filename == '':
        return api_error("沒有選擇檔案", 400)

    upload_error = validate_upload_file(file)
    if upload_error:
        return api_error(upload_error, 400)

    try:
        count = ShippingService.import_data(file)
        return jsonify({"success": True, "message": f"匯入成功，共 {count} 筆資料"})
    except ValueError as e:
        return api_error(str(e), 400)
    except Exception as e:
        return api_error(str(e), 500)

@shipping_bp.route('/api/export/excel')
@auth_required
def export_excel():
    try:
        output = ShippingService.export_excel(request.args)
        return send_file(
            output, 
            as_attachment=True, 
            download_name='出貨檢驗數據.xlsx', 
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        return api_error(str(e), 500)

