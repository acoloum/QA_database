"""MSA 正式報告：只讀已保存的結果版本，絕不重新分析。

報告的價值在於「當時的證據長什麼樣」，因此所有內容都來自結果版本
自己保存的快照；設備改名、準則改版或觀測後續修正，都不會改變已
產生的報告。
"""

from decimal import Decimal
from io import BytesIO
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (
    KeepTogether,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from ..extensions import db
from ..models import MsaResultVersion, MsaWorkflowDecision, User
from .msa_errors import MsaNotFound, MsaValidationError


SHEET_NAMES = (
    "研究摘要", "研究設計", "原始讀值", "統計結果",
    "圖表資料", "設備與校驗", "準則與判定", "核准歷程", "版本稽核",
)

# 尚未核准的結果必須在報告上明確標示，避免被當成正式結論引用
UNAPPROVED_NOTICE = "草稿／未核准：本報告尚未完成核准，不得作為正式判定依據"

_HEADER_FILL = PatternFill("solid", fgColor="1F3B39")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_TITLE_FONT = Font(bold=True, size=14)
_WARNING_FONT = Font(color="A63440", bold=True)

_DISPOSITION_LABELS = {
    "acceptable": "可接受",
    "conditionally_acceptable": "條件接受",
    "unacceptable": "不可接受",
    "indeterminate": "無法判定",
}


def _disposition_label(conclusion: dict):
    disposition = conclusion.get("system_disposition")
    return _DISPOSITION_LABELS.get(disposition, disposition)

# 正式 PDF 必須能正確呈現繁體中文；找不到字型時不得以 Helvetica 輸出亂碼
CJK_FONT_CANDIDATES = (
    "C:/Windows/Fonts/msjh.ttc",
    "C:/Windows/Fonts/mingliu.ttc",
    "C:/Windows/Fonts/kaiu.ttf",
    "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
    "/usr/share/fonts/truetype/wqy/wqy-microhei.ttc",
)

# 各發行版的 Noto CJK 檔名會隨版本改變，因此除了明確候選之外再掃描
# 標準字型目錄；只要裝了任一 CJK 字型套件就能找到，不必綁死檔名。
CJK_FONT_SEARCH_DIRS = (
    "/usr/share/fonts",
    "/usr/local/share/fonts",
    "C:/Windows/Fonts",
)
CJK_FONT_PATTERNS = (
    "NotoSansCJK*", "NotoSerifCJK*", "NotoSansTC*", "SourceHanSans*",
    "wqy-*", "msjh*", "mingliu*", "kaiu*",
)

CJK_FONT_NAME = "MSA-CJK"
PDF_WATERMARK_TEXT = "未核准"


def _discover_cjk_fonts():
    """掃描標準字型目錄，回傳可能的 CJK 字型檔路徑。"""
    from glob import glob

    found = []
    for directory in CJK_FONT_SEARCH_DIRS:
        if not os.path.isdir(directory):
            continue
        for pattern in CJK_FONT_PATTERNS:
            for suffix in (".ttc", ".otf", ".ttf"):
                found.extend(sorted(glob(
                    os.path.join(directory, "**", pattern + suffix),
                    recursive=True,
                )))
    return found


def register_cjk_font() -> str:
    """註冊可用的繁體中文字型；找不到就明確失敗而不是輸出亂碼。"""
    if CJK_FONT_NAME in pdfmetrics.getRegisteredFontNames():
        return CJK_FONT_NAME

    for path in (*CJK_FONT_CANDIDATES, *_discover_cjk_fonts()):
        if not os.path.exists(path):
            continue
        try:
            # .ttc 內含多個字體，取第一個子字體即可
            font = (
                TTFont(CJK_FONT_NAME, path, subfontIndex=0)
                if path.lower().endswith(".ttc")
                else TTFont(CJK_FONT_NAME, path)
            )
            pdfmetrics.registerFont(font)
            return CJK_FONT_NAME
        except Exception:
            continue

    raise MsaValidationError(
        "MSA_REPORT_FONT_MISSING",
        "伺服器缺少可用的繁體中文字型，無法產生正式 PDF",
        details={
            "candidates": list(CJK_FONT_CANDIDATES),
            "searched_dirs": list(CJK_FONT_SEARCH_DIRS),
        },
    )


class MsaReportService:
    """由不可變結果版本產生可稽核的報告檔。"""

    @staticmethod
    def generate_excel(result_version_id: int) -> BytesIO:
        snapshot = MsaReportService._load_snapshot(result_version_id)
        workbook = Workbook()
        workbook.remove(workbook.active)
        for name in SHEET_NAMES:
            workbook.create_sheet(name)

        MsaReportService._write_summary(workbook["研究摘要"], snapshot)
        MsaReportService._write_design(workbook["研究設計"], snapshot)
        MsaReportService._write_observations(workbook["原始讀值"], snapshot)
        MsaReportService._write_statistics(workbook["統計結果"], snapshot)
        MsaReportService._write_charts(workbook["圖表資料"], snapshot)
        MsaReportService._write_equipment(workbook["設備與校驗"], snapshot)
        MsaReportService._write_criteria(workbook["準則與判定"], snapshot)
        MsaReportService._write_workflow(workbook["核准歷程"], snapshot)
        MsaReportService._write_audit(workbook["版本稽核"], snapshot)

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return output

    @staticmethod
    def generate_pdf(result_version_id: int) -> BytesIO:
        snapshot = MsaReportService._load_snapshot(result_version_id)
        font = register_cjk_font()
        result = snapshot["result"]
        study = snapshot["study"]
        approved = result.status == "approved"

        output = BytesIO()
        document = SimpleDocTemplate(
            output,
            pagesize=A4,
            title=f"MSA {study.get('study_no')} 量測系統分析報告",
            author="QC Database",
            subject=f"結果版本 {result.id}／資料雜湊 {result.data_hash[:12]}",
            leftMargin=18 * mm, rightMargin=18 * mm,
            topMargin=20 * mm, bottomMargin=18 * mm,
        )
        styles = MsaReportService._pdf_styles(font)
        story = MsaReportService._pdf_story(snapshot, styles, approved)

        footer = (
            f"研究 {study.get('study_no')}｜結果版本 {result.id}"
            f"｜資料雜湊 {result.data_hash[:12]}"
        )
        document.build(
            story,
            onFirstPage=lambda canvas, doc: _draw_page_furniture(
                canvas, doc, font, footer, approved,
            ),
            onLaterPages=lambda canvas, doc: _draw_page_furniture(
                canvas, doc, font, footer, approved,
            ),
        )
        output.seek(0)
        return output

    @staticmethod
    def _pdf_styles(font: str) -> dict:
        base = getSampleStyleSheet()
        return {
            "title": ParagraphStyle(
                "MsaTitle", parent=base["Title"], fontName=font,
                fontSize=18, leading=24,
            ),
            "heading": ParagraphStyle(
                "MsaHeading", parent=base["Heading2"], fontName=font,
                fontSize=13, leading=18, spaceBefore=10, spaceAfter=6,
                textColor=colors.HexColor("#1F3B39"),
            ),
            "body": ParagraphStyle(
                "MsaBody", parent=base["BodyText"], fontName=font,
                fontSize=9.5, leading=14,
            ),
            "notice": ParagraphStyle(
                "MsaNotice", parent=base["BodyText"], fontName=font,
                fontSize=11, leading=16,
                textColor=colors.HexColor("#A63440"),
            ),
        }

    @staticmethod
    def _pdf_story(snapshot, styles, approved: bool) -> list:
        result = snapshot["result"]
        study = snapshot["study"]
        plan = snapshot["plan"]
        conclusion = snapshot["conclusion"]
        criteria = snapshot["criteria"]
        font = styles["body"].fontName
        story = []

        # 1. 封面與研究結論
        story.append(Paragraph("MSA 量測系統分析報告", styles["title"]))
        if not approved:
            story.append(Paragraph(UNAPPROVED_NOTICE, styles["notice"]))
        story.append(Spacer(1, 6))
        story.append(_pdf_table([
            ("研究編號", study.get("study_no")),
            ("品質特性", study.get("characteristic")),
            ("單位", study.get("unit")),
            ("研究類型", study.get("study_type")),
            ("結果狀態", result.status),
            ("系統處置", _disposition_label(conclusion)),
            ("百分比口徑", conclusion.get("percent_basis")),
            ("工程判斷", conclusion.get("engineering_judgment") or "未附加"),
        ], font))
        for reason in conclusion.get("reasons") or []:
            story.append(Paragraph(f"• {reason}", styles["body"]))

        # 2. 研究設計與適用性
        story.append(Paragraph("研究設計與適用性", styles["heading"]))
        story.append(_pdf_table([
            ("方法代碼", plan.get("method_code")),
            ("方法版本", plan.get("method_version")),
            ("設計", (
                f"{plan.get('part_count')} 零件 × "
                f"{plan.get('appraiser_count')} 評價人 × "
                f"{plan.get('trial_count')} 試驗"
            )),
            ("計畫雜湊", plan.get("plan_hash")),
            ("凍結時間", plan.get("frozen_at")),
        ], font))
        applicability = result.applicability_result or {}
        story.append(_pdf_table(
            [(str(key), _pdf_text(value))
             for key, value in sorted(applicability.items())],
            font,
        ))

        # 3. 設備／校驗／解析度
        story.append(Paragraph("設備、校驗與解析度", styles["heading"]))
        equipment = plan.get("equipment_snapshot") or {}
        assessment = equipment.get("resolution_assessment") or {}
        story.append(_pdf_table([
            ("資格檢查日", equipment.get("checked_on")),
            ("解析度評估", assessment.get("level")),
            ("評估說明", assessment.get("reason")),
        ], font))
        story.append(_pdf_grid(
            ["角色", "設備編號", "名稱", "解析度", "校驗結果", "下次校驗日"],
            [
                [
                    item.get("role"), item.get("equipment_no"),
                    item.get("name"), item.get("resolution"),
                    (item.get("calibration") or {}).get("result"),
                    (item.get("calibration") or {}).get("next_due_date"),
                ]
                for item in equipment.get("items") or []
            ],
            font,
        ))

        story.append(PageBreak())

        # 4. 原始資料摘要
        story.append(Paragraph("原始資料摘要", styles["heading"]))
        observations = snapshot["observations"]
        story.append(Paragraph(
            f"有效觀測共 {len(observations)} 筆，依實際輸入順序保存。",
            styles["body"],
        ))
        story.append(_pdf_grid(
            ["順序", "零件", "評價人", "試驗", "讀值"],
            [
                [
                    row.get("actual_entry_order"), row.get("part_blind_code"),
                    row.get("appraiser_blind_code"), row.get("trial_no"),
                    row.get("numeric_value") or row.get("attribute_value"),
                ]
                for row in observations[:40]
            ],
            font,
        ))
        if len(observations) > 40:
            story.append(Paragraph(
                f"（此處僅列前 40 筆，完整讀值請見 Excel 報告）",
                styles["body"],
            ))

        # 5. 方法統計證據
        story.append(Paragraph("統計證據", styles["heading"]))
        rows = []
        _flatten(snapshot["statistics"], "", rows)
        story.append(_pdf_grid(
            ["統計項目", "值"],
            [[path, _pdf_text(value)] for path, value in rows],
            font,
        ))

        # 6. 圖表資料
        story.append(PageBreak())
        story.append(Paragraph("圖表資料", styles["heading"]))
        story.append(Paragraph(
            "以下數值直接取自結果版本保存的圖表資料，與畫面顯示相同。",
            styles["body"],
        ))
        chart_rows = []
        _flatten(snapshot["charts"], "", chart_rows)
        story.append(_pdf_grid(
            ["圖表資料路徑", "值"],
            [[path, _pdf_text(value)] for path, value in chart_rows[:60]],
            font,
        ))

        # 7. 準則與三層判定
        story.append(Paragraph("準則與三層判定", styles["heading"]))
        story.append(_pdf_table([
            ("準則設定", criteria.get("profile_name")),
            ("準則版本", criteria.get("version_no")),
            ("生效日", criteria.get("effective_date")),
            ("百分比口徑", criteria.get("percent_basis")),
            ("口徑來源", criteria.get("percent_basis_source")),
        ], font))
        story.append(_pdf_grid(
            ["門檻", "值"],
            [
                [key, _pdf_text(value)]
                for key, value in sorted(
                    (criteria.get("thresholds") or {}).items()
                )
            ],
            font,
        ))
        story.append(_pdf_table([
            ("客觀統計結果", _pdf_text(
                conclusion.get("statistical_result")
            )),
            ("系統處置", conclusion.get("system_disposition")),
            ("人工工程判斷", conclusion.get("engineering_judgment") or "未附加"),
        ], font))

        # 8. 核准與稽核歷程
        story.append(Paragraph("核准與稽核歷程", styles["heading"]))
        story.append(_pdf_grid(
            ["動作", "狀態轉移", "執行者", "時間", "理由"],
            [
                [
                    row.get("action"),
                    f"{row.get('from_status')} → {row.get('to_status')}",
                    row.get("actor"), row.get("created_at"),
                    row.get("reason"),
                ]
                for row in snapshot["workflow"]
            ],
            font,
        ))
        story.append(KeepTogether(_pdf_table([
            ("結果版本ID", result.id),
            ("資料雜湊", result.data_hash),
            ("方法代碼", result.method_code),
            ("方法版本", result.method_version),
            ("程式版本", result.code_version),
        ], font)))
        return story

    # ------------------------------------------------------------------
    # 快照載入
    # ------------------------------------------------------------------

    @staticmethod
    def _load_snapshot(result_version_id: int) -> dict:
        result = db.session.get(MsaResultVersion, result_version_id)
        if result is None:
            raise MsaNotFound(
                "MSA_RESULT_NOT_FOUND",
                "找不到 MSA 結果版本",
                details={"result_version_id": result_version_id},
            )

        summary = result.raw_data_summary or {}
        missing = [
            key for key in ("study", "plan", "effective_observations")
            if not summary.get(key)
        ]
        if missing:
            raise MsaValidationError(
                "MSA_REPORT_EVIDENCE_INCOMPLETE",
                "結果版本缺少完整證據快照，且不得回頭重算",
                details={"missing": missing},
            )

        return {
            "result": result,
            "study": summary["study"],
            "plan": summary["plan"],
            "observations": summary["effective_observations"],
            "method": summary.get("method") or {},
            "statistics": result.statistics or {},
            "charts": result.chart_data or {},
            "criteria": result.criteria_snapshot or {},
            "conclusion": result.conclusion or {},
            "warnings": result.warnings or [],
            "blockers": result.blockers or [],
            "workflow": MsaReportService._serialize_decisions(result),
        }

    @staticmethod
    def _serialize_decisions(result: MsaResultVersion) -> list[dict]:
        decisions = (
            MsaWorkflowDecision.query
            .filter_by(study_id=result.study_id)
            .order_by(MsaWorkflowDecision.id.asc())
            .all()
        )
        actor_names = {
            user.id: user.username
            for user in User.query.filter(
                User.id.in_({row.actor_id for row in decisions} or {0})
            ).all()
        }
        return [
            {
                "action": row.action,
                "from_status": row.from_status,
                "to_status": row.to_status,
                "reason": row.reason,
                "actor_id": row.actor_id,
                "actor": actor_names.get(row.actor_id),
                "created_at": (
                    row.created_at.isoformat() if row.created_at else None
                ),
                "separation_check": row.separation_check or {},
                "extra_data": row.extra_data or {},
            }
            for row in decisions
        ]

    # ------------------------------------------------------------------
    # 工作表
    # ------------------------------------------------------------------

    @staticmethod
    def _write_summary(sheet, snapshot) -> None:
        result = snapshot["result"]
        study = snapshot["study"]
        conclusion = snapshot["conclusion"]

        sheet["A1"] = "MSA 量測系統分析報告"
        sheet["A1"].font = _TITLE_FONT
        row = 3
        if result.status != "approved":
            sheet.cell(row=row, column=1, value=UNAPPROVED_NOTICE).font = (
                _WARNING_FONT
            )
            row += 2

        pairs = [
            ("研究編號", study.get("study_no")),
            ("研究類型", study.get("study_type")),
            ("量測目的", study.get("measurement_purpose")),
            ("品質特性", study.get("characteristic")),
            ("單位", study.get("unit")),
            ("規格下限", _number(study.get("lsl"))),
            ("規格上限", _number(study.get("usl"))),
            ("結果狀態", result.status),
            ("系統處置", _disposition_label(conclusion)),
            ("百分比口徑", conclusion.get("percent_basis")),
            ("工程判斷", conclusion.get("engineering_judgment") or "未附加"),
        ]
        row = _write_pairs(sheet, pairs, start_row=row)

        row += 1
        sheet.cell(row=row, column=1, value="判定理由").font = _HEADER_FONT
        sheet.cell(row=row, column=1).fill = _HEADER_FILL
        for reason in conclusion.get("reasons") or []:
            row += 1
            sheet.cell(row=row, column=1, value=_safe(reason))
        for item in snapshot["blockers"]:
            row += 1
            sheet.cell(
                row=row, column=1,
                value=_safe(f"阻擋：{item.get('message')}"),
            ).font = _WARNING_FONT
        for item in snapshot["warnings"]:
            row += 1
            sheet.cell(row=row, column=1, value=_safe(f"警告：{item.get('message')}"))
        _autosize(sheet)

    @staticmethod
    def _write_design(sheet, snapshot) -> None:
        plan = snapshot["plan"]
        row = _write_pairs(sheet, [
            ("計畫版本號", plan.get("plan_version_no")),
            ("方法代碼", plan.get("method_code")),
            ("方法版本", plan.get("method_version")),
            ("設計型態", plan.get("design_type")),
            ("零件數", plan.get("part_count")),
            ("評價人數", plan.get("appraiser_count")),
            ("試驗次數", plan.get("trial_count")),
            ("隨機種子", plan.get("random_seed")),
            ("計畫雜湊", plan.get("plan_hash")),
            ("凍結時間", plan.get("frozen_at")),
        ], start_row=1)

        row += 2
        _write_table(
            sheet,
            headers=("零件序號", "盲碼", "零件識別", "參考值", "參考分類"),
            rows=[
                (
                    part.get("part_no"), part.get("blind_code"),
                    _safe(part.get("part_identifier")),
                    _number(part.get("reference_value")),
                    _safe(part.get("reference_category")),
                )
                for part in plan.get("parts") or []
            ],
            start_row=row,
        )
        row += len(plan.get("parts") or []) + 3
        _write_table(
            sheet,
            headers=("評價人序號", "盲碼", "姓名"),
            rows=[
                (
                    row_data.get("appraiser_no"), row_data.get("blind_code"),
                    _safe(row_data.get("name")),
                )
                for row_data in plan.get("appraisers") or []
            ],
            start_row=row,
        )
        _autosize(sheet)

    @staticmethod
    def _write_observations(sheet, snapshot) -> None:
        """原始讀值直接來自結果快照，不重新查詢目前有效觀測。"""
        _write_table(
            sheet,
            headers=(
                "要求順序", "實際輸入順序", "零件盲碼", "評價人盲碼",
                "試驗次數", "計量讀值", "計數分類", "量測時間", "來源",
            ),
            rows=[
                (
                    row.get("requested_order"), row.get("actual_entry_order"),
                    row.get("part_blind_code"), row.get("appraiser_blind_code"),
                    row.get("trial_no"), _number(row.get("numeric_value")),
                    _safe(row.get("attribute_value")),
                    _safe(row.get("measured_at")), row.get("source"),
                )
                for row in snapshot["observations"]
            ],
            start_row=1,
            freeze=True,
            autofilter=True,
        )
        _autosize(sheet)

    @staticmethod
    def _write_statistics(sheet, snapshot) -> None:
        rows = []
        _flatten(snapshot["statistics"], "", rows)
        _write_table(
            sheet,
            headers=("統計項目", "值"),
            rows=[(_safe(path), _cell(value)) for path, value in rows],
            start_row=1,
            freeze=True,
            autofilter=True,
        )
        _autosize(sheet)

    @staticmethod
    def _write_charts(sheet, snapshot) -> None:
        """保存畫面實際使用的 series，報告與畫面必須是同一份數字。"""
        rows = []
        _flatten(snapshot["charts"], "", rows)
        _write_table(
            sheet,
            headers=("圖表資料路徑", "值"),
            rows=[(_safe(path), _cell(value)) for path, value in rows],
            start_row=1,
            freeze=True,
            autofilter=True,
        )
        _autosize(sheet)

    @staticmethod
    def _write_equipment(sheet, snapshot) -> None:
        equipment = (snapshot["plan"].get("equipment_snapshot") or {})
        items = equipment.get("items") or []
        row = _write_pairs(sheet, [
            ("資格檢查日", equipment.get("checked_on")),
            (
                "解析度評估",
                (equipment.get("resolution_assessment") or {}).get("level"),
            ),
            (
                "評估說明",
                (equipment.get("resolution_assessment") or {}).get("reason"),
            ),
        ], start_row=1)

        row += 2
        _write_table(
            sheet,
            headers=(
                "角色", "量測模式", "設備編號", "設備名稱", "狀態",
                "解析度", "單位", "校驗結果", "下次校驗日",
            ),
            rows=[
                (
                    item.get("role"), item.get("measurement_mode"),
                    _safe(item.get("equipment_no")), _safe(item.get("name")),
                    item.get("status"), _number(item.get("resolution")),
                    _safe(item.get("unit")),
                    _safe((item.get("calibration") or {}).get("result")),
                    _safe((item.get("calibration") or {}).get("next_due_date")),
                )
                for item in items
            ],
            start_row=row,
        )
        _autosize(sheet)

    @staticmethod
    def _write_criteria(sheet, snapshot) -> None:
        criteria = snapshot["criteria"]
        row = _write_pairs(sheet, [
            ("準則設定", _safe(criteria.get("profile_name"))),
            ("準則版本ID", criteria.get("version_id")),
            ("準則版本號", criteria.get("version_no")),
            ("方法版本", _safe(criteria.get("method_version"))),
            ("生效日", _safe(criteria.get("effective_date"))),
            ("百分比口徑", criteria.get("percent_basis")),
            ("口徑來源", criteria.get("percent_basis_source")),
            ("依據", _safe(criteria.get("basis"))),
        ], start_row=1)

        row += 2
        _write_table(
            sheet,
            headers=("門檻", "值"),
            rows=[
                (_safe(key), _cell(value))
                for key, value in sorted(
                    (criteria.get("thresholds") or {}).items()
                )
            ],
            start_row=row,
        )
        row += len(criteria.get("thresholds") or {}) + 3
        _write_table(
            sheet,
            headers=("條件接受必要措施",),
            rows=[
                (_safe(action),)
                for action in criteria.get("conditional_actions") or []
            ],
            start_row=row,
        )
        _autosize(sheet)

    @staticmethod
    def _write_workflow(sheet, snapshot) -> None:
        _write_table(
            sheet,
            headers=(
                "動作", "原狀態", "新狀態", "執行者", "執行者ID",
                "時間", "理由", "職責分離檢查",
            ),
            rows=[
                (
                    row.get("action"), row.get("from_status"),
                    row.get("to_status"), _safe(row.get("actor")),
                    row.get("actor_id"), _safe(row.get("created_at")),
                    _safe(row.get("reason")),
                    _safe(_compact(row.get("separation_check"))),
                )
                for row in snapshot["workflow"]
            ],
            start_row=1,
            freeze=True,
            autofilter=True,
        )
        _autosize(sheet)

    @staticmethod
    def _write_audit(sheet, snapshot) -> None:
        result = snapshot["result"]
        _write_pairs(sheet, [
            ("結果版本ID", result.id),
            ("結果版本號", result.result_version_no),
            ("研究ID", result.study_id),
            ("計畫版本ID", result.plan_version_id),
            ("資料雜湊", result.data_hash),
            ("方法代碼", result.method_code),
            ("方法版本", result.method_version),
            ("程式版本", result.code_version),
            ("結果狀態", result.status),
            ("建立者ID", result.created_by_id),
            (
                "建立時間",
                result.created_at.isoformat() if result.created_at else None,
            ),
        ], start_row=1)
        _autosize(sheet)


# ----------------------------------------------------------------------
# 共用寫入
# ----------------------------------------------------------------------


def _safe(value):
    """避免試算表把使用者文字當公式執行。"""
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


def _number(value):
    """數值保留原精度，不先轉成顯示字串。"""
    if value is None or value == "":
        return None
    if isinstance(value, (int, float, Decimal)):
        return value
    try:
        return float(Decimal(str(value)))
    except Exception:
        return _safe(value)


def _cell(value):
    if isinstance(value, (dict, list, tuple)):
        return _safe(_compact(value))
    if isinstance(value, bool) or value is None:
        return value
    if isinstance(value, (int, float, Decimal)):
        return value
    return _safe(str(value))


def _compact(value) -> str:
    import json

    return json.dumps(value, ensure_ascii=False, sort_keys=True)


def _flatten(value, prefix: str, rows: list) -> None:
    """把巢狀統計輸出攤平成可稽核的路徑／值配對。"""
    if isinstance(value, dict):
        for key in sorted(value, key=str):
            _flatten(value[key], f"{prefix}.{key}" if prefix else str(key), rows)
        return
    if isinstance(value, (list, tuple)):
        if value and all(
            not isinstance(item, (dict, list, tuple)) for item in value
        ):
            rows.append((prefix, _compact(list(value))))
            return
        for index, item in enumerate(value):
            _flatten(item, f"{prefix}[{index}]", rows)
        return
    rows.append((prefix, value))


def _write_pairs(sheet, pairs, *, start_row: int) -> int:
    row = start_row
    for label, value in pairs:
        sheet.cell(row=row, column=1, value=label).font = Font(bold=True)
        sheet.cell(row=row, column=2, value=_cell(value))
        row += 1
    return row


def _write_table(
    sheet, *, headers, rows, start_row: int,
    freeze: bool = False, autofilter: bool = False,
) -> None:
    for index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=start_row, column=index, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(vertical="center")
    for offset, values in enumerate(rows, start=1):
        for index, value in enumerate(values, start=1):
            sheet.cell(row=start_row + offset, column=index, value=value)
    if autofilter and rows:
        last_column = get_column_letter(len(headers))
        sheet.auto_filter.ref = (
            f"A{start_row}:{last_column}{start_row + len(rows)}"
        )
    if freeze:
        sheet.freeze_panes = sheet.cell(row=start_row + 1, column=1)


# ----------------------------------------------------------------------
# PDF 版面
# ----------------------------------------------------------------------


def _pdf_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, (dict, list, tuple)):
        return _compact(value)
    if isinstance(value, float):
        return f"{value:.10g}"
    return str(value)


def _pdf_table(pairs, font: str) -> Table:
    table = Table(
        [[str(label), Paragraph(
            _pdf_text(value),
            ParagraphStyle("cell", fontName=font, fontSize=9, leading=12),
        )] for label, value in pairs],
        colWidths=[38 * mm, 132 * mm],
        hAlign="LEFT",
    )
    table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), font),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#42605E")),
        ("LINEBELOW", (0, 0), (-1, -1), 0.25, colors.HexColor("#D5DEDE")),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
    ]))
    return table


def _pdf_grid(headers, rows, font: str) -> Table:
    style = ParagraphStyle("grid", fontName=font, fontSize=8, leading=11)
    data = [[Paragraph(str(header), style) for header in headers]]
    for row in rows:
        data.append([Paragraph(_pdf_text(cell), style) for cell in row])
    if len(data) == 1:
        data.append([Paragraph("（無資料）", style)] + [
            Paragraph("", style) for _ in headers[1:]
        ])

    width = 174 * mm
    table = Table(
        data, colWidths=[width / len(headers)] * len(headers),
        hAlign="LEFT", repeatRows=1,
    )
    table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), font),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F3B39")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D5DEDE")),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
    ]))
    return table


def _draw_page_furniture(canvas, doc, font, footer, approved) -> None:
    """每頁都帶研究編號、結果版本與資料雜湊短碼；未核准另加浮水印。"""
    canvas.saveState()
    if not approved:
        canvas.setFont(font, 60)
        canvas.setFillColor(colors.HexColor("#A63440"))
        canvas.setFillAlpha(0.16)
        canvas.translate(A4[0] / 2, A4[1] / 2)
        canvas.rotate(45)
        canvas.drawCentredString(0, 0, PDF_WATERMARK_TEXT)
        canvas.restoreState()
        canvas.saveState()

    canvas.setFont(font, 7.5)
    canvas.setFillColor(colors.HexColor("#5D6F70"))
    canvas.drawString(18 * mm, 10 * mm, footer)
    canvas.drawRightString(
        A4[0] - 18 * mm, 10 * mm, f"第 {canvas.getPageNumber()} 頁",
    )
    canvas.restoreState()


def _autosize(sheet) -> None:
    widths = {}
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            length = min(len(str(cell.value)), 60)
            widths[cell.column] = max(widths.get(cell.column, 10), length + 2)
    for column, width in widths.items():
        sheet.column_dimensions[get_column_letter(column)].width = width
