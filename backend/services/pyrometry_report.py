"""CQI-9 爐溫測試報告產生器 — 比照 QRA073(TUS) / QRA074(SAT) 格式

依現行報告格式輸出：
  TUS → QRA073 均勻性紀錄表 + 原始數據與曲線圖（超限點紅標）
  SAT → QRA074 系統準確度紀錄表
"""
from typing import Dict, Any, Optional
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.chart import LineChart, Reference

_THIN = Side(style='thin', color='999999')
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
_HDR_FILL = PatternFill('solid', fgColor='D9E1F2')
_LABEL_FILL = PatternFill('solid', fgColor='F2F2F2')
_RED_FILL = PatternFill('solid', fgColor='F8D7DA')    # 超上限
_BLUE_FILL = PatternFill('solid', fgColor='CFE2FF')   # 低於下限
_TITLE_FONT = Font(bold=True, size=14)
_BOLD = Font(bold=True)


def _num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _set(ws, coord, value, *, font=None, fill=None, align=None, border=True):
    c = ws[coord]
    c.value = value
    if font:
        c.font = font
    if fill:
        c.fill = fill
    c.alignment = align or _CENTER
    if border:
        c.border = _BORDER
    return c


# ----------------------------------------------------------------------------
def build_tus_sheet(wb, detail: Dict[str, Any], meta: Dict[str, Any], tc_corr: float):
    """QRA073 爐具溫度均勻性評估(TUS)紀錄表"""
    main = detail["main"]
    ws = wb.create_sheet("QRA073-TUS均勻性")
    setpoint = _num(main.get("設定溫度")) or 0
    tol = _num(main.get("允許公差")) or 0
    meta = meta or {}

    widths = [16, 11, 11, 8, 14, 14, 11, 11, 11, 11, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    # 標題
    ws.merge_cells("A1:K1")
    _set(ws, "A1", "必榮實業股份有限公司", font=_TITLE_FONT)
    ws.merge_cells("A2:I2")
    _set(ws, "A2", "爐具溫度均勻性評估(TUS)紀錄表", font=Font(bold=True, size=12))
    _set(ws, "J2", "版次：A0", font=_BOLD)
    ws.merge_cells("J2:K2")

    # 參數區
    _set(ws, "A3", "爐具測試溫度 (D) °C", fill=_LABEL_FILL)
    _set(ws, "B3", setpoint)
    _set(ws, "C3", "測試條件", fill=_LABEL_FILL)
    _set(ws, "D3", meta.get("測試條件", ""))
    _set(ws, "E3", "溫度控制器設定值(SV)", fill=_LABEL_FILL)
    _set(ws, "F3", meta.get("控制器設定值", ""))
    _set(ws, "G3", "控制器器示值補償", fill=_LABEL_FILL)
    _set(ws, "H3", meta.get("控制器補償", ""))
    _set(ws, "I3", "執行單位/TAF", fill=_LABEL_FILL)
    ws.merge_cells("J3:K3")
    _set(ws, "J3", meta.get("TAF編號", meta.get("執行單位", "")))

    _set(ws, "A4", "客戶名稱", fill=_LABEL_FILL)
    _set(ws, "B4", meta.get("客戶名稱", ""))
    _set(ws, "C4", "工件料號", fill=_LABEL_FILL)
    _set(ws, "D4", meta.get("工件料號", ""))
    _set(ws, "E4", "生產日期/批號", fill=_LABEL_FILL)
    _set(ws, "F4", meta.get("生產批號", ""))
    _set(ws, "G4", "內外徑尺寸", fill=_LABEL_FILL)
    _set(ws, "H4", meta.get("內外徑尺寸", ""))
    _set(ws, "I4", "長度/支數", fill=_LABEL_FILL)
    _set(ws, "J4", meta.get("長度支數", ""))
    _set(ws, "K4", meta.get("預估總重量", ""))

    _set(ws, "A5", "爐具編號", fill=_LABEL_FILL)
    _set(ws, "B5", main.get("爐號", ""))
    _set(ws, "C5", "測試日期", fill=_LABEL_FILL)
    _set(ws, "D5", main.get("測試日期", ""))
    _set(ws, "E5", "季別", fill=_LABEL_FILL)
    _set(ws, "F5", main.get("季別", ""))
    _set(ws, "G5", "測試人員", fill=_LABEL_FILL)
    _set(ws, "H5", main.get("測試人員姓名", ""))
    _set(ws, "I5", "測試儀器編號", fill=_LABEL_FILL)
    ws.merge_cells("J5:K5")
    _set(ws, "J5", main.get("測試儀器編號", ""))

    # 測試時間（由曲線資料推估）
    cd = detail.get("曲線資料") or {}
    soak_txt = ""
    if cd.get("時間"):
        s = cd.get("穩定開始", 0)
        e = cd.get("穩定結束", len(cd["時間"]) - 1)
        try:
            soak_txt = f"恆溫穩定期：{cd['時間'][s]} ~ {cd['時間'][e]}（共 {e - s + 1} 點）"
        except Exception:
            soak_txt = ""
    ws.merge_cells("A6:K6")
    _set(ws, "A6", soak_txt, align=_LEFT, fill=_LABEL_FILL)

    # 量測點表頭
    hdr_row = 8
    headers = [
        "測試熱電偶\n位置/編號",
        "綜合顯示\nMin (A)", "綜合顯示\nMax (A)", "單位",
        "溫度計補償\n(B)", "熱電偶補償\n(C)",
        "修正結果 Min\n(A+B+C=F)", "修正結果 Max\n(A+B+C=F)",
        "誤差 Min\n(F-D=E)", "誤差 Max\n(F-D=E)", "判定",
    ]
    for i, h in enumerate(headers, start=1):
        _set(ws, f"{chr(64 + i)}{hdr_row}", h, font=_BOLD, fill=_HDR_FILL)

    # 各點資料
    r = hdr_row + 1
    worst = None      # 全數最大誤差（絕對值最大、保留正負）
    for p in detail["tus_points"]:
        raw_min = _num(p.get("最低溫"))
        raw_max = _num(p.get("最高溫"))
        total = _num(p.get("修正值")) or 0
        c_tc = tc_corr or 0
        b_rec = round(total - c_tc, 2)
        g_min = round(raw_min + total, 2) if raw_min is not None else None
        g_max = round(raw_max + total, 2) if raw_max is not None else None
        e_min = round(g_min - setpoint, 2) if g_min is not None else None
        e_max = round(g_max - setpoint, 2) if g_max is not None else None
        for v in (e_min, e_max):
            if v is not None and (worst is None or abs(v) > abs(worst)):
                worst = v
        ok = p.get("是否合格", True)
        vals = [p.get("點位", ""), raw_min, raw_max, "°C", b_rec, c_tc,
                g_min, g_max, e_min, e_max, "合格" if ok else "不合格"]
        for i, v in enumerate(vals, start=1):
            cell = _set(ws, f"{chr(64 + i)}{r}", v)
            if not ok and i == 11:
                cell.fill = _RED_FILL
        r += 1

    # 判定列
    r += 1
    _set(ws, f"A{r}", "TUS全範圍最大誤差值(°C)", font=_BOLD, fill=_LABEL_FILL)
    ws.merge_cells(f"A{r}:C{r}")
    _set(ws, f"D{r}", worst)
    _set(ws, f"E{r}", "CQI-9允許(°C)", fill=_LABEL_FILL)
    _set(ws, f"F{r}", f"±{tol}")
    _set(ws, f"G{r}", "判定", fill=_LABEL_FILL)
    ws.merge_cells(f"H{r}:K{r}")
    jc = _set(ws, f"H{r}", "合格" if main.get("是否合格") else "不合格", font=_BOLD)
    jc.fill = _RED_FILL if not main.get("是否合格") else PatternFill('solid', fgColor='D1E7DD')

    # 均勻度資訊
    r += 1
    _set(ws, f"A{r}", "均勻度極差", fill=_LABEL_FILL)
    _set(ws, f"B{r}", main.get("TUS均勻度極差"))
    _set(ws, f"C{r}", "最大正偏差", fill=_LABEL_FILL)
    _set(ws, f"D{r}", main.get("TUS最大正偏差"))
    _set(ws, f"E{r}", "最大負偏差", fill=_LABEL_FILL)
    _set(ws, f"F{r}", main.get("TUS最大負偏差"))

    # 簽核列
    r += 2
    _set(ws, f"A{r}", "核准", fill=_LABEL_FILL)
    _set(ws, f"B{r}", meta.get("核准", ""))
    _set(ws, f"C{r}", "製表", fill=_LABEL_FILL)
    _set(ws, f"D{r}", meta.get("製表", ""))
    _set(ws, f"E{r}", "備註", fill=_LABEL_FILL)
    ws.merge_cells(f"F{r}:J{r}")
    _set(ws, f"F{r}", main.get("備註", ""), align=_LEFT)
    _set(ws, f"K{r}", "QRA073", font=_BOLD)


# ----------------------------------------------------------------------------
def build_sat_sheet(wb, detail: Dict[str, Any], meta: Dict[str, Any], tc_corr: float):
    """QRA074 爐具系統準確度測試(SAT)紀錄表"""
    main = detail["main"]
    ws = wb.create_sheet("QRA074-SAT準確度")
    setpoint = _num(main.get("設定溫度")) or 0
    tol = _num(main.get("允許公差")) or 0
    meta = meta or {}

    widths = [16, 13, 11, 11, 8, 14, 11, 11, 11, 11, 11, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    ws.merge_cells("A1:L1")
    _set(ws, "A1", "必榮實業股份有限公司", font=_TITLE_FONT)
    ws.merge_cells("A2:J2")
    _set(ws, "A2", "爐具系統準確度測試(SAT)紀錄表", font=Font(bold=True, size=12))
    ws.merge_cells("K2:L2")
    _set(ws, "K2", "版次：A0", font=_BOLD)

    _set(ws, "A3", "設定測試溫度(E)", fill=_LABEL_FILL)
    _set(ws, "B3", setpoint)
    _set(ws, "C3", "測試日期", fill=_LABEL_FILL)
    _set(ws, "D3", main.get("測試日期", ""))
    _set(ws, "E3", "季別", fill=_LABEL_FILL)
    _set(ws, "F3", main.get("季別", ""))
    _set(ws, "G3", "溫度/濕度", fill=_LABEL_FILL)
    ws.merge_cells("H3:I3")
    _set(ws, "H3", meta.get("溫濕度", ""))
    _set(ws, "J3", "執行單位", fill=_LABEL_FILL)
    ws.merge_cells("K3:L3")
    _set(ws, "K3", meta.get("執行單位", ""))

    _set(ws, "A4", "爐具編號", fill=_LABEL_FILL)
    _set(ws, "B4", main.get("爐號", ""))
    _set(ws, "C4", "控制器型號/規格", fill=_LABEL_FILL)
    ws.merge_cells("D4:F4")
    _set(ws, "D4", meta.get("控制器型號", ""))
    _set(ws, "G4", "測試人員", fill=_LABEL_FILL)
    _set(ws, "H4", main.get("測試人員姓名", ""))
    _set(ws, "I4", "測試儀器編號", fill=_LABEL_FILL)
    ws.merge_cells("J4:L4")
    _set(ws, "J4", main.get("測試儀器編號", ""))

    hdr_row = 6
    headers = [
        "控溫區", "控制儀表\n讀值", "校正測試\n讀值", "差值", "單位",
        "熱電偶補償\n(C)", "修正後\n讀值", "偏差", "判定",
    ]
    cols = ["A", "B", "C", "D", "E", "F", "G", "H", "I"]
    for col, h in zip(cols, headers):
        _set(ws, f"{col}{hdr_row}", h, font=_BOLD, fill=_HDR_FILL)

    r = hdr_row + 1
    worst = None
    for p in detail["sat_points"]:
        ctrl = _num(p.get("控制儀表讀值"))
        test = _num(p.get("校正測試儀表讀值"))
        diff = _num(p.get("差值"))
        corr = _num(p.get("修正值")) or 0
        dev = _num(p.get("偏差"))
        corrected = round(test + corr, 2) if test is not None else None
        if dev is not None and (worst is None or abs(dev) > abs(worst)):
            worst = dev
        ok = p.get("是否合格", True)
        vals = [p.get("控溫區", ""), ctrl, test, diff, "°C", corr, corrected, dev,
                "合格" if ok else "不合格"]
        for col, v in zip(cols, vals):
            cell = _set(ws, f"{col}{r}", v)
            if not ok and col == "I":
                cell.fill = _RED_FILL
        r += 1

    r += 1
    _set(ws, f"A{r}", "SAT最大誤差值(°C)", font=_BOLD, fill=_LABEL_FILL)
    ws.merge_cells(f"A{r}:B{r}")
    _set(ws, f"C{r}", worst)
    _set(ws, f"D{r}", "CQI-9允許(°C)", fill=_LABEL_FILL)
    _set(ws, f"E{r}", f"±{tol}")
    _set(ws, f"F{r}", "判定", fill=_LABEL_FILL)
    ws.merge_cells(f"G{r}:I{r}")
    jc = _set(ws, f"G{r}", "合格" if main.get("是否合格") else "不合格", font=_BOLD)
    jc.fill = _RED_FILL if not main.get("是否合格") else PatternFill('solid', fgColor='D1E7DD')

    r += 2
    _set(ws, f"A{r}", "核准", fill=_LABEL_FILL)
    _set(ws, f"B{r}", meta.get("核准", ""))
    _set(ws, f"C{r}", "製表", fill=_LABEL_FILL)
    _set(ws, f"D{r}", meta.get("製表", ""))
    _set(ws, f"E{r}", "備註", fill=_LABEL_FILL)
    ws.merge_cells(f"F{r}:H{r}")
    _set(ws, f"F{r}", main.get("備註", ""), align=_LEFT)
    _set(ws, f"I{r}", "QRA074", font=_BOLD)


# ----------------------------------------------------------------------------
def build_raw_chart_sheet(wb, chart_data: Dict[str, Any], setpoint: float, tol: float):
    """原始數據（時間×通道）+ 溫度曲線圖；恆溫穩定期內超限點以紅底標示"""
    if not chart_data or not chart_data.get("時間"):
        return
    times = chart_data["時間"]
    values = chart_data.get("數值") or {}
    channels = list(values.keys())
    if not channels:
        return
    s_start = chart_data.get("穩定開始", 0)
    s_end = chart_data.get("穩定結束", len(times) - 1)
    upper = setpoint + tol
    lower = setpoint - tol

    ws = wb.create_sheet("原始數據")
    ws.column_dimensions['A'].width = 14
    # 表頭：時間 + 通道 + 上限/下限
    header = ["時間"] + channels + ["上限", "下限"]
    for i, h in enumerate(header, start=1):
        _set(ws, f"{ws.cell(row=1, column=i).coordinate}", h, font=_BOLD, fill=_HDR_FILL)

    for j, t in enumerate(times):
        row = j + 2
        _set(ws, f"A{row}", t)
        for k, ch in enumerate(channels, start=2):
            arr = values.get(ch) or []
            v = arr[j] if j < len(arr) else None
            cell = _set(ws, ws.cell(row=row, column=k).coordinate, v)
            if (s_start <= j <= s_end) and (v is not None):
                if v > upper:
                    cell.fill = _RED_FILL    # 超上限
                elif v < lower:
                    cell.fill = _BLUE_FILL   # 低於下限
        _set(ws, ws.cell(row=row, column=len(channels) + 2).coordinate, upper)
        _set(ws, ws.cell(row=row, column=len(channels) + 3).coordinate, lower)

    # 折線圖
    chart = LineChart()
    chart.title = "TUS 溫度曲線"
    chart.style = 2
    chart.y_axis.title = "溫度 (°C)"
    chart.x_axis.title = "時間"
    chart.height = 10
    chart.width = 24
    last_col = len(channels) + 3
    data = Reference(ws, min_col=2, max_col=last_col, min_row=1, max_row=len(times) + 1)
    cats = Reference(ws, min_col=1, min_row=2, max_row=len(times) + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    anchor_col = chr(64 + last_col + 2)
    ws.add_chart(chart, f"{anchor_col}2")
