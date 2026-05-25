"""AIAG 8D 報表 — Excel 格式（openpyxl）"""
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def _thin_border():
    thin = Side(style='thin')
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _hdr_font():
    return Font(name='微軟正黑體', size=11, bold=True, color='FFFFFF')


def _hdr_fill():
    return PatternFill(start_color='2B579A', end_color='2B579A', fill_type='solid')


def _lbl_fill():
    return PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')


def _cell(ws, row, col, value, bold=False, center=False, fill=None, border=True, wrap=False, font_size=10):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name='微軟正黑體', size=font_size, bold=bold)
    if center:
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=wrap)
    else:
        c.alignment = Alignment(vertical='center', wrap_text=wrap)
    if fill:
        c.fill = fill
    if border:
        c.border = _thin_border()
    return c


def generate_8d_excel(capa_data: dict) -> BytesIO:
    """從 CAPAService._to_dict() 回傳的 dict 產生 AIAG 8D Excel 報表"""
    wb = Workbook()
    ws = wb.active
    ws.title = 'AIAG 8D 報表'

    # 欄寬設定
    col_widths = [6, 20, 50, 20, 15]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── 標題列 ────────────────────────────────────────────────
    ws.row_dimensions[1].height = 30
    ws.merge_cells('A1:E1')
    c = ws.cell(row=1, column=1, value='AIAG 8D 問題解決報告')
    c.font = Font(name='微軟正黑體', size=16, bold=True, color='FFFFFF')
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')

    # ── 基本資訊 ──────────────────────────────────────────────
    row = 2
    meta = [
        ('8D 單號',   capa_data.get('no', '')),
        ('來源類型',  'NCMR' if capa_data.get('source_type') == 'ncmr' else '客訴'),
        ('嚴重度',    capa_data.get('D0_severity') or ''),
        ('嚴格度',    capa_data.get('rigor', '')),
        ('狀態',      capa_data.get('status', '')),
        ('建立日期',  (capa_data.get('created_at') or '')[:10]),
        ('結案日期',  capa_data.get('D8_close_date') or ''),
    ]
    ws.merge_cells(f'A{row}:E{row}')
    c2 = ws.cell(row=row, column=1, value='基本資訊')
    c2.font = _hdr_font()
    c2.fill = _hdr_fill()
    c2.alignment = Alignment(horizontal='left', vertical='center')
    c2.border = _thin_border()
    row += 1
    for label, val in meta:
        ws.merge_cells(f'A{row}:B{row}')
        _cell(ws, row, 1, label, bold=True, fill=_lbl_fill())
        ws.merge_cells(f'C{row}:E{row}')
        _cell(ws, row, 3, val)
        row += 1

    # 來源資訊
    src = capa_data.get('source_info') or {}
    if src:
        ws.merge_cells(f'A{row}:E{row}')
        c3 = ws.cell(row=row, column=1, value='來源資訊')
        c3.font = _hdr_font()
        c3.fill = _hdr_fill()
        c3.alignment = Alignment(horizontal='left', vertical='center')
        c3.border = _thin_border()
        row += 1
        for k, v in src.items():
            ws.merge_cells(f'A{row}:B{row}')
            _cell(ws, row, 1, k, bold=True, fill=_lbl_fill())
            ws.merge_cells(f'C{row}:E{row}')
            _cell(ws, row, 3, v or '')
            row += 1

    def section_header(title: str):
        nonlocal row
        ws.row_dimensions[row].height = 22
        ws.merge_cells(f'A{row}:E{row}')
        ch = ws.cell(row=row, column=1, value=title)
        ch.font = _hdr_font()
        ch.fill = _hdr_fill()
        ch.alignment = Alignment(horizontal='left', vertical='center')
        ch.border = _thin_border()
        row += 1

    def text_row(label: str, value: str, height: int = 15):
        nonlocal row
        ws.row_dimensions[row].height = height
        ws.merge_cells(f'A{row}:B{row}')
        _cell(ws, row, 1, label, bold=True, fill=_lbl_fill())
        ws.merge_cells(f'C{row}:E{row}')
        _cell(ws, row, 3, value or '', wrap=True)
        row += 1

    # ── D0 ───────────────────────────────────────────────────
    section_header('D0 緊急應對（Emergency Response）')
    text_row('症狀描述', capa_data.get('D0_symptom') or '', height=40)
    criteria = capa_data.get('D0_criteria') or []
    text_row('判斷準則', '、'.join(criteria) if criteria else '')
    text_row('客戶要求結案日', capa_data.get('D0_deadline') or '')

    # ── D1 ───────────────────────────────────────────────────
    section_header('D1 成立團隊（Team Formation）')
    text_row('Champion', capa_data.get('D1_champion_name') or '')
    text_row('Team Leader', capa_data.get('D1_leader_name') or '')

    # ── D2 ───────────────────────────────────────────────────
    section_header('D2 問題描述 5W2H（Problem Description）')
    for label, key in [
        ('What（是什麼）',   'D2_what'),
        ('Where（在哪裡）',  'D2_where'),
        ('When（何時）',     'D2_when'),
        ('Who（誰）',        'D2_who'),
        ('Why（為何出現）',  'D2_why'),
        ('How（如何發現）',  'D2_how'),
        ('How Many（數量）', 'D2_how_many'),
    ]:
        text_row(label, capa_data.get(key) or '', height=30)

    # ── D3 ───────────────────────────────────────────────────
    section_header('D3 暫時對策（Containment Action）')
    text_row('暫時對策內容', capa_data.get('D3_action') or '', height=50)
    text_row('生效日期', capa_data.get('D3_effective_date') or '')
    text_row('有效性驗證', capa_data.get('D3_verification') or '', height=30)

    # ── D4 ───────────────────────────────────────────────────
    section_header('D4 根本原因分析（Root Cause Analysis）')
    text_row('分析工具', capa_data.get('D4_tool') or '')

    tool = capa_data.get('D4_tool', '5why')
    if tool == '5why':
        five_why = capa_data.get('D4_five_why') or []
        for i, item in enumerate(five_why, 1):
            if isinstance(item, dict):
                text_row(f'Why {i}', f"{item.get('why', '')} → {item.get('answer', '')}", height=25)
    else:
        fishbone = capa_data.get('D4_fishbone') or {}
        for m_key, items in fishbone.items():
            if items:
                text_row(m_key, '、'.join(str(x) for x in items if x))

    text_row('根本原因（彙整）', capa_data.get('D4_root_cause') or '', height=50)

    # ── D5 ───────────────────────────────────────────────────
    section_header('D5 永久對策（Permanent Corrective Action）')
    text_row('永久對策內容', capa_data.get('D5_action') or '', height=50)
    text_row('預計實施日', capa_data.get('D5_planned_date') or '')
    text_row('驗證計畫', capa_data.get('D5_verify_plan') or '', height=30)

    # ── D6 ───────────────────────────────────────────────────
    section_header('D6 實施與驗證（Implementation & Verification）')
    text_row('實施日期', capa_data.get('D6_implement_date') or '')
    text_row('驗證結果', capa_data.get('D6_result') or '', height=50)
    text_row('驗證通過', '是' if capa_data.get('D6_verified') else '否')

    # ── D7 ───────────────────────────────────────────────────
    section_header('D7 橫向展開（Prevent Recurrence）')
    d7_actions = capa_data.get('D7_actions') or []
    checked = [a for a in d7_actions if isinstance(a, dict) and a.get('checked')]
    if checked:
        for a in checked:
            text_row(a.get('type', ''), a.get('description') or '')
    else:
        text_row('橫展項目', '（無）')

    # ── D8 ───────────────────────────────────────────────────
    section_header('D8 結案確認（Closure）')
    text_row('結案確認聲明', capa_data.get('D8_confirmation') or '', height=60)
    text_row('團隊表揚與心得', capa_data.get('D8_recognition') or '', height=40)
    text_row('結案日期', capa_data.get('D8_close_date') or '')

    # ── 頁尾 ──────────────────────────────────────────────────
    row += 1
    ws.merge_cells(f'A{row}:E{row}')
    cf = ws.cell(row=row, column=1, value=f'產出日期：{datetime.now().strftime("%Y-%m-%d %H:%M")}')
    cf.font = Font(name='微軟正黑體', size=9, italic=True, color='808080')
    cf.alignment = Alignment(horizontal='right')

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
