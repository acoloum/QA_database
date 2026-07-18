
import pandas as pd
from io import BytesIO
from datetime import datetime, timezone
from typing import List, Dict, Any, Optional, Union
from sqlalchemy import func, text
from sqlalchemy.orm import selectinload
from ..extensions import db
from ..models import PatrolMain, PatrolDetail, Machine, Operator, Inspector, Vendor, SPCCache
from .extrusion_tolerance_service import ExtrusionToleranceService
from .patrol_excel_utils import (
    build_patrol_measurements_from_row,
    copy_spc_workbook_sheets,
    sanitize_sheet_name,
)
from ..utils import (
    bounded_int,
    format_value,
    validate_excel_shape,
    validate_patrol_data,
    handle_db_error,
    log_audit,
)

class PatrolService:
    @staticmethod
    def _invalidate_spc_cache() -> None:
        """巡檢來源異動後移除即時預覽快取；正式研究版本不受影響。"""
        SPCCache.query.filter(
            SPCCache.cache_key.like('spc2026|patrol|%')
        ).delete(synchronize_session=False)

    @staticmethod
    def _limits_key_filter(key: Dict[str, str]):
        from ..models import SpcControlLimit
        return SpcControlLimit.query.filter_by(
            source='patrol',
            vendor='',
            material=key.get('material') or '',
            spec=key.get('spec') or '',
            field=key['item'],
            position=key.get('position') or '',
        )

    @staticmethod
    def get_frozen_limits(key: Dict[str, str]) -> Optional[Dict[str, float]]:
        """查詢巡檢是否已凍結管制界限（§9.4）；若無則回傳 None。"""
        rec = PatrolService._limits_key_filter(key).first()
        if rec is None:
            return None
        return {
            "x_cl": float(rec.x_cl), "x_ucl": float(rec.x_ucl), "x_lcl": float(rec.x_lcl),
            "r_cl": float(rec.r_cl), "r_ucl": float(rec.r_ucl), "r_lcl": float(rec.r_lcl),
            "avg_n": rec.avg_n, "note": rec.note,
            "updated_at": rec.updated_at.isoformat() if rec.updated_at else None,
        }

    @staticmethod
    def get_options() -> Dict[str, List[Dict[str, Any]]]:
        """獲取下拉選單選項"""
        try:
            machines = [{"id": r.id, "name": r.name.strip()} for r in Machine.query.all()]
            operators = [{"id": r.id, "name": r.name.strip()} for r in Operator.query.all()]
            inspectors = [{"id": r.id, "name": r.name.strip()} for r in Inspector.query.filter_by(group='品保').order_by(Inspector.name).all()]
            customers = [{"id": r.id, "name": r.name.strip()} for r in Vendor.query.all()]
            
            return {
                "machines": machines,
                "operators": operators,
                "inspectors": inspectors,
                "customers": customers
            }
        except Exception as e:
            raise e

    @staticmethod
    def get_spc(args: Dict[str, Any], skip_frozen_limits: bool = False) -> Dict[str, Any]:
        """使用 2026 共用 SPC 引擎產生即時預覽；舊凍結界限不再覆蓋計算。"""
        from .spc_study_service import SpcStudyService

        return SpcStudyService.preview("patrol", args)

    @staticmethod
    def get_patrol_details(main_id: int) -> List[Dict[str, Any]]:
        """取得單筆巡檢記錄的全部量測明細（供離群值管理 UI）"""
        rows = PatrolDetail.query.filter_by(main_id=main_id).order_by(
            PatrolDetail.item, PatrolDetail.group, PatrolDetail.position
        ).all()
        return [{
            "識別碼": d.id, "組別": d.group, "測量項目": d.item, "測量位置": d.position,
            "最小值": float(d.min_val) if d.min_val is not None else None,
            "最大值": float(d.max_val) if d.max_val is not None else None,
            "排除統計": d.excluded, "排除原因": d.exclusion_reason,
            "排除者ID": d.exclusion_user_id,
            "排除時間": d.excluded_at.isoformat() if d.excluded_at else None,
        } for d in rows]

    @staticmethod
    def set_patrol_detail_exclusion(
        detail_id: int,
        excluded: bool,
        reason: Optional[str],
        actor_id: Optional[int] = None,
    ) -> Dict[str, Any]:
        """標示/解除巡檢量測明細離群排除（§6.6：不刪除、保留追溯、排除統計）"""
        d = db.session.get(PatrolDetail, detail_id)
        if d is None:
            raise ValueError("量測明細不存在")
        normalized_reason = (reason or "").strip()
        if not normalized_reason:
            action = "標示離群" if excluded else "恢復統計"
            raise ValueError(f"{action}必須填寫原因（§6.6）")
        old_value = {
            "excluded": bool(d.excluded),
            "reason": d.exclusion_reason,
            "actor_id": d.exclusion_user_id,
            "at": d.excluded_at.isoformat() if d.excluded_at else None,
        }
        d.excluded = excluded
        d.exclusion_reason = normalized_reason if excluded else None
        d.exclusion_user_id = actor_id if excluded else None
        d.excluded_at = datetime.now(timezone.utc) if excluded else None
        new_value = {
            "excluded": bool(d.excluded),
            "reason": d.exclusion_reason,
            "actor_id": d.exclusion_user_id,
            "at": d.excluded_at.isoformat() if d.excluded_at else None,
            "action_reason": normalized_reason,
        }
        log_audit(
            actor_id,
            "exclude" if excluded else "restore",
            "patrol_detail",
            d.id,
            old_val=old_value,
            new_val=new_value,
        )
        PatrolService._invalidate_spc_cache()
        db.session.commit()
        return {
            "id": d.id, "排除統計": d.excluded, "排除原因": d.exclusion_reason,
            "排除者ID": d.exclusion_user_id,
            "排除時間": d.excluded_at.isoformat() if d.excluded_at else None,
        }

    @staticmethod
    def get_detail(id: int) -> Optional[Dict[str, Any]]:
        """獲取巡檢詳細資料（主檔+子檔）"""
        patrol = db.session.get(PatrolMain, id)
        if not patrol:
            return None

        # Main data
        main = {
            "識別碼": patrol.id,
            "檢驗日期": format_value(patrol.date),
            "機台": patrol.machine_id,
            "主機手": patrol.operator_id,
            "檢驗人員": patrol.inspector_id,
            "材質": format_value(patrol.material),
            "擠壓規格": format_value(patrol.spec),
            "客戶名稱": patrol.customer_id,
            "原料批號": format_value(patrol.batch_num)
        }

        # Details
        details_list = []
        # Sort details by group, item, position if needed, though list usually follows creation order or DB default
        # Implementing explicit sort just in case
        sorted_details = sorted(patrol.details, key=lambda x: (x.group, x.item, x.position))
        
        for d in sorted_details:
            group_val = str(d.group).strip()
            if group_val.isdigit():
                group_val = f"第{group_val}組"
            
            details_list.append({
                "group": group_val,
                "item": d.item.strip() if d.item else "",
                "pos": d.position.strip() if d.position else "",
                "min": float(d.min_val) if d.min_val else None,
                "max": float(d.max_val) if d.max_val else None
            })

        return {
            "main": main,
            "details": details_list
        }

    @staticmethod
    def _parse_id(val: Any) -> Optional[int]:
        """Convert value to int or None if empty/invalid"""
        if not val and val != 0:
            return None
        if isinstance(val, str) and not val.strip():
            return None
        try:
            return int(val)
        except (ValueError, TypeError):
            return None

    @staticmethod
    def _compute_is_ng(details: list, material: str, spec: str, customer_id) -> bool:
        """
        依量測明細與擠壓公差計算是否超差。
        複用 get_history 內的判斷邏輯，於新增/更新時即時儲存結果。
        """
        if not material:
            return False
        result = ExtrusionToleranceService.check({
            'material': material,
            'spec': spec or '',
            'vendor_id': customer_id
        })
        if not result.get('found'):
            return False
        tol_map = {t['項目']: t for t in result.get('tolerances', [])}

        # 從規格解析標準值（如 "85*2.8" → {外徑:85, 厚度:2.8}）
        spec_nominal: dict = {}
        if spec:
            parts = (spec or '').replace('×', '*').replace('x', '*').replace('X', '*').split('*')
            try:
                if len(parts) >= 1:
                    spec_nominal['外徑'] = float(parts[0])
                if len(parts) >= 2:
                    spec_nominal['厚度'] = float(parts[1])
            except ValueError:
                pass

        for d in details:
            tol = tol_map.get(d.item)
            if tol:
                dim_min = tol.get('尺寸下限')
                dim_max = tol.get('尺寸上限')
                if dim_min is None and dim_max is None:
                    std_val = tol.get('標準值') or spec_nominal.get(d.item)
                    tol_min = tol.get('公差下限')
                    tol_max = tol.get('公差上限')
                    if std_val is not None:
                        if tol_min is not None:
                            dim_min = std_val - abs(tol_min)
                        if tol_max is not None:
                            dim_max = std_val + abs(tol_max)
                for val in [
                    float(d.min_val) if d.min_val is not None else None,
                    float(d.max_val) if d.max_val is not None else None,
                ]:
                    if val is None:
                        continue
                    if dim_min is not None and val < dim_min:
                        return True
                    if dim_max is not None and val > dim_max:
                        return True

            # 同心度判斷
            if d.item == '厚度' and d.min_val is not None and d.max_val is not None:
                conc_tol = tol_map.get('同心度')
                if conc_tol:
                    conc_dim_min = conc_tol.get('尺寸下限')
                    conc_dim_max = conc_tol.get('尺寸上限')
                    if conc_dim_min is None and conc_dim_max is None:
                        std_val = conc_tol.get('標準值')
                        tol_min = conc_tol.get('公差下限')
                        tol_max = conc_tol.get('公差上限')
                        if std_val is not None:
                            if tol_min is not None:
                                conc_dim_min = std_val - abs(tol_min)
                            if tol_max is not None:
                                conc_dim_max = std_val + abs(tol_max)
                        elif tol_min == 0.0 and tol_max is not None:
                            conc_dim_min = 0.0
                            conc_dim_max = float(tol_max)
                        elif tol_max is not None:
                            # 同心度：僅有公差上限時，直接作為絕對上限
                            conc_dim_max = float(tol_max)
                    concentricity = float(d.max_val) - float(d.min_val)
                    if conc_dim_min is not None and concentricity < conc_dim_min:
                        return True
                    if conc_dim_max is not None and concentricity > conc_dim_max:
                        return True
        return False

    @staticmethod
    def add_patrol(data: Dict[str, Any]) -> int:
        """新增巡檢資料"""
        errors = validate_patrol_data(data)
        if errors:
            raise ValueError(", ".join(errors))

        try:
            new_patrol = PatrolMain(
                date=data.get('檢驗日期'),
                machine_id=PatrolService._parse_id(data.get('機台')),
                operator_id=PatrolService._parse_id(data.get('主機手')),
                inspector_id=PatrolService._parse_id(data.get('檢驗人員')),
                material=data.get('材質'),
                spec=data.get('擠壓規格'),
                customer_id=PatrolService._parse_id(data.get('客戶名稱')),
                batch_num=data.get('原料批號')
            )
            db.session.add(new_patrol)
            db.session.flush() # Get ID

            for d in data.get('details', []):
                group_raw = str(d.get('group', '')).strip()
                group_val = group_raw.replace('第', '').replace('組', '')
                group_val = int(group_val) if group_val.isdigit() else 1

                min_val = float(d.get('min')) if d.get('min') is not None else None
                max_val = float(d.get('max')) if d.get('max') is not None else None

                new_detail = PatrolDetail(
                    main_id=new_patrol.id,
                    group=group_val,
                    item=d.get('item'),
                    position=d.get('pos'),
                    min_val=min_val,
                    max_val=max_val
                )
                db.session.add(new_detail)

            # 計算並儲存 is_ng
            db.session.flush()
            new_patrol.is_ng = PatrolService._compute_is_ng(
                new_patrol.details,
                new_patrol.material or '',
                new_patrol.spec or '',
                new_patrol.customer_id
            )

            PatrolService._invalidate_spc_cache()
            db.session.commit()
            return new_patrol.id
        except Exception as e:
            try:
                import traceback
                with open('debug_error.log', 'a') as f:
                    f.write(f"[{datetime.now()}] ERROR IN ADD_PATROL: {str(e)}\n")
                    f.write(traceback.format_exc())
                    f.write("\n" + "="*30 + "\n")
            except:
                pass
            db.session.rollback()
            raise e

    @staticmethod
    def update_patrol(data: Dict[str, Any]) -> bool:
        """更新巡檢資料"""
        record_id = data.get('id')
        if not record_id:
            raise ValueError("缺少記錄 ID")

        errors = validate_patrol_data(data)
        if errors:
            raise ValueError(", ".join(errors))

        try:
            patrol = db.session.get(PatrolMain, record_id)
            if not patrol:
                raise ValueError("找不到該筆資料")

            # Update main fields
            patrol.date = data.get('檢驗日期')
            patrol.machine_id = PatrolService._parse_id(data.get('機台'))
            patrol.operator_id = PatrolService._parse_id(data.get('主機手'))
            patrol.inspector_id = PatrolService._parse_id(data.get('檢驗人員'))
            patrol.material = data.get('材質')
            patrol.spec = data.get('擠壓規格')
            patrol.customer_id = PatrolService._parse_id(data.get('客戶名稱'))
            patrol.batch_num = data.get('原料批號')

            # Update details: Delete all and re-insert
            PatrolDetail.query.filter_by(main_id=record_id).delete()

            for d in data.get('details', []):
                group_raw = str(d.get('group', '')).strip()
                group_val = group_raw.replace('第', '').replace('組', '')
                group_val = int(group_val) if group_val.isdigit() else 1

                min_val = float(d.get('min')) if d.get('min') is not None else None
                max_val = float(d.get('max')) if d.get('max') is not None else None

                new_detail = PatrolDetail(
                    main_id=record_id,
                    group=group_val,
                    item=d.get('item'),
                    position=d.get('pos'),
                    min_val=min_val,
                    max_val=max_val
                )
                db.session.add(new_detail)

            # 重新計算並儲存 is_ng
            db.session.flush()
            patrol.is_ng = PatrolService._compute_is_ng(
                patrol.details,
                patrol.material or '',
                patrol.spec or '',
                patrol.customer_id
            )

            PatrolService._invalidate_spc_cache()
            db.session.commit()
            return True
        except Exception as e:
            db.session.rollback()
            raise e

    @staticmethod
    def delete_patrol(record_id: int) -> bool:
        """刪除巡檢資料"""
        try:
            # Use query to find so we can delete. 
            # Note: Cascade delete on relationship should handle details, 
            # but sometimes implicit delete is safer to be explicit or if cascade not working.
            # I defined cascade="all, delete-orphan" in PatrolMain, so just deleting main is enough.
            patrol = db.session.get(PatrolMain, record_id)
            if patrol:
                db.session.delete(patrol)
                PatrolService._invalidate_spc_cache()
                db.session.commit()
            return True
        except Exception as e:
            db.session.rollback()
            raise e

    @staticmethod
    def get_history(args: Dict[str, Any]) -> Dict[str, Any]:
        """獲取巡檢歷史列表（分頁）"""
        try:
            query = db.session.query(
                PatrolMain,
                Machine.name.label('m_name'),
                Operator.name.label('op_name'),
                Vendor.name.label('cust_name')
            )\
                .outerjoin(Machine, PatrolMain.machine_id == Machine.id)\
                .outerjoin(Operator, PatrolMain.operator_id == Operator.id)\
                .outerjoin(Vendor, PatrolMain.customer_id == Vendor.id)\
                .options(selectinload(PatrolMain.details))

            if args.get('s_date'): query = query.filter(PatrolMain.date >= args['s_date'])
            if args.get('e_date'): query = query.filter(PatrolMain.date <= args['e_date'])
            if args.get('m_id'):   query = query.filter(PatrolMain.machine_id == args['m_id'])
            if args.get('op_id'):  query = query.filter(PatrolMain.operator_id == args['op_id'])
            if args.get('cust_id'): query = query.filter(PatrolMain.customer_id == args['cust_id'])
            if args.get('mat'):    query = query.filter(PatrolMain.material.like(f"%{args['mat']}%"))
            if args.get('spec'):   query = query.filter(PatrolMain.spec.like(f"%{args['spec']}%"))

            query = query.order_by(PatrolMain.id.desc())

            page = bounded_int(args.get('page'), 1, 1, 1000000)
            per_page = bounded_int(args.get('per_page'), 20, 1, 100)

            pagination = query.paginate(page=page, per_page=per_page, error_out=False)

            # --- 批次查詢押出公差（以避免 N+1） ---
            # 延遲匯入以避免循環依賴（patrol_service ↔ extrusion_tolerance_service）
            from ..services.extrusion_tolerance_service import ExtrusionToleranceService

            unique_combos = {
                (patrol_item.material or '', patrol_item.spec or '', patrol_item.customer_id)
                for patrol_item, *_ in pagination.items
                if patrol_item.material
            }

            tol_cache: dict = {}
            for mat, sp, vid in unique_combos:
                result = ExtrusionToleranceService.check({'material': mat, 'spec': sp, 'vendor_id': vid})
                if result.get('found'):
                    tol_cache[(mat, sp, vid)] = {t['項目']: t for t in result.get('tolerances', [])}
                else:
                    tol_cache[(mat, sp, vid)] = None

            data = []
            for item in pagination.items:
                patrol, m_name, op_name, cust_name = item

                # --- NG 計算 ---
                mat = patrol.material or ''
                sp = patrol.spec or ''
                tol_map = tol_cache.get((mat, sp, patrol.customer_id)) if mat else None
                tol_found = tol_map is not None
                is_ng = False

                if tol_found:
                    # 從規格字串解析標準值（如 "85*2.8" → {'外徑': 85.0, '厚度': 2.8}）
                    # 供無標準值的相對公差記錄使用
                    spec_nominal: dict = {}
                    if sp:
                        parts = sp.replace('×', '*').replace('x', '*').replace('X', '*').split('*')
                        try:
                            if len(parts) >= 1:
                                spec_nominal['外徑'] = float(parts[0])
                            if len(parts) >= 2:
                                spec_nominal['厚度'] = float(parts[1])
                        except ValueError:
                            pass

                    for d in patrol.details:
                        tol = tol_map.get(d.item)
                        if tol:
                            # 優先使用尺寸下限/尺寸上限（絕對限制），其次使用公差下限/公差上限（相對限制）
                            dim_min = tol.get('尺寸下限')
                            dim_max = tol.get('尺寸上限')

                            # 如果沒有尺寸下限/上限，則從標準值和公差計算
                            # 標準值優先使用公差記錄本身，否則 fallback 到規格解析值
                            if dim_min is None and dim_max is None:
                                std_val = tol.get('標準值') or spec_nominal.get(d.item)
                                tol_min = tol.get('公差下限')
                                tol_max = tol.get('公差上限')
                                if std_val is not None:
                                    if tol_min is not None:
                                        dim_min = std_val - abs(tol_min)
                                    if tol_max is not None:
                                        dim_max = std_val + abs(tol_max)

                            # 進行 NG 判斷
                            for val in [
                                float(d.min_val) if d.min_val is not None else None,
                                float(d.max_val) if d.max_val is not None else None,
                            ]:
                                if val is None:
                                    continue
                                if dim_min is not None and val < dim_min:
                                    is_ng = True
                                if dim_max is not None and val > dim_max:
                                    is_ng = True

                        # 同心度：厚度行的 max_val - min_val 與同心度公差比對
                        if d.item == '厚度' and d.min_val is not None and d.max_val is not None:
                            conc_tol = tol_map.get('同心度')
                            if conc_tol:
                                # 優先使用尺寸下限/尺寸上限
                                conc_dim_min = conc_tol.get('尺寸下限')
                                conc_dim_max = conc_tol.get('尺寸上限')

                                if conc_dim_min is None and conc_dim_max is None:
                                    std_val = conc_tol.get('標準值')
                                    tol_min = conc_tol.get('公差下限')
                                    tol_max = conc_tol.get('公差上限')
                                    if std_val is not None:
                                        if tol_min is not None:
                                            conc_dim_min = std_val - abs(tol_min)
                                        if tol_max is not None:
                                            conc_dim_max = std_val + abs(tol_max)
                                    elif tol_min == 0.0 and tol_max is not None:
                                        # 同心度：tolerance_min=0 表示「最小值=0」，tolerance_max 為絕對上限
                                        conc_dim_min = 0.0
                                        conc_dim_max = float(tol_max)
                                    elif tol_max is not None:
                                        # 同心度：僅有公差上限時，直接作為絕對上限
                                        conc_dim_max = float(tol_max)
                                
                                concentricity = float(d.max_val) - float(d.min_val)
                                if conc_dim_min is not None and concentricity < conc_dim_min:
                                    is_ng = True
                                if conc_dim_max is not None and concentricity > conc_dim_max:
                                    is_ng = True

                        if is_ng:
                            break

                date_str = patrol.date.strftime('%Y-%m-%d') if patrol.date else ''
                data.append({
                    'id': patrol.id,
                    'date': date_str,
                    'm_name': m_name.strip() if m_name else '',
                    'op_name': op_name.strip() if op_name else '',
                    'cust_name': cust_name.strip() if cust_name else '',
                    'mat': patrol.material,
                    'spec': patrol.spec,
                    'is_ng': is_ng,
                    'tol_found': tol_found,
                })

            return {
                "data": data,
                "pages": pagination.pages,
                "total": pagination.total
            }
        except Exception as e:
            raise e

    @staticmethod
    def export_excel(args: Dict[str, Any]) -> BytesIO:
        """匯出巡檢資料 Excel（含 SPC 統計與圖表，按規格+材質分組）"""
        from ..services.spc_report import SpcReportService
        from openpyxl import load_workbook

        spc_item = args.get('item', '')      # 測量項目（外徑/內徑/厚度）
        spc_position = args.get('position', '')  # 測量位置（前段/中段/後段/全段）

        try:
            # === 1. 查詢原始資料 ===
            query = db.session.query(
                PatrolMain,
                Machine.name.label('m_name'),
                Operator.name.label('op_name'),
                Inspector.name.label('i_name'),
                Vendor.name.label('v_name')
            )\
            .outerjoin(Machine, PatrolMain.machine_id == Machine.id)\
            .outerjoin(Operator, PatrolMain.operator_id == Operator.id)\
            .outerjoin(Inspector, PatrolMain.inspector_id == Inspector.id)\
            .outerjoin(Vendor, PatrolMain.customer_id == Vendor.id)\
            .options(selectinload(PatrolMain.details))

            if args.get('s_date'): query = query.filter(PatrolMain.date >= args['s_date'])
            if args.get('e_date'): query = query.filter(PatrolMain.date <= args['e_date'])
            if args.get('m_id'):   query = query.filter(PatrolMain.machine_id == args['m_id'])
            if args.get('op_id'):  query = query.filter(PatrolMain.operator_id == args['op_id'])
            if args.get('cust_id'): query = query.filter(PatrolMain.customer_id == args['cust_id'])
            if args.get('mat'):    query = query.filter(PatrolMain.material.like(f"%{args['mat']}%"))
            if args.get('spec'):   query = query.filter(PatrolMain.spec.like(f"%{args['spec']}%"))

            query = query.order_by(PatrolMain.id.desc())
            rows = query.all()

            if not rows:
                df = pd.DataFrame(columns=['識別碼', '檢驗日期', '擠壓機編號', '員工姓名', '材質', '擠壓規格', '廠商名稱', '原料批號', '檢驗人員'])
            else:
                export_data = []
                unique_groups = []

                for row in rows:
                    patrol, m_name, op_name, i_name, v_name = row
                    details = sorted(patrol.details, key=lambda x: (x.group, x.item, x.position))

                    measurements = {}
                    current_groups = []

                    for d in details:
                        group_val = str(d.group)
                        group_name = group_val if "組" in group_val else f"第{group_val}組"
                        item = d.item.strip() if d.item else ""
                        pos = d.position.strip() if d.position else ""
                        min_val = float(d.min_val) if d.min_val else ""
                        max_val = float(d.max_val) if d.max_val else ""
                        key = f"{group_name}_{item}_{pos}"
                        measurements[key] = {"min": min_val, "max": max_val}
                        if group_name not in current_groups:
                            current_groups.append(group_name)

                    for g in current_groups:
                        if g not in unique_groups:
                            unique_groups.append(g)

                    if not current_groups:
                        current_groups = ["第1組"]
                        if "第1組" not in unique_groups:
                            unique_groups.append("第1組")

                    row_dict = {
                        '識別碼': patrol.id,
                        '檢驗日期': patrol.date.strftime('%Y-%m-%d') if patrol.date else '',
                        '擠壓機編號': m_name.strip() if m_name else '',
                        '員工姓名': op_name.strip() if op_name else '',
                        '材質': patrol.material,
                        '擠壓規格': patrol.spec,
                        '廠商名稱': v_name.strip() if v_name else '',
                        '原料批號': patrol.batch_num,
                        '檢驗人員': i_name.strip() if i_name else ''
                    }

                    for group in current_groups:
                        for item in ["外徑", "內徑", "厚度"]:
                            for pos in ["前段", "中段", "後段"]:
                                key = f"{group}_{item}_{pos}"
                                min_val = measurements.get(key, {}).get("min", "")
                                max_val = measurements.get(key, {}).get("max", "")
                                row_dict[f"{group}{item}{pos}最小"] = min_val
                                row_dict[f"{group}{item}{pos}最大"] = max_val

                    export_data.append(row_dict)

                df = pd.DataFrame(export_data)

            # 先產生原始數據工作表
            output = BytesIO()
            df.to_excel(output, index=False, engine='openpyxl', sheet_name='原始數據')
            output.seek(0)

            # === 2. 如果有選擇 SPC 項目，按規格+材質分組產生 SPC 報表 ===
            if spc_item and rows:
                wb = load_workbook(output)

                # 收集所有不重複的「規格+材質」組合
                spec_mat_set: Dict[str, Dict[str, Any]] = {}
                for row in rows:
                    patrol = row[0]
                    v_name = row[4]  # Vendor.name
                    mat = (patrol.material or '').strip()
                    spec_val = (patrol.spec or '').strip()
                    cust = (v_name.strip() if v_name else '')
                    if mat or spec_val:
                        group_key = f"{spec_val}_{mat}"
                        if group_key not in spec_mat_set:
                            spec_mat_set[group_key] = {'spec': spec_val, 'material': mat, 'customers': set()}
                        if cust:
                            spec_mat_set[group_key]['customers'].add(cust)

                for group_key, group_info in spec_mat_set.items():
                    # 為每個組合呼叫 get_spc 取得統計數據
                    spc_args = {
                        'item': spc_item,
                        'pos': spc_position if spc_position != '全段' else '',
                        'mat': group_info['material'],
                        'spec': group_info['spec'],
                        's_date': args.get('s_date', ''),
                        'e_date': args.get('e_date', ''),
                        'm_id': args.get('m_id', ''),
                        'op_id': args.get('op_id', ''),
                    }
                    stats_data = PatrolService.get_spc(spc_args)

                    # 跳過無數據的組合
                    if not stats_data.get('avgs'):
                        continue

                    # 產生 SPC 報表（獨立 Workbook）
                    pos_label = spc_position if spc_position else '全段'
                    field_label = f"{spc_item}-{pos_label}"
                    # 從該分組的實際資料中取得客戶名稱
                    group_customers = group_info.get('customers', set())
                    cust_display = '、'.join(sorted(group_customers)) if group_customers else '無'
                    filters = {
                        'material': group_info['material'] or '全部',
                        'spec': group_info['spec'] or '全部',
                        'customer': cust_display,
                        'start_date': args.get('s_date', '不限'),
                        'end_date': args.get('e_date', '不限'),
                    }
                    spc_output = SpcReportService.generate_report(stats_data, field_label, filters)

                    spc_wb = load_workbook(spc_output)
                    # 建立工作表名稱前綴（截斷以避免超過 31 字元限制）
                    spec_short = sanitize_sheet_name(group_info['spec'][:8]) if group_info['spec'] else '無規格'
                    mat_short = sanitize_sheet_name(group_info['material'][:6]) if group_info['material'] else '無材質'
                    prefix = f"{spec_short}-{mat_short}"
                    copy_spc_workbook_sheets(wb, spc_wb, prefix)

                # 輸出最終結果
                output = BytesIO()
                wb.save(output)
                output.seek(0)

            return output
        except Exception as e:
            raise e

    @staticmethod
    def import_data(file: Any) -> int:
        """匯入巡檢資料 Excel"""
        try:
            df = pd.read_excel(file, engine='openpyxl')
        except Exception as e:
            raise ValueError(f"檔案讀取失敗: {str(e)}")
        validate_excel_shape(df)

        success_count = 0
        try:
            # We can use ORM for lookups but bulk insert might be faster if purely adding.
            # However, for consistency and relationship management, let's use ORM session.
            # Speed is usually acceptable for Excel imports of moderate size.
            
            for row_num, row in enumerate(df.iterrows()):
                main_data = row[1].to_dict()
                display_row_num = row_num + 2

                # Lookups
                machine_name = str(main_data.get('擠壓機編號', '')).strip()
                machine = Machine.query.filter_by(name=machine_name).first() if machine_name else None
                if not machine and machine_name:
                    raise ValueError(f"第 {display_row_num} 行: 找不到機台 '{machine_name}'")
                
                operator_name = str(main_data.get('員工姓名', '')).strip()
                operator = Operator.query.filter_by(name=operator_name).first() if operator_name else None
                if not operator and operator_name:
                    raise ValueError(f"第 {display_row_num} 行: 找不到員工 '{operator_name}'")

                customer_name = str(main_data.get('客戶名稱', '')).strip()
                customer = Vendor.query.filter_by(name=customer_name).first() if customer_name else None
                
                inspector_name = str(main_data.get('檢驗人員', '')).strip()
                inspector = Inspector.query.filter_by(name=inspector_name).first() if inspector_name else None
                if not inspector and inspector_name:
                    raise ValueError(f"第 {display_row_num} 行: 找不到檢驗人員 '{inspector_name}'")

                # Checks required fields
                if not machine: raise ValueError(f"第 {display_row_num} 行: 機台必填")
                if not operator: raise ValueError(f"第 {display_row_num} 行: 員工必填")
                if not inspector: raise ValueError(f"第 {display_row_num} 行: 檢驗人員必填")

                new_patrol = PatrolMain(
                    date=main_data.get('檢驗日期'),
                    machine_id=machine.id,
                    operator_id=operator.id,
                    material=main_data.get('材質'),
                    spec=main_data.get('擠壓規格'),
                    customer_id=customer.id if customer else None,
                    batch_num=main_data.get('原料批號'),
                    inspector_id=inspector.id
                )
                db.session.add(new_patrol)
                db.session.flush()

                # Process details from columns
                for measurement in build_patrol_measurements_from_row(main_data):
                    new_detail = PatrolDetail(
                        main_id=new_patrol.id,
                        group=1, # Default to group 1 for flattened Excel structure usually
                        item=measurement["item"],
                        position=measurement["position"],
                        min_val=measurement["min_val"],
                        max_val=measurement["max_val"],
                    )
                    db.session.add(new_detail)

                success_count += 1
            
            PatrolService._invalidate_spc_cache()
            db.session.commit()
            return success_count
        except Exception as e:
            db.session.rollback()
            raise e
