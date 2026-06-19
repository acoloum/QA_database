"""CQI-9 爐溫測試服務 — 設備主檔、測試紀錄、判定、到期、趨勢"""
from typing import Dict, Any, List
from datetime import date, timedelta
from dateutil.relativedelta import relativedelta
from ..extensions import db
from ..models import Furnace, Recorder, RecorderCalPoint, Thermocouple, ThermocoupleCalPoint
from ..utils import format_value


def _quarter_of(d) -> str:
    return f"{d.year}Q{(d.month - 1) // 3 + 1}"


def _to_float(v) -> float | None:
    """安全轉換為 float；None 或空字串回傳 None。"""
    if v is None or str(v).strip() == '':
        return None
    return float(v)


def _parse_date(v):
    if not v:
        return None
    if isinstance(v, date):
        return v
    from datetime import date as _date
    return _date.fromisoformat(str(v)[:10])


def _furnace_to_dict(f: Furnace) -> Dict[str, Any]:
    return {
        "識別碼": f.id, "爐號": f.code, "名稱": f.name,
        "製程類型": f.process_type or "",
        "TUS點數": f.tus_points, "SAT點數": f.sat_points,
        "TUS頻率_月": f.tus_freq_months, "SAT頻率_月": f.sat_freq_months,
        "TUS允許公差": format_value(f.tus_tolerance),
        "SAT允許誤差": format_value(f.sat_tolerance),
        "有效加熱區尺寸": f.work_zone or "",
        "儀器型式": f.instrument_type or "", "CQI9等級": f.cqi9_class or "",
        "啟用狀態": f.is_active, "備註": f.note or "",
    }


def _recorder_to_dict(r: Recorder, with_points: bool = False) -> Dict[str, Any]:
    d = {
        "識別碼": r.id, "編號": r.serial,
        "校正日期": format_value(r.cal_date), "到期日": format_value(r.cal_due_date),
        "熱電偶補正值": format_value(r.tc_correction),
        "啟用狀態": r.is_active, "備註": r.note or "",
    }
    if with_points:
        d["校正點"] = [
            {"頻道": cp.channel, "標準溫度": format_value(cp.std_temp), "器差值": format_value(cp.error)}
            for cp in sorted(r.cal_points, key=lambda x: (x.channel, float(x.std_temp)))
        ]
    return d


def _thermocouple_to_dict(t: Thermocouple, with_points: bool = False) -> Dict[str, Any]:
    d = {
        "識別碼": t.id, "編號": t.serial, "型式": t.tc_type or "",
        "校正日期": format_value(t.cal_date), "到期日": format_value(t.cal_due_date),
        "啟用狀態": t.is_active, "備註": t.note or "",
    }
    if with_points:
        d["校正點"] = [
            {"標準溫度": format_value(cp.std_temp), "器差值": format_value(cp.error)}
            for cp in sorted(t.cal_points, key=lambda x: float(x.std_temp))
        ]
    return d


def _interp_error(points: List, setpoint: float) -> float:
    """以(標準溫度, 器差值)點集對 setpoint 線性內插；超出範圍夾擠取端點。"""
    if not points:
        return 0.0
    pts = sorted(points, key=lambda p: p[0])
    if setpoint <= pts[0][0]:
        return pts[0][1]
    if setpoint >= pts[-1][0]:
        return pts[-1][1]
    for (t0, e0), (t1, e1) in zip(pts, pts[1:]):
        if t0 <= setpoint <= t1:
            if t1 == t0:
                return e0
            return e0 + (setpoint - t0) / (t1 - t0) * (e1 - e0)
    return pts[-1][1]


class PyrometryService:

    # ---------- 設備主檔 ----------
    @staticmethod
    def list_furnaces(active_only: bool = False) -> List[Dict[str, Any]]:
        q = Furnace.query
        if active_only:
            q = q.filter(Furnace.is_active.is_(True))
        return [_furnace_to_dict(f) for f in q.order_by(Furnace.code).all()]

    @staticmethod
    def get_furnace(furnace_id: int) -> Dict[str, Any]:
        f = db.session.get(Furnace, furnace_id)
        if not f:
            raise ValueError("找不到該爐子設備")
        return _furnace_to_dict(f)

    @staticmethod
    def add_furnace(data: Dict[str, Any]) -> int:
        try:
            f = Furnace(
                code=data.get("爐號"), name=data.get("名稱"),
                process_type=data.get("製程類型") or None,
                tus_points=int(data.get("TUS點數", 12) or 12),
                sat_points=int(data.get("SAT點數", 2) or 2),
                tus_freq_months=int(data.get("TUS頻率_月", 3) or 3),
                sat_freq_months=int(data.get("SAT頻率_月", 3) or 3),
                tus_tolerance=data.get("TUS允許公差") or None,
                sat_tolerance=data.get("SAT允許誤差") or None,
                work_zone=data.get("有效加熱區尺寸") or None,
                instrument_type=data.get("儀器型式") or None,
                cqi9_class=data.get("CQI9等級") or None,
                is_active=bool(data.get("啟用狀態", True)),
                note=data.get("備註") or None,
            )
            db.session.add(f)
            db.session.commit()
            return f.id
        except Exception as e:
            db.session.rollback()
            raise e

    @staticmethod
    def update_furnace(furnace_id: int, data: Dict[str, Any]) -> bool:
        try:
            f = db.session.get(Furnace, furnace_id)
            if not f:
                raise ValueError("找不到該爐子設備")
            f.code = data.get("爐號")
            f.name = data.get("名稱")
            f.process_type = data.get("製程類型") or None
            f.tus_points = int(data.get("TUS點數", 12) or 12)
            f.sat_points = int(data.get("SAT點數", 2) or 2)
            f.tus_freq_months = int(data.get("TUS頻率_月", 3) or 3)
            f.sat_freq_months = int(data.get("SAT頻率_月", 3) or 3)
            f.tus_tolerance = data.get("TUS允許公差") or None
            f.sat_tolerance = data.get("SAT允許誤差") or None
            f.work_zone = data.get("有效加熱區尺寸") or None
            f.instrument_type = data.get("儀器型式") or None
            f.cqi9_class = data.get("CQI9等級") or None
            f.is_active = bool(data.get("啟用狀態", True))
            f.note = data.get("備註") or None
            db.session.commit()
            return True
        except Exception as e:
            db.session.rollback()
            raise e

    @staticmethod
    def delete_furnace(furnace_id: int) -> bool:
        try:
            f = db.session.get(Furnace, furnace_id)
            if not f:
                raise ValueError("找不到該爐子設備")
            db.session.delete(f)
            db.session.commit()
            return True
        except ValueError:
            raise
        except Exception as e:
            db.session.rollback()
            raise e

    # ---------- 溫度記錄器校正 ----------
    @staticmethod
    def list_recorders(active_only: bool = False) -> List[Dict[str, Any]]:
        q = Recorder.query
        if active_only:
            q = q.filter(Recorder.is_active.is_(True))
        return [_recorder_to_dict(r) for r in q.order_by(Recorder.serial).all()]

    @staticmethod
    def get_recorder(recorder_id: int) -> Dict[str, Any]:
        r = db.session.get(Recorder, recorder_id)
        if not r:
            raise ValueError("找不到該記錄器")
        return _recorder_to_dict(r, with_points=True)

    @staticmethod
    def add_recorder(data: Dict[str, Any]) -> int:
        try:
            r = Recorder(
                serial=data.get("編號"),
                cal_date=_parse_date(data.get("校正日期")),
                cal_due_date=_parse_date(data.get("到期日")),
                tc_correction=data.get("熱電偶補正值") or 0,
                is_active=bool(data.get("啟用狀態", True)),
                note=data.get("備註") or None,
            )
            db.session.add(r)
            db.session.flush()
            for cp in data.get("校正點", []):
                db.session.add(RecorderCalPoint(
                    recorder_id=r.id, channel=int(cp["頻道"]),
                    std_temp=cp["標準溫度"], error=cp["器差值"],
                ))
            db.session.commit()
            return r.id
        except Exception as e:
            db.session.rollback()
            raise e

    @staticmethod
    def update_recorder(recorder_id: int, data: Dict[str, Any]) -> bool:
        try:
            r = db.session.get(Recorder, recorder_id)
            if not r:
                raise ValueError("找不到該記錄器")
            r.serial = data.get("編號")
            r.cal_date = _parse_date(data.get("校正日期"))
            r.cal_due_date = _parse_date(data.get("到期日"))
            r.tc_correction = data.get("熱電偶補正值") or 0
            r.is_active = bool(data.get("啟用狀態", True))
            r.note = data.get("備註") or None
            if "校正點" in data:
                RecorderCalPoint.query.filter_by(recorder_id=recorder_id).delete()
                for cp in data.get("校正點", []):
                    db.session.add(RecorderCalPoint(
                        recorder_id=recorder_id, channel=int(cp["頻道"]),
                        std_temp=cp["標準溫度"], error=cp["器差值"],
                    ))
            db.session.commit()
            return True
        except ValueError:
            raise
        except Exception as e:
            db.session.rollback()
            raise e

    @staticmethod
    def delete_recorder(recorder_id: int) -> bool:
        try:
            r = db.session.get(Recorder, recorder_id)
            if not r:
                raise ValueError("找不到該記錄器")
            db.session.delete(r)
            db.session.commit()
            return True
        except ValueError:
            raise
        except Exception as e:
            db.session.rollback()
            raise e

    # ---------- 熱電偶校正 ----------
    @staticmethod
    def list_thermocouples(active_only: bool = False) -> List[Dict[str, Any]]:
        q = Thermocouple.query
        if active_only:
            q = q.filter(Thermocouple.is_active.is_(True))
        return [_thermocouple_to_dict(t) for t in q.order_by(Thermocouple.serial).all()]

    @staticmethod
    def get_thermocouple(tc_id: int) -> Dict[str, Any]:
        t = db.session.get(Thermocouple, tc_id)
        if not t:
            raise ValueError("找不到該熱電偶")
        return _thermocouple_to_dict(t, with_points=True)

    @staticmethod
    def add_thermocouple(data: Dict[str, Any]) -> int:
        try:
            t = Thermocouple(
                serial=data.get("編號"), tc_type=data.get("型式") or None,
                cal_date=_parse_date(data.get("校正日期")),
                cal_due_date=_parse_date(data.get("到期日")),
                is_active=bool(data.get("啟用狀態", True)),
                note=data.get("備註") or None,
            )
            db.session.add(t)
            db.session.flush()
            for cp in data.get("校正點", []):
                db.session.add(ThermocoupleCalPoint(
                    thermocouple_id=t.id, std_temp=cp["標準溫度"], error=cp["器差值"],
                ))
            db.session.commit()
            return t.id
        except Exception as e:
            db.session.rollback()
            raise e

    @staticmethod
    def update_thermocouple(tc_id: int, data: Dict[str, Any]) -> bool:
        try:
            t = db.session.get(Thermocouple, tc_id)
            if not t:
                raise ValueError("找不到該熱電偶")
            t.serial = data.get("編號")
            t.tc_type = data.get("型式") or None
            t.cal_date = _parse_date(data.get("校正日期"))
            t.cal_due_date = _parse_date(data.get("到期日"))
            t.is_active = bool(data.get("啟用狀態", True))
            t.note = data.get("備註") or None
            if "校正點" in data:
                ThermocoupleCalPoint.query.filter_by(thermocouple_id=tc_id).delete()
                for cp in data.get("校正點", []):
                    db.session.add(ThermocoupleCalPoint(
                        thermocouple_id=tc_id, std_temp=cp["標準溫度"], error=cp["器差值"],
                    ))
            db.session.commit()
            return True
        except ValueError:
            raise
        except Exception as e:
            db.session.rollback()
            raise e

    @staticmethod
    def delete_thermocouple(tc_id: int) -> bool:
        try:
            t = db.session.get(Thermocouple, tc_id)
            if not t:
                raise ValueError("找不到該熱電偶")
            db.session.delete(t)
            db.session.commit()
            return True
        except ValueError:
            raise
        except Exception as e:
            db.session.rollback()
            raise e

    @staticmethod
    def thermocouple_correction(setpoint: float) -> float:
        """啟用中熱電偶於設定溫度的補正值（−內插器差值）；無則回傳 0。"""
        tc = Thermocouple.query.filter(Thermocouple.is_active.is_(True)).order_by(Thermocouple.id.desc()).first()
        if not tc:
            return 0.0
        pts = [(float(cp.std_temp), float(cp.error)) for cp in tc.cal_points]
        return round(-_interp_error(pts, float(setpoint)), 2)

    @staticmethod
    def compute_corrections(setpoint: float, test_type: str, count: int,
                            recorder_id: int = None,
                            channels: List[int] = None) -> List[float]:
        """依設定溫度算各量測點的修正值。
        修正值 = 熱電偶補正(−內插器差值) + 記錄器補正(−內插器差值)。
        頻道預設對應：TUS-i→CH i；SAT-i→CH (12+i)。
        channels 參數可明確指定每個量測點使用的頻道編號，取代自動推算。
        熱電偶：取啟用中熱電偶的校正曲線內插；若無則回退記錄器舊固定值。
        無啟用記錄器時回傳全 0（向後相容）。
        """
        if recorder_id is not None:
            r = db.session.get(Recorder, recorder_id)
        else:
            r = Recorder.query.filter(Recorder.is_active.is_(True)).order_by(Recorder.id.desc()).first()
        if not r:
            return [0.0] * count

        # 熱電偶補正：優先用啟用中熱電偶的校正曲線內插
        tc_obj = Thermocouple.query.filter(Thermocouple.is_active.is_(True)).order_by(Thermocouple.id.desc()).first()
        if tc_obj:
            tc_points = [(float(cp.std_temp), float(cp.error)) for cp in tc_obj.cal_points]
            tc_corr = -_interp_error(tc_points, float(setpoint))
        else:
            tc_corr = float(r.tc_correction or 0)   # 向後相容：舊固定值

        offset = 0 if test_type == "TUS" else 12
        # 預先依頻道整理記錄器校正點
        by_channel: Dict[int, List] = {}
        for cp in r.cal_points:
            by_channel.setdefault(cp.channel, []).append((float(cp.std_temp), float(cp.error)))
        out = []
        for i in range(count):
            ch = channels[i] if (channels and i < len(channels)) else (offset + i + 1)
            err = _interp_error(by_channel.get(ch, []), float(setpoint))
            out.append(round(tc_corr - err, 2))
        return out

    # ---------- 判定邏輯 ----------
    @staticmethod
    def evaluate_tus(setpoint: float, tolerance: float, points: List[Dict[str, Any]]) -> Dict[str, Any]:
        sp = float(setpoint)
        tol = float(tolerance)
        out_points = []
        all_max, all_min = [], []
        overall_pass = True
        for p in points:
            tmax = _to_float(p.get("最高溫"))
            tmin = _to_float(p.get("最低溫"))
            corr = _to_float(p.get("修正值")) or 0.0   # 熱電偶+記錄器補正
            # 校正後溫度 = 量測值 + 修正值
            tmax = tmax + corr if tmax is not None else None
            tmin = tmin + corr if tmin is not None else None
            dev_candidates = []
            if tmax is not None:
                dev_candidates.append(tmax - sp); all_max.append(tmax)
            if tmin is not None:
                dev_candidates.append(tmin - sp); all_min.append(tmin)
            # 最大偏差 = 絕對值最大的偏差（保留正負號）
            max_dev = max(dev_candidates, key=abs) if dev_candidates else None
            pt_pass = max_dev is None or abs(max_dev) <= tol
            overall_pass = overall_pass and pt_pass
            np_point = dict(p)
            np_point["最大偏差"] = round(max_dev, 2) if max_dev is not None else None
            np_point["是否合格"] = pt_pass
            out_points.append(np_point)
        tus_range = round(max(all_max) - min(all_min), 2) if all_max and all_min else None
        max_pos = round(max(all_max) - sp, 2) if all_max else None
        max_neg = round(min(all_min) - sp, 2) if all_min else None
        return {
            "是否合格": overall_pass, "TUS均勻度極差": tus_range,
            "TUS最大正偏差": max_pos, "TUS最大負偏差": max_neg, "points": out_points,
        }

    @staticmethod
    def evaluate_sat(tolerance: float, points: List[Dict[str, Any]]) -> Dict[str, Any]:
        """SAT 判定：每控溫區含多筆取樣讀值，所有讀值均需在公差內。
        兼容舊格式（單筆 控制儀表讀值/校正測試讀值）與新格式（readings 陣列）。
        """
        tol = float(tolerance)
        out_points = []
        overall_pass = True
        for p in points:
            corr = _to_float(p.get("修正值")) or 0.0
            # 取得讀值列表（新格式 readings 或舊格式單筆）
            raw_readings: List[Dict[str, Any]] = p.get("readings") or []
            if not raw_readings:
                ctrl_s = p.get("控制儀表讀值")
                test_s = p.get("校正測試讀值")
                if ctrl_s is not None or test_s is not None:
                    raw_readings = [{"控制儀表讀值": ctrl_s, "校正測試讀值": test_s}]
            computed_readings = []
            worst_diff = None
            worst_dev = None
            zone_pass = True
            for r in raw_readings:
                ctrl = _to_float(r.get("控制儀表讀值"))
                test = _to_float(r.get("校正測試讀值"))
                diff      = round(test - ctrl, 2) if (ctrl is not None and test is not None) else None
                deviation = round(diff + corr, 2) if diff is not None else None
                r_pass = deviation is None or abs(deviation) <= tol
                zone_pass = zone_pass and r_pass
                if deviation is not None and (worst_dev is None or abs(deviation) > abs(worst_dev)):
                    worst_dev = deviation
                    worst_diff = diff
                cr = {k: v for k, v in r.items() if k not in ("差值", "偏差", "是否合格")}
                cr["差值"] = diff
                cr["偏差"] = deviation
                cr["是否合格"] = r_pass
                computed_readings.append(cr)
            overall_pass = overall_pass and zone_pass
            np_point = {k: v for k, v in p.items() if k not in ("readings", "差值", "偏差", "是否合格")}
            np_point["readings"] = computed_readings
            np_point["差值"] = worst_diff
            np_point["偏差"] = worst_dev
            np_point["是否合格"] = zone_pass
            out_points.append(np_point)
        return {"是否合格": overall_pass, "points": out_points}

    # ---------- 測試紀錄 ----------
    @staticmethod
    def create_test(data: Dict[str, Any]) -> int:
        from ..models import PyrometryTest, TusPoint, SatPoint
        try:
            test_date = _parse_date(data.get("測試日期"))
            test_type = data.get("測試類型")
            tolerance = float(data.get("允許公差") or 0)
            setpoint = float(data.get("設定溫度") or 0)
            raw_points = data.get("points", [])

            if test_type == "TUS":
                judged = PyrometryService.evaluate_tus(setpoint, tolerance, raw_points)
            else:
                judged = PyrometryService.evaluate_sat(tolerance, raw_points)

            t = PyrometryTest(
                furnace_id=data.get("爐子ID"), test_type=test_type,
                quarter=data.get("季別") or _quarter_of(test_date),
                test_date=test_date, setpoint=setpoint, tolerance=tolerance,
                tester_id=data.get("測試人員") or None,
                test_instrument=data.get("測試儀器編號") or None,
                std_instrument=data.get("標準校正儀器編號") or None,
                cal_due_date=_parse_date(data.get("儀器校正到期日")),
                is_pass=judged["是否合格"],
                tus_range=judged.get("TUS均勻度極差"),
                tus_max_pos=judged.get("TUS最大正偏差"),
                tus_max_neg=judged.get("TUS最大負偏差"),
                chart_data=data.get("曲線資料") or None,
                report_meta=data.get("報告欄位") or None,
                note=data.get("備註") or None,
                created_by=data.get("建立人") or None,
            )
            db.session.add(t)
            db.session.flush()

            for p in judged["points"]:
                if test_type == "TUS":
                    tch = p.get("頻道")
                    db.session.add(TusPoint(
                        test_id=t.id, position=p.get("點位"), tc_no=p.get("熱電偶編號"),
                        channel=int(tch) if tch is not None else None,
                        correction=_to_float(p.get("修正值")),
                        temp_max=_to_float(p.get("最高溫")),
                        temp_min=_to_float(p.get("最低溫")),
                        max_dev=_to_float(p.get("最大偏差")),
                        is_pass=p.get("是否合格", True),
                    ))
                else:
                    ch = p.get("頻道")
                    db.session.add(SatPoint(
                        test_id=t.id, zone=p.get("控溫區"),
                        channel=int(ch) if ch is not None else None,
                        readings=p.get("readings") or None,
                        diff=_to_float(p.get("差值")),
                        correction=_to_float(p.get("修正值")),
                        deviation=_to_float(p.get("偏差")),
                        is_pass=p.get("是否合格", True),
                    ))
            db.session.commit()
            return t.id
        except Exception as e:
            db.session.rollback()
            raise e

    @staticmethod
    def get_test(test_id: int) -> Dict[str, Any]:
        from ..models import PyrometryTest
        t = PyrometryTest.query.filter(
            PyrometryTest.id == test_id, PyrometryTest.deleted_at.is_(None)
        ).first()
        if not t:
            raise ValueError("找不到該筆爐溫測試")
        main = {
            "識別碼": t.id, "爐子ID": t.furnace_id,
            "爐號": t.furnace.code if t.furnace else "",
            "測試類型": t.test_type, "季別": t.quarter or "",
            "測試日期": format_value(t.test_date), "設定溫度": format_value(t.setpoint),
            "允許公差": format_value(t.tolerance),
            "測試人員": t.tester_id, "測試人員姓名": t.tester.name if t.tester else "",
            "測試儀器編號": t.test_instrument or "", "標準校正儀器編號": t.std_instrument or "",
            "儀器校正到期日": format_value(t.cal_due_date),
            "是否合格": t.is_pass, "TUS均勻度極差": format_value(t.tus_range),
            "TUS最大正偏差": format_value(t.tus_max_pos), "TUS最大負偏差": format_value(t.tus_max_neg),
            "備註": t.note or "",
        }
        tus_points = [{
            "識別碼": p.id, "點位": p.position or "", "熱電偶編號": p.tc_no or "",
            "頻道": p.channel,
            "修正值": format_value(p.correction), "最高溫": format_value(p.temp_max),
            "最低溫": format_value(p.temp_min), "最大偏差": format_value(p.max_dev),
            "是否合格": p.is_pass,
        } for p in sorted(t.tus_points, key=lambda x: x.id)]
        def _sat_readings(p):
            """取得讀值列表；舊格式（單欄）轉為一筆 reading 相容。"""
            if p.readings:
                return p.readings
            if p.control_read is not None or p.test_read is not None:
                return [{"控制儀表讀值": format_value(p.control_read),
                         "校正測試讀值": format_value(p.test_read)}]
            return []

        sat_points = [{
            "識別碼": p.id, "控溫區": p.zone or "",
            "頻道": p.channel,
            "修正值": format_value(p.correction),
            "差值": format_value(p.diff),
            "偏差": format_value(p.deviation),
            "是否合格": p.is_pass,
            "readings": _sat_readings(p),
        } for p in sorted(t.sat_points, key=lambda x: x.id)]
        return {"success": True, "main": main, "tus_points": tus_points,
                "sat_points": sat_points, "曲線資料": t.chart_data, "報告欄位": t.report_meta}

    @staticmethod
    def update_test(test_id: int, data: Dict[str, Any]) -> bool:
        from ..models import PyrometryTest, TusPoint, SatPoint
        try:
            t = db.session.get(PyrometryTest, test_id)
            if not t or t.deleted_at is not None:
                raise ValueError("找不到該筆爐溫測試")
            test_date = _parse_date(data.get("測試日期"))
            tolerance = float(data.get("允許公差") or 0)
            setpoint = float(data.get("設定溫度") or 0)
            raw_points = data.get("points", [])
            if t.test_type == "TUS":
                judged = PyrometryService.evaluate_tus(setpoint, tolerance, raw_points)
            else:
                judged = PyrometryService.evaluate_sat(tolerance, raw_points)

            t.furnace_id = data.get("爐子ID")
            t.quarter = data.get("季別") or _quarter_of(test_date)
            t.test_date = test_date
            t.setpoint = setpoint
            t.tolerance = tolerance
            t.tester_id = data.get("測試人員") or None
            t.test_instrument = data.get("測試儀器編號") or None
            t.std_instrument = data.get("標準校正儀器編號") or None
            t.cal_due_date = _parse_date(data.get("儀器校正到期日"))
            t.is_pass = judged["是否合格"]
            t.tus_range = judged.get("TUS均勻度極差")
            t.tus_max_pos = judged.get("TUS最大正偏差")
            t.tus_max_neg = judged.get("TUS最大負偏差")
            if "曲線資料" in data:
                t.chart_data = data.get("曲線資料") or None
            if "報告欄位" in data:
                t.report_meta = data.get("報告欄位") or None
            t.note = data.get("備註") or None

            TusPoint.query.filter_by(test_id=test_id).delete()
            SatPoint.query.filter_by(test_id=test_id).delete()
            for p in judged["points"]:
                if t.test_type == "TUS":
                    tch = p.get("頻道")
                    db.session.add(TusPoint(
                        test_id=t.id, position=p.get("點位"), tc_no=p.get("熱電偶編號"),
                        channel=int(tch) if tch is not None else None,
                        correction=_to_float(p.get("修正值")),
                        temp_max=_to_float(p.get("最高溫")),
                        temp_min=_to_float(p.get("最低溫")),
                        max_dev=_to_float(p.get("最大偏差")),
                        is_pass=p.get("是否合格", True)))
                else:
                    ch = p.get("頻道")
                    db.session.add(SatPoint(
                        test_id=t.id, zone=p.get("控溫區"),
                        channel=int(ch) if ch is not None else None,
                        readings=p.get("readings") or None,
                        diff=_to_float(p.get("差值")),
                        correction=_to_float(p.get("修正值")),
                        deviation=_to_float(p.get("偏差")),
                        is_pass=p.get("是否合格", True)))
            db.session.commit()
            return True
        except Exception as e:
            db.session.rollback()
            raise e

    @staticmethod
    def delete_test(test_id: int) -> bool:
        from ..models import PyrometryTest
        try:
            t = db.session.get(PyrometryTest, test_id)
            if not t or t.deleted_at is not None:
                raise ValueError("找不到該筆爐溫測試")
            t.soft_delete()
            db.session.commit()
            return True
        except ValueError:
            raise
        except Exception as e:
            db.session.rollback()
            raise e

    @staticmethod
    def search_tests(args: Dict[str, Any]) -> Dict[str, Any]:
        from ..models import PyrometryTest
        q = PyrometryTest.query.filter(PyrometryTest.deleted_at.is_(None))
        if args.get("furnace_id"):
            q = q.filter(PyrometryTest.furnace_id == int(args["furnace_id"]))
        if args.get("test_type"):
            q = q.filter(PyrometryTest.test_type == args["test_type"])
        if args.get("quarter"):
            q = q.filter(PyrometryTest.quarter == args["quarter"])
        if args.get("is_pass") in ("0", "1"):
            q = q.filter(PyrometryTest.is_pass.is_(args["is_pass"] == "1"))
        if args.get("date_from"):
            q = q.filter(PyrometryTest.test_date >= _parse_date(args["date_from"]))
        if args.get("date_to"):
            q = q.filter(PyrometryTest.test_date <= _parse_date(args["date_to"]))
        page = int(args.get("page", 1))
        page_size = int(args.get("page_size", 20))
        total = q.count()
        pg = q.order_by(PyrometryTest.test_date.desc(), PyrometryTest.id.desc()).paginate(
            page=page, per_page=page_size, error_out=False)
        data = [{
            "識別碼": t.id, "爐號": t.furnace.code if t.furnace else "",
            "測試類型": t.test_type, "季別": t.quarter or "",
            "測試日期": format_value(t.test_date), "是否合格": t.is_pass,
            "測試人員姓名": t.tester.name if t.tester else "",
        } for t in pg.items]
        return {"success": True, "data": data, "total": total, "page": page,
                "page_size": page_size, "total_pages": pg.pages}

    # ---------- 到期計算 ----------
    @staticmethod
    def _due_from_last(last, freq_months, today=None) -> Dict[str, Any]:
        """依「最近一筆測試」計算到期狀態。抽出共用邏輯，供逐筆查詢與看板批次查詢共用。"""
        today = today or date.today()
        if not last:
            return {"最近測試日": None, "下次應測日": None, "狀態": "尚無紀錄"}
        next_due = last.test_date + relativedelta(months=int(freq_months or 3))
        if not last.is_pass:
            status = "不合格"
        elif next_due < today:
            status = "逾期"
        elif next_due - today <= timedelta(days=14):
            status = "即將到期"
        else:
            status = "正常"
        return {"最近測試日": format_value(last.test_date),
                "下次應測日": format_value(next_due), "狀態": status}

    @staticmethod
    def _due_for(furnace_id: int, test_type: str, freq_months: int, today=None) -> Dict[str, Any]:
        from ..models import PyrometryTest
        last = PyrometryTest.query.filter(
            PyrometryTest.furnace_id == furnace_id,
            PyrometryTest.test_type == test_type,
            PyrometryTest.deleted_at.is_(None),
        ).order_by(PyrometryTest.test_date.desc()).first()
        return PyrometryService._due_from_last(last, freq_months, today)

    @staticmethod
    def furnace_due_status(furnace_id: int, today=None) -> Dict[str, Any]:
        f = db.session.get(Furnace, furnace_id)
        if not f:
            raise ValueError("找不到該爐子設備")
        return {
            "TUS": PyrometryService._due_for(furnace_id, "TUS", f.tus_freq_months, today),
            "SAT": PyrometryService._due_for(furnace_id, "SAT", f.sat_freq_months, today),
        }

    # ---------- 看板與趨勢 ----------
    @staticmethod
    def dashboard(today=None) -> List[Dict[str, Any]]:
        from ..models import PyrometryTest
        furnaces = Furnace.query.filter(Furnace.is_active.is_(True)).order_by(Furnace.code).all()
        if not furnaces:
            return []

        # 一次撈出所有相關測試（依日期新→舊），在記憶體分組，取代原本逐爐 N+1 查詢
        furnace_ids = [f.id for f in furnaces]
        tests = PyrometryTest.query.filter(
            PyrometryTest.furnace_id.in_(furnace_ids),
            PyrometryTest.deleted_at.is_(None),
        ).order_by(PyrometryTest.test_date.desc(), PyrometryTest.id.desc()).all()

        latest_by_type: Dict[Any, Any] = {}   # (爐子ID, 測試類型) → 最近一筆
        latest_overall: Dict[int, Any] = {}   # 爐子ID → 最近一筆（不分類型）
        for t in tests:
            latest_overall.setdefault(t.furnace_id, t)
            latest_by_type.setdefault((t.furnace_id, t.test_type), t)

        rows = []
        for f in furnaces:
            last = latest_overall.get(f.id)
            rows.append({
                "爐子ID": f.id, "爐號": f.code, "名稱": f.name,
                "製程類型": f.process_type or "",
                "TUS": PyrometryService._due_from_last(latest_by_type.get((f.id, "TUS")), f.tus_freq_months, today),
                "SAT": PyrometryService._due_from_last(latest_by_type.get((f.id, "SAT")), f.sat_freq_months, today),
                "最近結果": ({"測試類型": last.test_type, "測試日期": format_value(last.test_date),
                            "是否合格": last.is_pass} if last else None),
            })
        return rows

    @staticmethod
    def tus_trend(furnace_id: int) -> List[Dict[str, Any]]:
        from ..models import PyrometryTest
        tests = PyrometryTest.query.filter(
            PyrometryTest.furnace_id == furnace_id,
            PyrometryTest.test_type == "TUS",
            PyrometryTest.deleted_at.is_(None),
        ).order_by(PyrometryTest.test_date.asc()).all()
        return [{
            "測試日期": format_value(t.test_date), "季別": t.quarter or "",
            "均勻度極差": format_value(t.tus_range),
            "最大正偏差": format_value(t.tus_max_pos),
            "最大負偏差": format_value(t.tus_max_neg),
            "是否合格": t.is_pass,
        } for t in tests]

    # ---------- 報告匯出 ----------
    @staticmethod
    def export_test_xlsx(test_id: int) -> bytes:
        import io
        from openpyxl import Workbook
        from . import pyrometry_report as rpt
        detail = PyrometryService.get_test(test_id)
        main = detail["main"]
        meta = detail.get("報告欄位") or {}
        setpoint = float(main.get("設定溫度") or 0)
        tol = float(main.get("允許公差") or 0)
        tc_corr = PyrometryService.thermocouple_correction(setpoint)

        wb = Workbook()
        wb.remove(wb.active)   # 移除預設空白表
        if main["測試類型"] == "TUS":
            rpt.build_tus_sheet(wb, detail, meta, tc_corr)
            cd = detail.get("曲線資料") or {}
            if cd.get("時間"):
                rpt.build_raw_chart_sheet(
                    wb, cd.get("時間"), cd.get("數值"), setpoint, tol,
                    sheet_name="原始數據-記錄器", title="測試儀器（記錄器）溫度曲線",
                    s_start=cd.get("穩定開始", 0), s_end=cd.get("穩定結束", len(cd["時間"]) - 1))
            if cd.get("爐體數值"):
                rpt.build_raw_chart_sheet(
                    wb, cd.get("爐體時間") or cd.get("時間"), cd.get("爐體數值"), setpoint, tol,
                    sheet_name="原始數據-爐體", title="爐體記錄溫度曲線")
        else:
            rpt.build_sat_sheet(wb, detail, meta, tc_corr)
        if not wb.sheetnames:    # 保險：至少一張表
            wb.create_sheet("報告")
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
