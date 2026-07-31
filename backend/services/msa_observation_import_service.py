"""MSA 觀測 Excel 的兩階段受控匯入。

預覽只解析與比對，不寫入任何觀測；確認以檔案指紋維持冪等，並在
單一交易內寫入所有無問題的列。公式儲存格不得作為正式讀值。
"""

from datetime import datetime, timezone
from pathlib import Path
import hashlib

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from ..extensions import db
from ..models import (
    MsaObservation,
    MsaObservationImportBatch,
    MsaPlanVersion,
)
from ..utils import log_audit
from .msa_errors import MsaConflict, MsaNotFound, MsaValidationError
from .msa_observation_service import MsaObservationService
from .msa_plan_lookup import require_frozen_plan


PARSER_VERSION = "msa-observation-xlsx-1"
MAX_FILE_BYTES = 5 * 1024 * 1024
MAX_ROWS = 10_000
MAX_COLUMNS = 200

# 欄位順序固定，讓問題回報可以直接指向 Excel 儲存格。
COLUMN_TASK_ORDER = 1
COLUMN_PART_BLIND_CODE = 2
COLUMN_APPRAISER_BLIND_CODE = 3
COLUMN_VALUE = 4
COLUMN_MEASURED_AT = 5

EXPECTED_HEADER = (
    "任務順序", "零件盲碼", "評價人盲碼", "讀值", "量測時間",
)


class MsaObservationImportService:
    """以 plan 盲測矩陣為準比對 Excel，逐格回報問題。"""

    # ------------------------------------------------------------------
    # 預覽
    # ------------------------------------------------------------------

    @staticmethod
    def preview(
        plan_id: int, upload, actor_id: int,
    ) -> MsaObservationImportBatch:
        path, file_name = _resolve_upload(upload)
        plan = MsaObservationImportService._frozen_plan(plan_id)
        digest, size = _hash_file(path)
        if size > MAX_FILE_BYTES:
            raise MsaValidationError(
                "MSA_IMPORT_FILE_TOO_LARGE",
                "匯入檔案超過允許大小",
                details={"max_bytes": MAX_FILE_BYTES, "size": size},
            )

        existing = MsaObservationImportBatch.query.filter_by(
            plan_version_id=plan.id, file_sha256=digest,
        ).one_or_none()
        if existing is not None:
            return existing

        rows, issues = MsaObservationImportService._parse(path, plan)
        batch = MsaObservationImportBatch(
            plan_version_id=plan.id,
            file_name=file_name,
            file_sha256=digest,
            plan_hash=plan.plan_hash,
            status="previewed",
            issues=issues,
            stats={
                "parser_version": PARSER_VERSION,
                "total_rows": len(rows) + len(
                    {issue["row"] for issue in issues}
                ),
                "clean_rows": len(rows),
                "issue_rows": len({issue["row"] for issue in issues}),
                "rows": rows,
            },
            created_by_id=actor_id,
        )
        try:
            db.session.add(batch)
            db.session.flush()
            log_audit(
                actor_id,
                "preview_observation_import",
                "msa_observation_import",
                batch.id,
                new_val={
                    "plan_version_id": plan.id,
                    "file_sha256": digest,
                    "clean_rows": len(rows),
                    "issue_rows": batch.stats["issue_rows"],
                },
            )
            db.session.commit()
            return batch
        except Exception:
            db.session.rollback()
            raise

    # ------------------------------------------------------------------
    # 確認
    # ------------------------------------------------------------------

    @staticmethod
    def confirm(batch_id: int, actor_id: int) -> MsaObservationImportBatch:
        try:
            batch = (
                MsaObservationImportBatch.query
                .filter_by(id=batch_id)
                .with_for_update()
                .one_or_none()
            )
            if batch is None:
                raise MsaNotFound(
                    "MSA_IMPORT_BATCH_NOT_FOUND",
                    "找不到觀測匯入批次",
                    details={"batch_id": batch_id},
                )
            if batch.status == "confirmed":
                return batch

            # 鎖定計畫列，避免與逐筆盲測輸入（record）並發時搶到相同的
            # 實際輸入順序；持有鎖期間，順序可安全地在記憶體中遞增。
            plan = MsaObservationImportService._frozen_plan(
                batch.plan_version_id, for_update=True,
            )
            if batch.plan_hash != plan.plan_hash:
                raise MsaConflict(
                    "MSA_IMPORT_PLAN_CHANGED",
                    "計畫內容已變更，請重新預覽匯入檔",
                    details={
                        "batch_plan_hash": batch.plan_hash,
                        "plan_hash": plan.plan_hash,
                    },
                )

            written = 0
            next_entry_order = MsaObservationService._next_entry_order(plan.id)
            for row in batch.stats.get("rows", []):
                observation = MsaObservation(
                    plan_version_id=plan.id,
                    part_id=row["part_id"],
                    appraiser_id=row["appraiser_id"],
                    trial_no=row["trial_no"],
                    requested_order=row["requested_order"],
                    actual_entry_order=next_entry_order,
                    numeric_value=row.get("numeric_value"),
                    attribute_value=row.get("attribute_value"),
                    measured_at=_parse_timestamp(row.get("measured_at")),
                    source="excel_import",
                    entered_by_id=actor_id,
                    is_effective=True,
                    import_batch_id=batch.id,
                )
                db.session.add(observation)
                next_entry_order += 1
                written += 1
            db.session.flush()

            batch.status = "confirmed"
            batch.stats = {
                **batch.stats,
                "written": written,
                "skipped": batch.stats.get("issue_rows", 0),
            }
            MsaObservationService._advance_study_to_collecting(
                plan.study_id, actor_id,
            )
            log_audit(
                actor_id,
                "confirm_observation_import",
                "msa_observation_import",
                batch.id,
                new_val={"written": written, "plan_version_id": plan.id},
            )
            db.session.commit()
            return batch
        except Exception:
            db.session.rollback()
            raise

    # ------------------------------------------------------------------
    # 解析
    # ------------------------------------------------------------------

    @staticmethod
    def _parse(path: Path, plan: MsaPlanVersion):
        workbook = load_workbook(path, read_only=True, data_only=False)
        try:
            sheet = workbook.worksheets[0]
            sheet_name = sheet.title
            if sheet.max_row and sheet.max_row > MAX_ROWS + 1:
                raise MsaValidationError(
                    "MSA_IMPORT_TOO_MANY_ROWS",
                    "匯入檔案列數超過允許上限",
                    details={"max_rows": MAX_ROWS},
                )
            if sheet.max_column and sheet.max_column > MAX_COLUMNS:
                raise MsaValidationError(
                    "MSA_IMPORT_TOO_MANY_COLUMNS",
                    "匯入檔案欄數超過允許上限",
                    details={"max_columns": MAX_COLUMNS},
                )

            tasks = {
                item["requested_order"]: item
                for item in plan.randomized_order
            }
            part_codes = {part.id: part.blind_code for part in plan.parts}
            appraiser_codes = {
                row.id: row.blind_code for row in plan.appraisers
            }
            categories = list(plan.category_set or [])

            rows = []
            issues = []
            for row_index, cells in enumerate(
                sheet.iter_rows(min_row=2, max_col=COLUMN_MEASURED_AT), start=2
            ):
                values = [cell.value for cell in cells]
                if all(value is None for value in values):
                    continue
                MsaObservationImportService._parse_row(
                    row_index, values, sheet_name, tasks, part_codes,
                    appraiser_codes, categories, rows, issues,
                )
            return rows, issues
        finally:
            workbook.close()

    @staticmethod
    def _parse_row(
        row_index, values, sheet_name, tasks, part_codes, appraiser_codes,
        categories, rows, issues,
    ) -> None:
        def add_issue(column, code, message):
            issues.append({
                "sheet": sheet_name,
                "row": row_index,
                "cell": f"{get_column_letter(column)}{row_index}",
                "code": code,
                "message": message,
            })

        raw_order = values[COLUMN_TASK_ORDER - 1]
        try:
            task_order = int(raw_order)
        except (TypeError, ValueError):
            add_issue(
                COLUMN_TASK_ORDER,
                "MSA_OBSERVATION_TASK_UNKNOWN",
                "任務順序必須是整數",
            )
            return
        task = tasks.get(task_order)
        if task is None:
            add_issue(
                COLUMN_TASK_ORDER,
                "MSA_OBSERVATION_TASK_UNKNOWN",
                "任務順序不存在於此計畫版本",
            )
            return

        expected_part = part_codes.get(task["part_id"])
        expected_appraiser = appraiser_codes.get(task["appraiser_id"])
        supplied_part = _text(values[COLUMN_PART_BLIND_CODE - 1])
        supplied_appraiser = _text(values[COLUMN_APPRAISER_BLIND_CODE - 1])
        if supplied_part is not None and supplied_part != expected_part:
            add_issue(
                COLUMN_PART_BLIND_CODE,
                "MSA_OBSERVATION_BLIND_CODE_MISMATCH",
                f"零件盲碼與計畫不符，應為 {expected_part}",
            )
            return
        if (
            supplied_appraiser is not None
            and supplied_appraiser != expected_appraiser
        ):
            add_issue(
                COLUMN_APPRAISER_BLIND_CODE,
                "MSA_OBSERVATION_BLIND_CODE_MISMATCH",
                f"評價人盲碼與計畫不符，應為 {expected_appraiser}",
            )
            return

        raw_value = values[COLUMN_VALUE - 1]
        if isinstance(raw_value, str) and raw_value.startswith("="):
            add_issue(
                COLUMN_VALUE,
                "MSA_OBSERVATION_FORMULA_NOT_ALLOWED",
                "公式儲存格不可作為正式讀值",
            )
            return

        numeric_value = None
        attribute_value = None
        if categories:
            attribute_value = _text(raw_value)
            if attribute_value not in categories:
                add_issue(
                    COLUMN_VALUE,
                    "MSA_OBSERVATION_CATEGORY_UNKNOWN",
                    "計數分類不在計畫保存的類別集合內",
                )
                return
        else:
            try:
                numeric_value = MsaObservationService._resolve_value(
                    _CategorylessPlan(),
                    {"numeric_value": raw_value},
                )[0]
            except MsaValidationError as error:
                add_issue(COLUMN_VALUE, error.code, error.message)
                return

        rows.append({
            "row": row_index,
            "requested_order": task_order,
            "part_id": task["part_id"],
            "appraiser_id": task["appraiser_id"],
            "trial_no": task["trial_no"],
            "numeric_value": (
                str(numeric_value) if numeric_value is not None else None
            ),
            "attribute_value": attribute_value,
            "measured_at": _text(values[COLUMN_MEASURED_AT - 1]),
        })

    @staticmethod
    def _frozen_plan(plan_id: int, *, for_update: bool = False) -> MsaPlanVersion:
        return require_frozen_plan(
            plan_id,
            message="計畫尚未凍結，不可匯入正式讀值",
            for_update=for_update,
        )


class _CategorylessPlan:
    """讓計量讀值重用觀測服務驗證時，明確表示沒有計數類別集合。"""

    id = None
    category_set: list = []


def _resolve_upload(upload):
    """接受檔案路徑或 Flask 上傳物件，一律回傳可讀路徑與原始檔名。"""
    if isinstance(upload, (str, Path)):
        path = Path(upload)
        if not path.exists():
            raise MsaValidationError(
                "MSA_IMPORT_FILE_MISSING",
                "找不到要匯入的檔案",
                details={"path": str(path)},
            )
        _assert_xlsx(path.name)
        return path, path.name

    file_name = getattr(upload, "filename", None)
    if not file_name:
        raise MsaValidationError(
            "MSA_IMPORT_FILE_MISSING", "請選擇要匯入的檔案", details={},
        )
    _assert_xlsx(file_name)
    import tempfile

    with tempfile.NamedTemporaryFile(
        delete=False, suffix=".xlsx"
    ) as temporary:
        upload.save(temporary.name)
        return Path(temporary.name), file_name


def _assert_xlsx(file_name: str) -> None:
    if not str(file_name).lower().endswith(".xlsx"):
        raise MsaValidationError(
            "MSA_IMPORT_FILE_TYPE_INVALID",
            "觀測匯入只接受 .xlsx 檔案",
            details={"file_name": str(file_name)[:120]},
        )


def _hash_file(path: Path):
    digest = hashlib.sha256()
    size = 0
    with open(path, "rb") as stream:
        for chunk in iter(lambda: stream.read(65536), b""):
            digest.update(chunk)
            size += len(chunk)
    return digest.hexdigest(), size


def _text(value):
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _parse_timestamp(value):
    if value is None:
        return None
    try:
        parsed = datetime.fromisoformat(str(value))
    except ValueError:
        return None
    return (
        parsed if parsed.tzinfo is not None
        else parsed.replace(tzinfo=timezone.utc)
    )
