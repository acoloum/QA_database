import numpy as np
import json
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.chart.layout import Layout, ManualLayout
from .spc_stability import RULE_LABELS
from ..extensions import db
from ..models import SpcStudyVersion
from .spc_errors import SpcNotFound, SpcValidationError


class SpcReportService:
    @staticmethod
    def _stats_from_version(version: SpcStudyVersion) -> dict:
        """只使用不可變研究版本與樣本快照建立報表資料。"""

        chart = version.chart_result or {}
        location = chart.get("location") or {}
        variation = chart.get("variation") or {}
        samples = list(version.samples)
        avgs = list(location.get("values") or [
            float(np.mean(sample.values)) for sample in samples
        ])
        variation_values = list(variation.get("values") or [
            float(np.ptp(sample.values)) for sample in samples
        ])
        x_cls = list(location.get("cl") or [])
        x_ucls = list(location.get("ucl") or [])
        x_lcls = list(location.get("lcl") or [])
        r_cls = list(variation.get("cl") or [])
        r_ucls = list(variation.get("ucl") or [])
        r_lcls = list(variation.get("lcl") or [])

        def first(values):
            return values[0] if values else None

        limits = version.limit_versions[-1] if version.limit_versions else None
        events = [event for item in version.limit_versions for event in item.events]
        excluded_count = sum(
            1
            for sample in samples
            for item in (sample.exclusion_snapshot or [])
            if item.get("excluded")
        )
        distribution_values = [
            float(value)
            for sample in samples
            for value in (sample.distribution_values or sample.values or [])
        ]
        metadata = {
            "研究版本ID": version.id,
            "研究ID": version.study_id,
            "研究類型": version.study.study_type,
            "資料來源": version.study.source,
            "完整篩選條件": json.dumps(
                version.study.filters or {}, ensure_ascii=False, sort_keys=True
            ),
            "規格快照": json.dumps(
                version.specification_snapshot or {}, ensure_ascii=False, sort_keys=True
            ),
            "每組樣本數": ", ".join(
                str(value) for value in (chart.get("subgroup_sizes") or [
                    len(sample.values or []) for sample in samples
                ])
            ),
            "圖表選型": chart.get("chart_type"),
            "時間模型": json.dumps(
                version.time_model_result or {}, ensure_ascii=False, sort_keys=True
            ),
            "分布證據": json.dumps(
                version.distribution_result or {}, ensure_ascii=False, sort_keys=True
            ),
            "穩定性證據": json.dumps(
                version.stability_result or {}, ensure_ascii=False, sort_keys=True
            ),
            "指標適用性": json.dumps(
                version.capability_result or {}, ensure_ascii=False, sort_keys=True
            ),
            "資料雜湊": version.data_hash,
            "方法版本": version.method_version,
            "程式版本": version.code_version or "未記錄",
            "版本狀態": version.status,
            "稽核不完整": "是" if version.audit_incomplete else "否",
            "建立者ID": version.created_by,
            "建立時間": version.created_at.isoformat() if version.created_at else None,
            "核准界限版次": limits.revision if limits else None,
            "核准者ID": limits.approved_by if limits else None,
            "核准時間": (
                limits.approved_at.isoformat() if limits and limits.approved_at else None
            ),
            "OCAP事件數": len(events),
            "OCAP結案數": sum(1 for event in events if event.status == "closed"),
        }
        return {
            "schema_version": version.method_version,
            "labels": [sample.subgroup_key for sample in samples],
            "dates": [sample.sample_timestamp or "" for sample in samples],
            "ids": [str(sample.source_record_id) for sample in samples],
            "avgs": avgs,
            "ranges": variation_values,
            "all_values": distribution_values,
            "subgroup_sizes": list(chart.get("subgroup_sizes") or []),
            "avg_subgroup_size": (
                int(round(float(np.mean(chart.get("subgroup_sizes")))))
                if chart.get("subgroup_sizes") else None
            ),
            "x_cl": first(x_cls), "x_ucl": first(x_ucls), "x_lcl": first(x_lcls),
            "r_cl": first(r_cls), "r_ucl": first(r_ucls), "r_lcl": first(r_lcls),
            "x_cls": x_cls, "x_ucls": x_ucls, "x_lcls": x_lcls,
            "r_cls": r_cls, "r_ucls": r_ucls, "r_lcls": r_lcls,
            "charts": chart,
            "stability": version.stability_result or {},
            "distribution": version.distribution_result or {},
            "time_model": version.time_model_result or {},
            "process_capability": version.capability_result or {},
            "capability": version.capability_result or {},
            "applicability": version.applicability_result or {},
            "tolerance": version.specification_snapshot or {},
            "excluded_count": excluded_count,
            "limits_frozen": bool(limits and limits.status == "active"),
            "study_version": {
                "id": version.id, "version_no": version.version_no,
                "status": version.status, "data_hash": version.data_hash,
            },
            "version_metadata": metadata,
            "version_samples": [{
                "子組順序": sample.subgroup_order,
                "子組鍵": sample.subgroup_key,
                "樣本時間": sample.sample_timestamp,
                "來源紀錄ID": json.dumps(sample.source_record_ids or [], ensure_ascii=False),
                "來源量測ID": json.dumps(sample.source_measurement_ids or [], ensure_ascii=False),
                "管制圖量測值": json.dumps(sample.values or [], ensure_ascii=False),
                "分布分析值": json.dumps(
                    sample.distribution_values or [], ensure_ascii=False
                ),
                "排除快照": json.dumps(
                    sample.exclusion_snapshot or [], ensure_ascii=False, sort_keys=True
                ),
            } for sample in samples],
        }

    @staticmethod
    def generate_version_report(
        version_id: int,
        *,
        expected_source: str | None = None,
        expected_filters: dict | None = None,
    ) -> BytesIO:
        """由已保存研究版本重建 Excel，不讀取目前來源資料。"""

        version = db.session.get(SpcStudyVersion, version_id)
        if version is None:
            raise SpcNotFound("SPC_STUDY_VERSION_NOT_FOUND", "找不到 SPC 研究版本")
        if expected_source is not None and version.study.source != expected_source:
            raise SpcValidationError(
                "SPC_REPORT_SOURCE_MISMATCH", "研究版本不屬於目前匯出頁面"
            )
        if expected_filters is not None:
            def normalize(values):
                result = {}
                for key, value in sorted(values.items()):
                    if value is None:
                        result[key] = ""
                    elif key in {"m_id", "op_id", "cust_id"} and str(value).isdigit():
                        result[key] = int(value)
                    elif hasattr(value, "isoformat"):
                        result[key] = value.isoformat()
                    else:
                        result[key] = value
                return result
            if normalize(version.study.filters or {}) != normalize(expected_filters):
                raise SpcValidationError(
                    "SPC_REPORT_FILTER_MISMATCH",
                    "研究版本的完整篩選條件與目前匯出條件不一致",
                )
        return SpcReportService.generate_report(
            SpcReportService._stats_from_version(version),
            version.study.characteristic,
            dict(version.study.filters or {}),
        )

    @staticmethod
    def generate_report(stats_data: dict, field: str, filters: dict) -> BytesIO:
        """Generate an SPC report as Excel file with embedded charts"""
        wb = Workbook()

        # --- Sheet 1: Summary ---
        ws = wb.active
        ws.title = "SPC 統計摘要"

        # Styles
        title_font = Font(name="微軟正黑體", size=14, bold=True)
        header_font = Font(name="微軟正黑體", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
        normal_font = Font(name="微軟正黑體", size=10)
        good_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        warn_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        bad_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # Title
        ws.merge_cells('A1:F1')
        ws['A1'] = f"SPC 統計分析報告 - {field}"
        ws['A1'].font = title_font

        # Report info
        ws['A3'] = "報告產生時間："
        ws['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws['A4'] = "檢驗項目："
        ws['B4'] = field
        ws['A5'] = "客戶名稱："
        ws['B5'] = filters.get('customer', filters.get('vendor', '全部'))
        ws['A6'] = "材質："
        ws['B6'] = filters.get('material', '全部')
        ws['A7'] = "規格："
        ws['B7'] = filters.get('spec', '全部')
        ws['A8'] = "日期範圍："
        ws['B8'] = f"{filters.get('start_date', '不限')} ~ {filters.get('end_date', '不限')}"
        for r in range(3, 9):
            ws[f'A{r}'].font = Font(name="微軟正黑體", size=10, bold=True)
            ws[f'B{r}'].font = normal_font

        # --- Basic Statistics ---
        row = 10
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = "基本統計量"
        ws[f'A{row}'].font = Font(name="微軟正黑體", size=12, bold=True)
        row += 1

        avgs = stats_data.get('avgs', [])
        pc = stats_data.get('process_capability', {})
        if avgs:
            stats_items = [
                ("有效樣本數", len(avgs)),
                ("平均值 (X̄)", round(float(np.mean(avgs)), 4)),
                ("標準差 (σ)", round(float(np.std(avgs, ddof=1)), 4) if len(avgs) > 1 else "N/A"),
                ("最小值", round(min(avgs), 4)),
                ("最大值", round(max(avgs), 4)),
                ("全距 (R)", round(max(avgs) - min(avgs), 4)),
            ]

            # Add distribution stats if available
            dist_stats = stats_data.get('distribution_stats', {})
            if dist_stats:
                stats_items.append(("偏態係數 (Skewness)", dist_stats.get('skewness', 'N/A')))
                stats_items.append(("峰態係數 (Kurtosis)", dist_stats.get('kurtosis', 'N/A')))
                stats_items.append(("常態性評估", dist_stats.get('normality_label', 'N/A')))

            headers = ["統計項目", "數值"]
            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col_idx, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
            row += 1
            for name, value in stats_items:
                cell_name = ws.cell(row=row, column=1, value=name)
                cell_name.font = normal_font
                cell_name.border = thin_border
                cell_val = ws.cell(row=row, column=2, value=value)
                cell_val.font = normal_font
                cell_val.border = thin_border
                cell_val.alignment = Alignment(horizontal='center')
                # Highlight normality assessment
                if name == "常態性評估":
                    normality = dist_stats.get('normality', '')
                    if normality == 'good':
                        cell_val.fill = good_fill
                    elif normality == 'moderate':
                        cell_val.fill = warn_fill
                    elif normality == 'poor':
                        cell_val.fill = bad_fill
                row += 1

        # --- 研究資訊（AIAG-VDA SPC 2026 §11.2 報告要素）---
        stability = stats_data.get('stability') or {}
        targets = pc.get('targets') or {}
        dist_info = pc.get('distribution') or {}
        applicable = pc.get('applicable')

        row += 1
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = "研究資訊（AIAG-VDA SPC 2026）"
        ws[f'A{row}'].font = Font(name="微軟正黑體", size=12, bold=True)
        row += 1

        version_metadata = stats_data.get('version_metadata') or {}
        study_items = [
            ("研究類型", version_metadata.get("研究類型") or "即時預覽（未保存版本）"),
            ("適用指數", "能力 Cp/Cpk（穩定）" if applicable == "capability"
                        else "績效 Pp/Ppk（不穩定或穩定性未證明）"),
            ("計算方法", f"{pc.get('method', 'G')} 法（分位數法；常態時等同 6s 公式）"),
            ("分布模型", dist_info.get('label', '常態分布')),
            ("穩定性判定", "穩定（統計受控）" if stability.get('stable')
                          else ("不穩定" if stability.get('stable') is False else "無法評估")),
            ("使用之穩定性準則", "、".join(RULE_LABELS.get(r, r) for r in stability.get('rules_used', [])) or "—"),
            ("特性重要度", targets.get('class', '其他')),
            ("Ppk/Cpk 目標值", targets.get('pk_target')),
            ("目標值樣本數調整", f"已依樣本數上修（{targets.get('confidence', '95%')}）" if targets.get('adjusted') else "無"),
            ("樣本數(個別值)", len(stats_data.get('all_values', []))),
            ("子組數", len(stats_data.get('avgs', []))),
            ("排除之離群值筆數", stats_data.get('excluded_count', 0)),
            ("初步值註記", "是（n<125 或子組<25）" if pc.get('preliminary') else "否"),
            ("圖表選型", version_metadata.get("圖表選型") or (stats_data.get('charts') or {}).get('chart_type')),
            ("時間模型", (stats_data.get('time_model') or {}).get('model') or (stats_data.get('time_model') or {}).get('candidate')),
            ("資料雜湊", version_metadata.get("資料雜湊") or stats_data.get('data_hash')),
            ("方法版本", version_metadata.get("方法版本") or stats_data.get('schema_version')),
        ]
        for name, value in study_items:
            ws.cell(row=row, column=1, value=name).font = normal_font
            ws.cell(row=row, column=1).border = thin_border
            cell_val = ws.cell(row=row, column=2, value=value if value is not None else 'N/A')
            cell_val.font = normal_font
            cell_val.border = thin_border
            row += 1
        row += 1

        # --- Control Limits ---
        row += 1
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = "管制界限"
        ws[f'A{row}'].font = Font(name="微軟正黑體", size=12, bold=True)
        row += 1

        chart_type = (stats_data.get('charts') or {}).get('chart_type', 'xbar_r')
        variation_label = {'xbar_s': 'S', 'xbar_r': 'R', 'i_mr': 'MR'}.get(
            chart_type, '變異'
        )
        avg_n = stats_data.get('avg_subgroup_size', 5)
        cl_items = [
            ("管制圖選型", chart_type),
            ("平均子群大小 (n)", avg_n),
            ("X̄ 中心線 (CL)", stats_data.get('x_cl', 0)),
            ("X̄ 管制上限 (UCL)", stats_data.get('x_ucl', 0)),
            ("X̄ 管制下限 (LCL)", stats_data.get('x_lcl', 0)),
            (f"{variation_label} 中心線", stats_data.get('r_cl')),
            (f"{variation_label} 管制上限 (UCL)", stats_data.get('r_ucl')),
            (f"{variation_label} 管制下限 (LCL)", stats_data.get('r_lcl')),
        ]
        for col_idx, h in enumerate(["管制項目", "數值"], 1):
            cell = ws.cell(row=row, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        row += 1
        for name, value in cl_items:
            ws.cell(row=row, column=1, value=name).font = normal_font
            ws.cell(row=row, column=1).border = thin_border
            cell_val = ws.cell(row=row, column=2, value=round(value, 4) if isinstance(value, (int, float)) else value)
            cell_val.font = normal_font
            cell_val.border = thin_border
            cell_val.alignment = Alignment(horizontal='center')
            row += 1

        # --- Process Capability ---
        if pc.get('available'):
            row += 1
            ws.merge_cells(f'A{row}:F{row}')
            ws[f'A{row}'] = "製程能力指標"
            ws[f'A{row}'].font = Font(name="微軟正黑體", size=12, bold=True)
            row += 1

            for col_idx, h in enumerate(["指標", "數值", "等級"], 1):
                cell = ws.cell(row=row, column=col_idx, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
            row += 1

            pk_target = targets.get('pk_target', 1.33)

            def get_cpk_grade(val):
                if val is None:
                    return ("N/A", None)
                if val >= pk_target:
                    return (f"達標 (≥{pk_target:.2f})", good_fill)
                return (f"未達標 (<{pk_target:.2f})", bad_fill)

            method = pc.get('method', 'G')
            cpk_label = f"Cpk.{method} (修正製程能力,僅穩定時)"
            ppk_label = f"Ppk.{method} (修正製程績效)"
            pc_items = [
                ("USL (規格上限)", pc.get('usl'), False),
                ("LSL (規格下限)", pc.get('lsl'), False),
                (f"Pp.{method} (製程績效)", pc.get('pp'), False),
                (ppk_label, pc.get('ppk'), False),
                (f"Cp.{method} (製程能力,僅穩定時)", pc.get('cp'), True),
                (cpk_label, pc.get('cpk'), True),
                ("Cw (組內能力參考)", pc.get('cw'), False),
                ("Cwk (組內能力參考)", pc.get('cwk'), False),
                ("σ_overall (整體標準差)", pc.get('sigma_overall'), False),
                ("σ_within (組內標準差)", pc.get('sigma_within'), False),
            ]

            for name, value, capability_only in pc_items:
                ws.cell(row=row, column=1, value=name).font = normal_font
                ws.cell(row=row, column=1).border = thin_border
                if value is None and capability_only:
                    display_val = "不適用(未證明穩定)"
                else:
                    display_val = round(value, 4) if isinstance(value, (int, float)) else "N/A"
                cell_val = ws.cell(row=row, column=2, value=display_val)
                cell_val.font = normal_font
                cell_val.border = thin_border
                cell_val.alignment = Alignment(horizontal='center')

                # Grade column for Cpk/Ppk
                if name in (cpk_label, ppk_label):
                    grade, fill = get_cpk_grade(value)
                    cell_grade = ws.cell(row=row, column=3, value=grade)
                    cell_grade.font = normal_font
                    cell_grade.border = thin_border
                    cell_grade.alignment = Alignment(horizontal='center')
                    if fill:
                        cell_grade.fill = fill
                row += 1

            # --- PPM Row ---
            ppm = pc.get('ppm', {})
            if ppm:
                row += 1
                ws.merge_cells(f'A{row}:F{row}')
                ws[f'A{row}'] = "PPM 不良率估算"
                ws[f'A{row}'].font = Font(name="微軟正黑體", size=12, bold=True)
                row += 1
                ppm_items = [
                    ("PPM 超上限", ppm.get('upper', 0)),
                    ("PPM 超下限", ppm.get('lower', 0)),
                    ("PPM 總計", ppm.get('total', 0)),
                ]
                for col_idx, h in enumerate(["項目", "數值"], 1):
                    cell = ws.cell(row=row, column=col_idx, value=h)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                row += 1
                for name, value in ppm_items:
                    ws.cell(row=row, column=1, value=name).font = normal_font
                    ws.cell(row=row, column=1).border = thin_border
                    cell_val = ws.cell(row=row, column=2, value=round(value, 1) if isinstance(value, (int, float)) else value)
                    cell_val.font = normal_font
                    cell_val.border = thin_border
                    cell_val.alignment = Alignment(horizontal='center')
                    if name == "PPM 總計" and isinstance(value, (int, float)):
                        if value <= 3.4:
                            cell_val.fill = good_fill
                        elif value <= 6210:
                            cell_val.fill = warn_fill
                        else:
                            cell_val.fill = bad_fill
                    row += 1

            # --- Process Capability Conclusion ---
            row += 1
            ws.merge_cells(f'A{row}:F{row}')
            ws[f'A{row}'] = "製程能力結論"
            ws[f'A{row}'].font = Font(name="微軟正黑體", size=12, bold=True)
            row += 1

            pk_val = pc.get('cpk') if applicable == 'capability' else pc.get('ppk')
            pk_label = cpk_label if applicable == 'capability' else ppk_label
            conclusion_lines = []
            if pk_val is not None:
                if pk_val >= pk_target:
                    conclusion_lines.append(f"✅ 指數 {pk_label}={pk_val} 達到特性重要度「{targets.get('class', '其他')}」目標值 {pk_target:.2f}。")
                    conclusion_lines.append("建議：維持現有製程管控。")
                else:
                    conclusion_lines.append(f"❌ 指數 {pk_label}={pk_val} 未達目標值 {pk_target:.2f}，需啟動改善（OCAP/根本原因分析）。")
                    conclusion_lines.append("建議：分析變異來源並採取矯正措施。")

                if stability.get('stable') is False:
                    conclusion_lines.append("⚠️ 製程未通過穩定性準則，僅能報告績效指數 Pp/Ppk；請先消除特殊原因後重新評估能力。")

                # Normality note
                dist_stats = stats_data.get('distribution_stats', {})
                normality = dist_stats.get('normality', '')
                if normality == 'poor':
                    conclusion_lines.append("⚠️ 注意：數據分佈明顯非常態，已改用 G 法分位數計算，但仍建議搭配其他分析方法確認。")
                elif normality == 'moderate':
                    conclusion_lines.append("📋 備註：數據分佈略偏常態，指數數值僅供參考。")

            for line in conclusion_lines:
                ws.cell(row=row, column=1, value=line).font = normal_font
                ws.merge_cells(f'A{row}:F{row}')
                row += 1

        # Column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 15

        # --- Sheet 2: Control Chart Data ---
        ws2 = wb.create_sheet(title="管制圖數據")
        # Headers include control limit columns for chart references
        data_headers = [
            "序號", "標籤", "日期", "位置統計量",
            f"變異統計量 ({variation_label})", "UCL", "CL", "LCL",
            f"{variation_label}_UCL", f"{variation_label}_CL",
            f"{variation_label}_LCL",
        ]
        for col_idx, h in enumerate(data_headers, 1):
            cell = ws2.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

        labels = stats_data.get('labels', [])
        dates = stats_data.get('dates', [])
        ranges = stats_data.get('ranges', [])
        x_ucl = stats_data.get('x_ucl')
        x_cl = stats_data.get('x_cl')
        x_lcl = stats_data.get('x_lcl')
        r_ucl = stats_data.get('r_ucl')
        r_cl = stats_data.get('r_cl')
        x_ucls = stats_data.get('x_ucls') or [x_ucl] * len(avgs)
        x_cls = stats_data.get('x_cls') or [x_cl] * len(avgs)
        x_lcls = stats_data.get('x_lcls') or [x_lcl] * len(avgs)
        r_ucls = stats_data.get('r_ucls') or [r_ucl] * len(avgs)
        r_cls = stats_data.get('r_cls') or [r_cl] * len(avgs)
        r_lcls = stats_data.get('r_lcls') or [stats_data.get('r_lcl')] * len(avgs)

        # USL/LSL columns if available
        has_spec_limits = pc.get('available') and pc.get('usl') is not None and pc.get('lsl') is not None
        if has_spec_limits:
            for col_idx, h in enumerate(["USL", "LSL"], len(data_headers) + 1):
                cell = ws2.cell(row=1, column=col_idx, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')

        for i in range(len(avgs)):
            r = i + 2
            ws2.cell(row=r, column=1, value=i + 1).border = thin_border
            ws2.cell(row=r, column=2, value=labels[i] if i < len(labels) else "").border = thin_border
            ws2.cell(row=r, column=3, value=dates[i] if i < len(dates) else "").border = thin_border
            cell_avg = ws2.cell(row=r, column=4, value=round(avgs[i], 4))
            cell_avg.border = thin_border
            point_x_ucl = x_ucls[i] if i < len(x_ucls) else x_ucl
            point_x_cl = x_cls[i] if i < len(x_cls) else x_cl
            point_x_lcl = x_lcls[i] if i < len(x_lcls) else x_lcl
            point_r_ucl = r_ucls[i] if i < len(r_ucls) else r_ucl
            point_r_cl = r_cls[i] if i < len(r_cls) else r_cl
            point_r_lcl = (
                r_lcls[i] if i < len(r_lcls) else stats_data.get('r_lcl')
            )
            if (
                point_x_ucl is not None and point_x_lcl is not None
                and (avgs[i] > point_x_ucl or avgs[i] < point_x_lcl)
            ):
                cell_avg.fill = bad_fill
            range_value = ranges[i] if i < len(ranges) else None
            cell_range = ws2.cell(
                row=r, column=5,
                value=round(range_value, 4) if range_value is not None else None,
            )
            cell_range.border = thin_border
            if (
                i < len(ranges) and ranges[i] is not None
                and (
                    (point_r_ucl is not None and ranges[i] > point_r_ucl)
                    or (point_r_lcl is not None and ranges[i] < point_r_lcl)
                )
            ):
                cell_range.fill = bad_fill
            # Control limit columns for chart series
            ws2.cell(row=r, column=6, value=round(point_x_ucl, 4) if point_x_ucl is not None else None)
            ws2.cell(row=r, column=7, value=round(point_x_cl, 4) if point_x_cl is not None else None)
            ws2.cell(row=r, column=8, value=round(point_x_lcl, 4) if point_x_lcl is not None else None)
            ws2.cell(row=r, column=9, value=round(point_r_ucl, 4) if point_r_ucl is not None else None)
            ws2.cell(row=r, column=10, value=round(point_r_cl, 4) if point_r_cl is not None else None)
            ws2.cell(row=r, column=11, value=round(point_r_lcl, 4) if point_r_lcl is not None else None)

            if has_spec_limits:
                ws2.cell(row=r, column=12, value=round(pc['usl'], 4))
                ws2.cell(row=r, column=13, value=round(pc['lsl'], 4))

        for col_idx in range(1, 13):
            ws2.column_dimensions[get_column_letter(col_idx)].width = 15

        # --- Sheet 3: Histogram Data ---
        all_values = stats_data.get('all_values', [])
        hist_data_rows = 0
        ws3 = None
        if all_values:
            ws3 = wb.create_sheet(title="直方圖數據")
            hist_headers = ["區間", "頻次", "累積百分比"]
            for col_idx, h in enumerate(hist_headers, 1):
                cell = ws3.cell(row=1, column=col_idx, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')

            arr = np.array(all_values)
            counts, bin_edges = np.histogram(arr, bins='auto')
            total = len(all_values)
            cumulative = 0
            hist_data_rows = len(counts)
            for i in range(len(counts)):
                label = f"{round(bin_edges[i], 3)} ~ {round(bin_edges[i + 1], 3)}"
                cumulative += counts[i]
                ws3.cell(row=i + 2, column=1, value=label).border = thin_border
                ws3.cell(row=i + 2, column=2, value=int(counts[i])).border = thin_border
                ws3.cell(row=i + 2, column=3, value=f"{round(cumulative / total * 100, 1)}%").border = thin_border

            for col_idx in range(1, 4):
                ws3.column_dimensions[get_column_letter(col_idx)].width = 20

        # --- Sheet 4: WECO Violations ---
        ws_weco = wb.create_sheet(title="WECO 異常清單")
        weco_headers = ["序號", "數據標籤", "類型", "量測值/全距", "違規規則"]
        for col_idx, h in enumerate(weco_headers, 1):
            cell = ws_weco.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

        # Detect WECO violations for the report (統一改用 spc_stability 的穩定性判定引擎，§9.2.2)
        weco_row = 2
        if len(avgs) > 0:
            saved_stability = stats_data.get('stability') or {}
            x_stability = saved_stability.get('location') or {"violations": []}
            r_stability = saved_stability.get('variation') or {"violations": []}

            def _stability_to_rows(stability_result, chart_type, values, point_labels):
                """將 evaluate_stability 的違規清單轉為報告列（保留原欄位結構）。

                注意：stability_result['violations'][i]['label'] 是規則的中文說明
                （例如「單點超出管制界限」），並非資料點本身的標籤；
                資料點標籤需另從 point_labels[index] 取得。
                """
                rows = []
                for v in stability_result.get('violations', []):
                    idx = v['index']
                    rows.append({
                        'label': point_labels[idx] if idx < len(point_labels) else str(idx),
                        'chart_type': chart_type,
                        'value': values[idx],
                        'reasons': [v['label']],
                    })
                return rows

            x_violations = _stability_to_rows(x_stability, 'X̄', avgs, labels)
            r_violations = _stability_to_rows(
                r_stability, variation_label, ranges, labels
            )

            for v in x_violations + r_violations:
                ws_weco.cell(row=weco_row, column=1, value=weco_row - 1).border = thin_border
                ws_weco.cell(row=weco_row, column=2, value=v['label']).border = thin_border
                ws_weco.cell(row=weco_row, column=3, value=v['chart_type']).border = thin_border
                ws_weco.cell(row=weco_row, column=4, value=round(v['value'], 4)).border = thin_border
                cell_rules = ws_weco.cell(row=weco_row, column=5, value=', '.join(v['reasons']))
                cell_rules.border = thin_border
                cell_rules.fill = bad_fill
                weco_row += 1

        if weco_row == 2:
            ws_weco.cell(row=2, column=1, value="無異常檢出").font = Font(name="微軟正黑體", size=11, color="28A745")
            ws_weco.merge_cells('A2:E2')

        for col_idx, w in enumerate([8, 15, 8, 15, 40], 1):
            ws_weco.column_dimensions[get_column_letter(col_idx)].width = w

        # ======== Sheet 5: Embedded Charts ========
        if avgs and len(avgs) >= 2:
            ws_chart = wb.create_sheet(title="圖表")
            data_count = len(avgs)

            # --- X-bar Control Chart ---
            xbar_chart = LineChart()
            xbar_chart.title = f"X̄ 平均值管制圖 - {field}"
            xbar_chart.y_axis.title = "量測值"
            xbar_chart.y_axis.delete = False
            xbar_chart.x_axis.delete = False
            xbar_chart.y_axis.tickLblPos = "low"
            xbar_chart.x_axis.tickLblPos = "low"
            xbar_chart.y_axis.numFmt = '0.000'
            xbar_chart.x_axis.tickLblSkip = 1
            # 縮小圖表尺寸
            xbar_chart.width = 20
            xbar_chart.height = 14
            xbar_chart.style = 10
            # 隱藏圖例（頁面已有圖例說明），避免與軸重疊
            xbar_chart.legend = None

            cats = Reference(ws2, min_col=1, min_row=2, max_row=data_count + 1)
            avg_data = Reference(ws2, min_col=4, min_row=1, max_row=data_count + 1)
            ucl_data = Reference(ws2, min_col=6, min_row=1, max_row=data_count + 1)
            cl_data = Reference(ws2, min_col=7, min_row=1, max_row=data_count + 1)
            lcl_data = Reference(ws2, min_col=8, min_row=1, max_row=data_count + 1)

            xbar_chart.add_data(avg_data, titles_from_data=True)
            xbar_chart.add_data(ucl_data, titles_from_data=True)
            xbar_chart.add_data(cl_data, titles_from_data=True)
            xbar_chart.add_data(lcl_data, titles_from_data=True)

            # Add USL/LSL if available
            if has_spec_limits:
                usl_data = Reference(ws2, min_col=12, min_row=1, max_row=data_count + 1)
                lsl_data = Reference(ws2, min_col=13, min_row=1, max_row=data_count + 1)
                xbar_chart.add_data(usl_data, titles_from_data=True)
                xbar_chart.add_data(lsl_data, titles_from_data=True)

            xbar_chart.set_categories(cats)

            # Style: Avg=blue, UCL/LCL=red dashed, CL=green
            s_avg = xbar_chart.series[0]
            s_avg.graphicalProperties.line.solidFill = "0D6EFD"
            s_avg.graphicalProperties.line.width = 22000

            s_ucl = xbar_chart.series[1]
            s_ucl.graphicalProperties.line.solidFill = "FF0000"
            s_ucl.graphicalProperties.line.dashStyle = "dash"
            s_ucl.graphicalProperties.line.width = 15000

            s_cl = xbar_chart.series[2]
            s_cl.graphicalProperties.line.solidFill = "00B050"
            s_cl.graphicalProperties.line.width = 15000

            s_lcl = xbar_chart.series[3]
            s_lcl.graphicalProperties.line.solidFill = "FF0000"
            s_lcl.graphicalProperties.line.dashStyle = "dash"
            s_lcl.graphicalProperties.line.width = 15000

            if has_spec_limits:
                s_usl = xbar_chart.series[4]
                s_usl.graphicalProperties.line.solidFill = "E83E8C"
                s_usl.graphicalProperties.line.dashStyle = "lgDash"
                s_usl.graphicalProperties.line.width = 18000

                s_lsl = xbar_chart.series[5]
                s_lsl.graphicalProperties.line.solidFill = "E83E8C"
                s_lsl.graphicalProperties.line.dashStyle = "lgDash"
                s_lsl.graphicalProperties.line.width = 18000

            ws_chart.add_chart(xbar_chart, "A1")

            # --- R Chart ---
            r_chart = LineChart()
            r_chart.title = f"{variation_label} 變異管制圖 - {field}"
            r_chart.y_axis.title = variation_label
            r_chart.y_axis.delete = False
            r_chart.x_axis.delete = False
            r_chart.y_axis.tickLblPos = "low"
            r_chart.x_axis.tickLblPos = "low"
            r_chart.y_axis.numFmt = '0.000'
            r_chart.x_axis.tickLblSkip = 1
            # 縮小圖表尺寸
            r_chart.width = 20
            r_chart.height = 14
            r_chart.style = 10
            # 隱藏圖例
            r_chart.legend = None

            r_data = Reference(ws2, min_col=5, min_row=1, max_row=data_count + 1)
            r_ucl_ref = Reference(ws2, min_col=9, min_row=1, max_row=data_count + 1)
            r_cl_ref = Reference(ws2, min_col=10, min_row=1, max_row=data_count + 1)
            r_lcl_ref = Reference(ws2, min_col=11, min_row=1, max_row=data_count + 1)

            r_chart.add_data(r_data, titles_from_data=True)
            r_chart.add_data(r_ucl_ref, titles_from_data=True)
            r_chart.add_data(r_cl_ref, titles_from_data=True)
            r_chart.add_data(r_lcl_ref, titles_from_data=True)
            r_chart.set_categories(cats)

            s_r = r_chart.series[0]
            s_r.graphicalProperties.line.solidFill = "6F42C1"
            s_r.graphicalProperties.line.width = 22000

            s_r_ucl = r_chart.series[1]
            s_r_ucl.graphicalProperties.line.solidFill = "FF0000"
            s_r_ucl.graphicalProperties.line.dashStyle = "dash"
            s_r_ucl.graphicalProperties.line.width = 15000

            s_r_cl = r_chart.series[2]
            s_r_cl.graphicalProperties.line.solidFill = "00B050"
            s_r_cl.graphicalProperties.line.width = 15000

            s_r_lcl = r_chart.series[3]
            s_r_lcl.graphicalProperties.line.solidFill = "FF0000"
            s_r_lcl.graphicalProperties.line.dashStyle = "dash"
            s_r_lcl.graphicalProperties.line.width = 15000

            # 調整 R 圖表位置（從 A50 改為 A30，與 X-bar 圖表並排）
            ws_chart.add_chart(r_chart, "A30")

            # --- Histogram ---
            if hist_data_rows > 0 and ws3 is not None:
                hist_chart = BarChart()
                hist_chart.title = f"量測值分佈直方圖 - {field}"
                hist_chart.y_axis.title = "頻次"
                hist_chart.y_axis.delete = False
                hist_chart.x_axis.delete = False
                hist_chart.y_axis.tickLblPos = "low"
                hist_chart.x_axis.tickLblPos = "low"
                # 縮小圖表尺寸
                hist_chart.width = 20
                hist_chart.height = 14
                hist_chart.style = 10
                hist_chart.type = "col"
                hist_chart.grouping = "clustered"
                # 圖例隱藏（因為只有一個數據系列）
                hist_chart.legend = None

                hist_cats = Reference(ws3, min_col=1, min_row=2, max_row=hist_data_rows + 1)
                hist_vals = Reference(ws3, min_col=2, min_row=1, max_row=hist_data_rows + 1)

                hist_chart.add_data(hist_vals, titles_from_data=True)
                hist_chart.set_categories(hist_cats)

                s_hist = hist_chart.series[0]
                s_hist.graphicalProperties.solidFill = "0D6EFD"

                # 調整直方圖位置（從 A100 改為 A55）
                ws_chart.add_chart(hist_chart, "A55")

        # --- 版本稽核與不可變樣本證據 ---
        if version_metadata:
            ws_audit = wb.create_sheet(title="版本稽核")
            for row_index, (name, value) in enumerate(version_metadata.items(), 1):
                ws_audit.cell(row=row_index, column=1, value=name)
                ws_audit.cell(row=row_index, column=2, value=value)
            ws_audit.column_dimensions['A'].width = 24
            ws_audit.column_dimensions['B'].width = 100

            ws_samples = wb.create_sheet(title="研究樣本")
            sample_rows = stats_data.get('version_samples') or []
            headers = list(sample_rows[0].keys()) if sample_rows else ["無樣本"]
            for column, header in enumerate(headers, 1):
                ws_samples.cell(row=1, column=column, value=header)
            for row_index, sample in enumerate(sample_rows, 2):
                for column, header in enumerate(headers, 1):
                    ws_samples.cell(row=row_index, column=column, value=sample.get(header))

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
