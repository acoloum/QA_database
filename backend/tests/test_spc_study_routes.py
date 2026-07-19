"""SPC 共用研究 API 契約與權限測試。"""

from datetime import date
from io import BytesIO

from sqlalchemy import event as sqlalchemy_event

import pytest
from openpyxl import load_workbook

from backend.models import (
    Role,
    SpcEvent,
    SpcLimitVersion,
    SpcStudy,
    SpcStudyVersion,
    User,
)
from backend.services.spc_contracts import SpcStudyInput, SpcSubgroup
from backend.services.spc_report import SpcReportService
from backend.services.spc_study_service import ADAPTERS
from backend.routes.spc_studies import _serialization_context
from backend.services.spc_study_service import SpcStudyService
from backend.utils import generate_token, hash_password


def _user(db_session, username, role_code):
    user = User(
        username=username, password=hash_password("pw12345678"),
        role=role_code, is_active=True,
    )
    db_session.add(user)
    db_session.commit()
    return user


def _headers(user):
    token = generate_token(user.id, user.username, user.role)
    return {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}


@pytest.fixture
def spc_roles(db_session):
    db_session.add_all([
        Role(code="spc_viewer", name="SPC檢視", permissions={"spc.view": True}),
        Role(
            code="spc_manager", name="SPC管理",
            permissions={"spc.view": True, "spc.manage": True},
        ),
        Role(
            code="spc_approver", name="SPC核准",
            permissions={"spc.view": True, "spc.manage": True, "spc.approve": True},
        ),
        Role(code="qa_supervisor", name="QA主管", permissions={"spc.manage": True}),
        Role(code="qc_manager", name="品管經理", permissions={"spc.manage": True}),
        Role(code="admin", name="系統管理員", permissions={"spc.manage": True}),
        Role(code="inspector", name="檢驗員", permissions={"spc.view": True}),
        Role(code="no_spc", name="無SPC權限", permissions={}),
    ])
    db_session.commit()


def _input(data_hash="a" * 64):
    return SpcStudyInput(
        source="shipping",
        filters={
            "vendor": "", "material": "6061", "spec": "10*1*100",
            "field": "外徑", "start_date": "", "end_date": "",
        },
        process_stream_key="api-stream", characteristic="外徑",
        subgroups=tuple(
            SpcSubgroup(
                key=f"shipping:{index}", timestamp=date(2026, 7, index + 1),
                values=(9.8 + index * 0.01, 10.2 + index * 0.01),
                record_ids=(index + 1,), measurement_ids=(index + 101,),
            )
            for index in range(5)
        ),
        specification={"found": True, "LSL": 9.5, "USL": 10.5},
        data_hash=data_hash,
    )


def test_viewer_can_analyze_and_read_study_contract(
    client, db_session, spc_roles, monkeypatch
):
    viewer = _user(db_session, "spc-viewer", "spc_viewer")
    monkeypatch.setitem(ADAPTERS, "shipping", lambda _filters: _input())

    response = client.post(
        "/api/spc/studies/analyze", headers=_headers(viewer),
        json={"source": "shipping", "filters": {"field": "外徑"}},
    )

    assert response.status_code == 200
    payload = response.get_json()
    assert payload["success"] is True
    version = payload["data"]
    assert version["source"] == "shipping"
    assert version["study_type"] == "retrospective"
    assert version["process_stream_key"] == "api-stream"
    assert version["filters"]["field"] == "外徑"
    assert version["method_version"] == "2026.2"
    assert version["data_hash"] == "a" * 64
    assert version["charts"]["chart_type"] == "xbar_r"
    assert "location" in version["stability"]
    assert "applicable" in version["applicability"]

    listing = client.get("/api/spc/studies", headers=_headers(viewer))
    detail = client.get(
        f"/api/spc/studies/{version['study_id']}", headers=_headers(viewer)
    )
    history = client.get(
        f"/api/spc/studies/{version['study_id']}/history", headers=_headers(viewer)
    )
    assert listing.status_code == detail.status_code == history.status_code == 200
    assert listing.get_json()["data"][0]["id"] == version["study_id"]
    history_data = history.get_json()["data"]
    assert history_data["items"][0]["version_no"] == 1
    assert history_data["page"] == 1
    assert history_data["per_page"] == 20
    assert history_data["total"] == 1


def test_analyze_rejects_non_string_analysis_family_from_api(
    client, db_session, spc_roles
):
    viewer = _user(db_session, "spc-family-api-viewer", "spc_viewer")

    response = client.post(
        "/api/spc/studies/analyze",
        headers=_headers(viewer),
        json={"source": "shipping", "filters": {}, "analysis_family": []},
    )

    assert response.status_code == 400
    assert response.get_json()["code"] == "SPC_ANALYSIS_FAMILY_INVALID"


def test_analyze_requires_spc_view_permission(client, db_session, spc_roles):
    user = _user(db_session, "no-spc", "no_spc")
    response = client.post(
        "/api/spc/studies/analyze", headers=_headers(user),
        json={"source": "shipping", "filters": {}},
    )
    assert response.status_code == 403


def test_machine_analyze_route_requires_spc_manage_permission(client, db_session, spc_roles):
    viewer = _user(db_session, "machine-route-viewer", "spc_viewer")
    response = client.post(
        "/api/spc/studies/analyze", headers=_headers(viewer),
        json={
            "source": "patrol", "analysis_family": "machine",
            "filters": {"m_id": 1, "mat": "6061", "spec": "10x1", "item": "外徑", "pos": "A"},
            "options": {"conditions_confirmed": True, "condition_reason": "已確認固定研究條件"},
        },
    )

    assert response.status_code == 403
    assert response.get_json()["code"] == "SPC_MANAGE_FORBIDDEN"


def test_malformed_machine_request_checks_manage_before_machine_payload_validation(
    client, db_session, spc_roles
):
    viewer = _user(db_session, "malformed-machine-viewer", "spc_viewer")
    manager = _user(db_session, "malformed-machine-manager", "spc_manager")
    payload = {
        "source": "patrol", "analysis_family": "machine",
        "filters": {"m_id": [], "mat": "6061", "spec": "10x1", "item": "外徑", "pos": "A"},
        "options": {"conditions_confirmed": True, "condition_reason": "已確認固定研究條件"},
    }

    forbidden = client.post("/api/spc/studies/analyze", headers=_headers(viewer), json=payload)
    invalid = client.post("/api/spc/studies/analyze", headers=_headers(manager), json=payload)

    assert forbidden.status_code == 403
    assert forbidden.get_json()["code"] == "SPC_MANAGE_FORBIDDEN"
    assert invalid.status_code == 400
    assert invalid.get_json()["code"] == "MACHINE_STREAM_NOT_FIXED"


def test_spc_assignees_require_manage_and_expose_minimal_active_quality_users(
    client, db_session, spc_roles
):
    manager = _user(db_session, "assignee-reader", "spc_manager")
    viewer = _user(db_session, "assignee-viewer", "spc_viewer")
    qa = _user(db_session, "qa-choice", "qa_supervisor")
    qc = _user(db_session, "qc-choice", "qc_manager")
    admin = _user(db_session, "admin-choice", "admin")
    _user(db_session, "inspector-choice", "inspector")
    disabled = _user(db_session, "disabled-choice", "qa_supervisor")
    disabled.is_active = False
    db_session.commit()

    forbidden = client.get("/api/spc/assignees", headers=_headers(viewer))
    response = client.get("/api/spc/assignees", headers=_headers(manager))

    assert forbidden.status_code == 403
    assert response.status_code == 200
    rows = response.get_json()["data"]
    assert {row["id"] for row in rows} == {qa.id, qc.id, admin.id}
    assert set(rows[0]) == {"id", "username", "role", "role_name"}


def test_submit_data_hash_conflict_returns_stable_409_contract(
    client, db_session, spc_roles, monkeypatch
):
    manager = _user(db_session, "spc-manager", "spc_manager")
    values = iter((_input(), _input("b" * 64)))
    monkeypatch.setitem(ADAPTERS, "shipping", lambda _filters: next(values))
    analyzed = client.post(
        "/api/spc/studies/analyze", headers=_headers(manager),
        json={"source": "shipping", "filters": {}},
    ).get_json()["data"]

    response = client.post(
        f"/api/spc/study-versions/{analyzed['id']}/submit",
        headers=_headers(manager), json={"reason": "建立正式基準"},
    )

    assert response.status_code == 409
    assert response.get_json() == {
        "success": False,
        "code": "STUDY_DATA_CHANGED",
        "message": "來源資料、規格或排除狀態已變更，請重新分析後再送審",
        "details": {"saved_hash": "a" * 64, "current_hash": "b" * 64},
    }


def test_manage_permission_cannot_call_approve_endpoint(
    client, db_session, spc_roles
):
    manager = _user(db_session, "manager-no-approve", "spc_manager")
    response = client.post(
        "/api/spc/study-versions/999/approve",
        headers=_headers(manager), json={"reason": "不應通過"},
    )
    assert response.status_code == 403


def test_time_model_route_requires_approve_permission_and_reason(
    client, db_session, spc_roles
):
    manager = _user(db_session, "time-model-manager", "spc_manager")
    approver = _user(db_session, "time-model-approver", "spc_approver")

    forbidden = client.post(
        "/api/spc/study-versions/999/time-model",
        headers=_headers(manager), json={"model": "A1", "reason": "確認"},
    )
    missing_reason = client.post(
        "/api/spc/study-versions/999/time-model",
        headers=_headers(approver), json={"model": "A1", "reason": " "},
    )

    assert forbidden.status_code == 403
    assert forbidden.get_json()["error"] == "權限不足"
    assert missing_reason.status_code == 422
    assert missing_reason.get_json()["code"] == "REASON_REQUIRED"


def test_time_model_route_returns_stable_invalid_override_error(
    client, db_session, spc_roles
):
    approver = _user(db_session, "invalid-time-model", "spc_approver")
    study = SpcStudy(
        source="shipping", study_type="retrospective", analysis_family="variable",
        process_stream_key="invalid-model", characteristic="外徑", filters={},
        created_by=approver.id,
    )
    version = SpcStudyVersion(
        study=study, version_no=1, method_version="2026.2", data_hash="a" * 64,
        analysis_options={}, specification_snapshot={}, chart_result={},
        stability_result={}, distribution_result={}, capability_result={},
        applicability_result={}, status="draft", created_by=approver.id,
        time_model_result={
            "candidate": "B", "confirmed": False,
            "diagnostic_version": "2026.2", "reason_code": None,
        },
    )
    db_session.add(version)
    db_session.commit()

    response = client.post(
        f"/api/spc/study-versions/{version.id}/time-model",
        headers=_headers(approver), json={"model": "C9", "reason": "人工檢查"},
    )

    assert response.status_code == 422
    assert response.get_json()["code"] == "TIME_MODEL_OVERRIDE_INVALID"


def test_transformation_route_requires_approve_permission_and_reason(
    client, db_session, spc_roles
):
    manager = _user(db_session, "transformation-manager", "spc_manager")
    approver = _user(db_session, "transformation-approver", "spc_approver")

    forbidden = client.post(
        "/api/spc/study-versions/999/transformation",
        headers=_headers(manager),
        json={"model": "johnson_su", "reason": "確認"},
    )
    missing_reason = client.post(
        "/api/spc/study-versions/999/transformation",
        headers=_headers(approver),
        json={"model": "johnson_su", "reason": " "},
    )

    assert forbidden.status_code == 403
    assert forbidden.get_json()["error"] == "權限不足"
    assert missing_reason.status_code == 422
    assert missing_reason.get_json()["code"] == "REASON_REQUIRED"


def test_legacy_limit_write_endpoints_are_gone(client, db_session, spc_roles):
    manager = _user(db_session, "legacy-manager", "spc_manager")
    for method in (client.post, client.delete):
        response = method(
            "/api/control-limits", headers=_headers(manager),
            json={} if method == client.post else None,
        )
        assert response.status_code == 410
        assert response.get_json()["code"] == "LEGACY_SPC_LIMITS_READ_ONLY"


def test_study_detail_exposes_limit_event_and_ocap_traceability(
    client, db_session, spc_roles, monkeypatch
):
    manager = _user(db_session, "trace-manager", "spc_manager")
    monkeypatch.setitem(ADAPTERS, "shipping", lambda _filters: _input())
    analyzed = client.post(
        "/api/spc/studies/analyze", headers=_headers(manager),
        json={"source": "shipping", "filters": {}},
    ).get_json()["data"]

    limit = SpcLimitVersion(
        study_version_id=analyzed["id"], process_stream_key="api-stream",
        characteristic="外徑", revision=1, chart_type="xbar_r",
        limits={"location": {}, "variation": {}}, status="active",
        approved_by=manager.id,
    )
    db_session.add(limit)
    db_session.flush()
    event = SpcEvent(
        limit_version_id=limit.id, study_version_id=analyzed["id"],
        chart_kind="variation", rule_code="beyond_limits", point_index=4,
        observed_value=0.8, status="investigating",
    )
    db_session.add(event)
    db_session.flush()
    db_session.commit()

    created = client.post(
        f"/api/spc/events/{event.id}/ocap",
        headers=_headers(manager),
        json={
            "investigation_6m": {"summary": "壓力波動"},
            "owner_id": None,
            "status": "open",
        },
    )
    assert created.status_code == 200
    created_data = created.get_json()["data"]
    assert created_data["event_id"] == event.id
    assert created_data["investigation_6m"] == {"summary": "壓力波動"}
    assert "created_at" in created_data
    assert "updated_at" in created_data

    updated = client.patch(
        f"/api/spc/ocap/{created_data['id']}",
        headers=_headers(manager),
        json={"process_adjustment": "調整壓力", "status": "open"},
    )
    assert updated.status_code == 200
    assert updated.get_json()["data"]["process_adjustment"] == "調整壓力"

    detail = client.get(
        f"/api/spc/studies/{analyzed['study_id']}", headers=_headers(manager)
    ).get_json()["data"]
    saved_limit = detail["versions"][0]["limit_versions"][0]

    assert saved_limit["id"] == limit.id
    assert saved_limit["status"] == "active"
    assert saved_limit["events"][0]["id"] == event.id
    assert saved_limit["events"][0]["ocap"]["investigation_6m"] == {
        "summary": "壓力波動"
    }

    event_response = client.get(
        f"/api/spc/events/{event.id}", headers=_headers(manager)
    )
    assert event_response.status_code == 200
    event_data = event_response.get_json()["data"]
    assert event_data["id"] == event.id
    assert event_data["ocap"]["id"] == created_data["id"]


@pytest.mark.parametrize(
    ("payload", "expected_code"),
    [
        ([], "SPC_OCAP_PAYLOAD_INVALID"),
        ({"status": "unknown"}, "SPC_OCAP_STATUS_INVALID"),
        ({"owner_id": True}, "SPC_OCAP_OWNER_INVALID"),
    ],
)
def test_ocap_route_returns_stable_validation_contract(
    client, db_session, spc_roles, payload, expected_code
):
    manager = _user(db_session, f"payload-{expected_code}", "spc_manager")
    study = SpcStudy(
        source="shipping",
        study_type="ongoing",
        process_stream_key=f"route-{expected_code}",
        characteristic="外徑",
        filters={},
        status="active",
        created_by=manager.id,
    )
    db_session.add(study)
    db_session.flush()
    version = SpcStudyVersion(
        study_id=study.id,
        version_no=1,
        method_version="2026.1",
        status="active",
        created_by=manager.id,
    )
    db_session.add(version)
    db_session.flush()
    limit = SpcLimitVersion(
        study_version_id=version.id,
        process_stream_key=study.process_stream_key,
        characteristic=study.characteristic,
        revision=1,
        chart_type="xbar_s",
        limits={},
        status="active",
        created_by=manager.id,
    )
    db_session.add(limit)
    db_session.flush()
    event = SpcEvent(
        limit_version_id=limit.id,
        study_version_id=version.id,
        chart_kind="location",
        rule_code="beyond_limits",
        point_index=0,
    )
    db_session.add(event)
    db_session.commit()

    response = client.post(
        f"/api/spc/events/{event.id}/ocap",
        headers=_headers(manager),
        json=payload,
    )

    assert response.status_code == 422
    assert response.get_json()["code"] == expected_code


def test_spc_report_requires_view_permission(client, db_session, spc_roles):
    user = _user(db_session, "report-no-spc", "no_spc")

    response = client.get("/api/spc-report", headers=_headers(user))

    assert response.status_code == 403


def test_shipping_report_rejects_other_source_and_uses_only_saved_version(
    client, db_session, spc_roles, monkeypatch
):
    viewer = _user(db_session, "report-viewer", "spc_viewer")
    monkeypatch.setitem(ADAPTERS, "shipping", lambda _filters: _input())
    analyzed = client.post(
        "/api/spc/studies/analyze", headers=_headers(viewer),
        json={"source": "shipping", "filters": {"field": "外徑"}},
    ).get_json()["data"]

    patrol_response = client.get(
        f"/api/patrol/export?item=外徑&study_version_id={analyzed['id']}",
        headers=_headers(viewer),
    )
    shipping_response = client.get(
        "/api/spc-report?field=外徑&material=6061&spec=10*1*100"
        f"&study_version_id={analyzed['id']}",
        headers=_headers(viewer),
    )

    assert patrol_response.status_code == 422
    assert shipping_response.status_code == 200
    workbook = load_workbook(BytesIO(shipping_response.data))
    assert "原始數據" not in workbook.sheetnames
    assert "研究樣本" in workbook.sheetnames


def test_patrol_export_normalizes_full_position_before_creating_version(
    client, db_session, spc_roles, monkeypatch
):
    viewer = _user(db_session, "patrol-export-viewer", "spc_viewer")
    captured = {}

    def analyze(source, filters, actor_id):
        captured["analyze"] = (source, filters, actor_id)
        return type("Version", (), {"id": 91})()

    def generate(version_id, *, expected_source, expected_filters):
        captured["report"] = (version_id, expected_source, expected_filters)
        return BytesIO(b"versioned patrol report")

    monkeypatch.setattr(SpcStudyService, "analyze", staticmethod(analyze))
    monkeypatch.setattr(
        SpcReportService, "generate_version_report", staticmethod(generate)
    )

    response = client.get(
        "/api/patrol/export?item=外徑&position=全段&mat=6061&spec=10*1*100",
        headers=_headers(viewer),
    )

    assert response.status_code == 200
    assert captured["analyze"][0] == "patrol"
    assert captured["analyze"][1]["pos"] == ""
    assert captured["report"] == (91, "patrol", captured["analyze"][1])


def test_event_detail_returns_attribute_immutable_evidence(
    client, db_session, spc_roles
):
    viewer = _user(db_session, "attribute-event-viewer", "spc_viewer")
    baseline_study = SpcStudy(source="shipping", study_type="retrospective", analysis_family="attribute", process_stream_key="attr-event", characteristic="不符合單位", filters={}, created_by=viewer.id)
    baseline = SpcStudyVersion(study=baseline_study, version_no=1, method_version="2026.2", data_hash="a" * 64, analysis_options={}, status="active", created_by=viewer.id)
    ongoing_study = SpcStudy(source="shipping", study_type="ongoing", analysis_family="attribute", process_stream_key="attr-event", characteristic="不符合單位", filters={}, created_by=viewer.id)
    ongoing = SpcStudyVersion(study=ongoing_study, version_no=1, method_version="2026.2", data_hash="b" * 64, analysis_options={}, chart_result={"pearson_residuals": [2.5]}, status="active", created_by=viewer.id)
    db_session.add_all([baseline, ongoing])
    db_session.flush()
    sample = __import__("backend.models", fromlist=["SpcStudySample"]).SpcStudySample(version_id=ongoing.id, source_record_type="ShippingData", source_record_id=1, source_record_ids=[1, 2], source_measurement_ids=[], subgroup_key="attribute:day", subgroup_order=0, values=[2.0, 10.0], distribution_values=[])
    limit = SpcLimitVersion(study_version_id=baseline.id, analysis_family="attribute", process_stream_key="attr-event", characteristic="不符合單位", revision=1, chart_type="p", limits={}, status="active", created_by=viewer.id)
    db_session.add_all([sample, limit])
    db_session.flush()
    event = SpcEvent(limit_version_id=limit.id, study_version_id=ongoing.id, sample_id=sample.id, chart_kind="location", rule_code="beyond_limits", point_index=0, source_point_key="ShippingData:attribute:day", observed_value=0.2)
    db_session.add(event)
    db_session.commit()

    response = client.get(f"/api/spc/events/{event.id}", headers=_headers(viewer))

    assert response.status_code == 200
    evidence = response.get_json()["data"]["attribute_evidence"]
    assert evidence == {"x": 2.0, "n": 10.0, "display_value": 0.2, "pearson_residual": 2.5, "subgroup_key": "attribute:day"}


def test_event_serialization_context_batches_version_and_sample_queries(app, db_session, spc_roles):
    """擴增 event 不可造成逐筆 version/sample lookup。"""
    with app.app_context():
        viewer = _user(db_session, "batch-event-viewer", "spc_viewer")
        base_study = SpcStudy(source="shipping", study_type="retrospective", analysis_family="attribute", process_stream_key="batch", characteristic="不符合單位", filters={}, created_by=viewer.id)
        base = SpcStudyVersion(study=base_study, version_no=1, method_version="2026.2", data_hash="d" * 64, analysis_options={}, status="active", created_by=viewer.id)
        db_session.add(base); db_session.flush()
        limit = SpcLimitVersion(study_version_id=base.id, analysis_family="attribute", process_stream_key="batch", characteristic="不符合單位", revision=1, chart_type="p", limits={}, status="active", created_by=viewer.id)
        db_session.add(limit); db_session.flush()
        versions = []
        for number in range(2):
            study = SpcStudy(source="shipping", study_type="ongoing", analysis_family="attribute", process_stream_key=f"batch-{number}", characteristic="不符合單位", filters={}, created_by=viewer.id)
            version = SpcStudyVersion(study=study, version_no=1, method_version="2026.2", data_hash=str(number) * 64, analysis_options={}, chart_result={"pearson_residuals": [1.0, 2.0]}, status="active", created_by=viewer.id)
            db_session.add(version); db_session.flush(); versions.append(version)
            for point in range(2):
                sample = __import__("backend.models", fromlist=["SpcStudySample"]).SpcStudySample(version_id=version.id, source_record_type="ShippingData", source_record_id=point + 1, source_record_ids=[point + 1], source_measurement_ids=[], subgroup_key=f"g{point}", subgroup_order=point, values=[point + 1, 10], distribution_values=[])
                db_session.add(sample); db_session.flush()
                db_session.add(SpcEvent(limit_version_id=limit.id, study_version_id=version.id, sample_id=sample.id, chart_kind="location", rule_code="beyond_limits", point_index=point, source_point_key=f"{number}-{point}", observed_value=0.1))
        db_session.commit(); db_session.expire_all()
        loaded_limit = SpcLimitVersion.query.options(__import__("sqlalchemy.orm", fromlist=["selectinload"]).selectinload(SpcLimitVersion.events)).filter_by(id=limit.id).first()
        count = [0]
        def listener(_conn, _cursor, statement, *_args):
            if statement.lstrip().upper().startswith("SELECT"): count[0] += 1
        sqlalchemy_event.listen(db_session.get_bind(), "before_cursor_execute", listener)
        try:
            version_map, sample_map = _serialization_context([], {loaded_limit.id: loaded_limit})
        finally:
            sqlalchemy_event.remove(db_session.get_bind(), "before_cursor_execute", listener)
        assert len(version_map) == 2 and len(sample_map) == 4
        assert count[0] <= 4
