
import pytest
from datetime import datetime
from backend.services.tolerance_service import ToleranceService
from backend.models import VendorToleranceMain, VendorToleranceDetail, Vendor

def test_add_tolerance(app, db_session):
    with app.app_context():
        # Setup Vendor
        v = Vendor(name="Test Vendor")
        db_session.add(v)
        db_session.commit()

        data = {
            "材質": "SS400",
            "規格": "100*100",
            "廠商ID": v.id,
            "備註": "Test Note",
            "建立日期": datetime.now(),
            "details": [
                {
                    "測量項目": "OD",
                    "測量位置": "Pos1",
                    "尺寸下限": 9.9,
                    "尺寸上限": 10.1,
                    "公差下限": -0.1,
                    "公差上限": 0.1,
                    "標準值": 10.0,
                    "單位": "mm",
                    "備註": "Detail Note"
                }
            ]
        }
        
        new_id = ToleranceService.add_tolerance(data)
        assert new_id is not None
        
        # Verify DB
        t = db_session.get(VendorToleranceMain, new_id)
        assert t.material == "SS400"
        assert len(t.details) == 1
        assert t.details[0].item == "OD"

def test_check_tolerance_match(app, db_session):
    with app.app_context():
        # Setup Data
        v1 = Vendor(name="Vendor A")
        db_session.add(v1)
        db_session.flush()
        
        # Case 1: Exact Match
        t1 = VendorToleranceMain(material="SUS304", spec="10*10", vendor_id=v1.id)
        db_session.add(t1)
        db_session.flush()

        # 附加一筆明細，並標註特性重要度，驗證 check_tolerance 的序列化會帶出此欄位
        d1 = VendorToleranceDetail(
            main_id=t1.id,
            item="外徑",
            position="",
            dim_min=9.9,
            dim_max=10.1,
            characteristic_class="關鍵"
        )
        db_session.add(d1)

        # Case 2: Partial Spec Match (Wildcard logic handled in code check_tolerance)
        # Service logic: input_spec.startswith(t_spec + '*')

        db_session.commit()

        # Test Exact Match
        args = {"material": "SUS304", "spec": "10*10", "vendor_id": v1.id}
        result = ToleranceService.check_tolerance(args)
        assert result["success"] is True
        assert result["found"] is True
        assert result["tolerance_id"] == t1.id
        assert result["matched_priority"] == 1 # Priority 1: Same Vendor + Same Material + Same Spec
        assert result["tolerances"][0]["特性重要度"] == "關鍵"

def test_check_tolerance_not_found(app, db_session):
    with app.app_context():
        args = {"material": "Unknown", "spec": "10*10"}
        result = ToleranceService.check_tolerance(args)
        assert result["success"] is True
        assert result["found"] is False


def test_search_tolerance_clamps_page_size(app, db_session):
    with app.app_context():
        result = ToleranceService.search_tolerance({"page_size": "5000"})
        assert result["page_size"] == 100


def test_tolerance_parse_spec_values_uses_shared_nominal_parser(app):
    """公差服務的規格解析應與共用解析器保持一致。"""
    with app.app_context():
        assert ToleranceService.parse_spec_values('31.9*2.2*589') == {
            '外徑': 31.9,
            '厚度': 2.2,
            '內徑': 27.5,
            '長度': 589.0,
        }


def test_tolerance_detail_roundtrips_characteristic_class(app, db_session):
    """特性重要度欄位應能透過 add_tolerance 寫入並經 get_tolerance_detail 讀出。"""
    with app.app_context():
        v = Vendor(name="Test Vendor CC")
        db_session.add(v)
        db_session.commit()

        data = {
            "材質": "TEST-CLS",
            "規格": "1*2*3",
            "廠商ID": v.id,
            "備註": "",
            "建立日期": datetime.now(),
            "details": [{
                "測量項目": "外徑", "測量位置": "",
                "尺寸下限": 1.0, "尺寸上限": 2.0,
                "公差下限": None, "公差上限": None,
                "標準值": None, "單位": "mm", "備註": "",
                "特性重要度": "關鍵",
            }],
        }
        new_id = ToleranceService.add_tolerance(data)
        got = ToleranceService.get_tolerance_detail(new_id)
        assert got["details"][0]["特性重要度"] == "關鍵"


# ---------------------------------------------------------------------------
# 匯出：改為串流寫入前後的輸出快照
# ---------------------------------------------------------------------------

EXPORT_COLUMNS = [
    '識別碼', '材質', '規格', '廠商名稱', '測量項目', '測量位置',
    '尺寸下限', '尺寸上限', '公差下限', '公差上限', '標準值', '單位', '備註',
]


def _seed_export_fixture(db_session):
    vendor = Vendor(name='廠商甲')
    db_session.add(vendor)
    db_session.flush()

    # 有兩筆明細：每筆明細各展開成一列
    main_a = VendorToleranceMain(material='6061', spec='10*2', vendor_id=vendor.id)
    db_session.add(main_a)
    db_session.flush()
    db_session.add_all([
        VendorToleranceDetail(
            main_id=main_a.id, item='外徑', position='前段',
            dim_min=9.5, dim_max=10.5, tolerance_min=-0.5, tolerance_max=0.5,
            std_val=10.0, unit='mm', note='備註A',
        ),
        VendorToleranceDetail(
            main_id=main_a.id, item='內徑', position='中段',
            dim_min=5.5, dim_max=6.5, tolerance_min=-0.5, tolerance_max=0.5,
            std_val=6.0, unit='mm', note=None,
        ),
    ])

    # 無明細：仿 LEFT JOIN 產生一列，明細欄位全空
    main_b = VendorToleranceMain(material='6063', spec='20*3', vendor_id=None)
    db_session.add(main_b)
    db_session.commit()
    return main_a, main_b


def _export_rows():
    from openpyxl import load_workbook

    sheet = load_workbook(ToleranceService.export_excel({})).active
    return [list(row) for row in sheet.iter_rows(values_only=True)]


def test_export_excel_expands_details_into_rows(app, db_session):
    """每筆明細展開成一列，無明細的主檔仍輸出一列且明細欄位留空。"""
    with app.app_context():
        main_a, main_b = _seed_export_fixture(db_session)

        rows = _export_rows()

        assert rows[0] == EXPORT_COLUMNS
        assert len(rows) == 4                      # 表頭 + 2 筆明細 + 1 筆無明細

        # 主檔以識別碼遞增排序，明細以識別碼遞增排序
        assert [row[0] for row in rows[1:]] == [main_a.id, main_a.id, main_b.id]
        assert [row[4] for row in rows[1:]] == ['外徑', '內徑', None]

        first = rows[1]
        assert first[1:4] == ['6061', '10*2', '廠商甲']
        assert first[5] == '前段'
        assert first[6] == 9.5      # 尺寸下限
        assert first[7] == 10.5     # 尺寸上限
        assert first[8] == -0.5     # 公差下限
        assert first[10] == 10.0    # 標準值
        assert first[11] == 'mm'
        assert first[12] == '備註A'

        # 明細的 note 為 NULL
        assert rows[2][12] is None

        # 無明細的主檔：廠商為 NULL，明細欄位全空
        last = rows[3]
        assert last[1:4] == ['6063', '20*3', None]
        assert all(value is None for value in last[4:])


def test_export_excel_with_no_data_still_writes_header(app, db_session):
    with app.app_context():
        rows = _export_rows()

        assert rows[0] == EXPORT_COLUMNS
        assert len(rows) == 1
