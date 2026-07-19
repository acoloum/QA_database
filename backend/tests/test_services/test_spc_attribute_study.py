"""屬性型 SPC 研究版本、持續監控與報表的生命週期測試。"""

from datetime import date

import pytest
from openpyxl import load_workbook

import backend.services.spc_study_service as study_module
from backend.models import Role, SpcEvent, SpcLimitVersion, User
from backend.services.spc_attribute_engine import AttributeSubgroup
from backend.services.spc_contracts import SpcStudyInput
from backend.services.spc_report import SpcReportService
from backend.services.spc_errors import SpcValidationError
from backend.services.spc_study_service import (
    ADAPTERS,
    ATTRIBUTE_ADAPTERS,
    SpcStudyService,
    _canonical_attribute_options,
)


def _viewer(db_session):
    db_session.add(Role(
        code="attribute_viewer", name="屬性研究檢視", permissions={"spc.view": True}
    ))
    user = User(username="attribute-viewer", password="hashed", role="attribute_viewer")
    db_session.add(user)
    db_session.flush()
    return user


def _admin(db_session):
    user = User(username="attribute-admin", password="hashed", role="admin")
    db_session.add(user)
    db_session.flush()
    return user


def _attribute_input(
    *, counts=(3, 4, 2, 3, 5), inspected=100, options=None, data_hash="a" * 64
):
    return SpcStudyInput(
        source="shipping",
        filters={"field": "外徑", "material": "6061", "spec": "10*1*100"},
        process_stream_key="attribute-shipping-stream",
        characteristic="不符合單位",
        analysis_family="attribute",
        options=dict(options or {}),
        data_hash=data_hash,
        specification={"found": True, "source": "measurement_and_specification_snapshot"},
        subgroups=tuple(
            AttributeSubgroup(
                key=f"attribute:{index}", timestamp=date(2026, 7, index + 1),
                inspected=inspected, nonconforming=count, record_ids=(index + 1,),
            )
            for index, count in enumerate(counts)
        ),
        metadata={
            "classification_evidence": [{"record_id": index + 1, "eligible": True}
                                        for index in range(len(counts))],
            "classification_snapshot": [],
            "eligible_record_count": len(counts),
            "excluded_record_count": 0,
        },
    )


def _adapter(_filters, *, analysis_family, options, input_contract_version):
    assert analysis_family == "attribute"
    assert input_contract_version == "2026.2"
    return _attribute_input(options=options)


def _attribute_adapter(source, filters, *, options, input_contract_version):
    assert source == "shipping"
    return _attribute_input(options=options)


def test_attribute_options_are_canonical_and_reject_invalid_values():
    assert _canonical_attribute_options(None) == {
        "interval": "day", "chart_type": "p", "alpha": 0.0027,
    }
    assert _canonical_attribute_options({"chart_type": "np", "alpha": 0.01}) == {
        "interval": "day", "chart_type": "np", "alpha": 0.01,
    }
    for options in (
        {"alpha": 0}, {"alpha": True}, {"alpha": "0.01"},
        {"alpha": float("nan")}, {"alpha": float("inf")}, {"unknown": 1},
    ):
        with pytest.raises(Exception):
            _canonical_attribute_options(options)


def test_attribute_analysis_persists_family_counts_and_exact_limits(
    app, db_session, monkeypatch
):
    with app.app_context():
        actor = _viewer(db_session)
        monkeypatch.setitem(ATTRIBUTE_ADAPTERS, "shipping", _attribute_adapter)

        version = SpcStudyService.analyze(
            "shipping", {"field": "外徑"}, actor.id,
            analysis_family="attribute",
            options={"interval": "day", "chart_type": "p"},
        )

        assert version.study.analysis_family == "attribute"
        assert version.method_version == "2026.2"
        assert version.chart_result["chart_type"] == "p"
        assert version.chart_result["x"] == [3, 4, 2, 3, 5]
        assert version.chart_result["n"] == [100] * 5
        assert version.chart_result["exact_alpha"] == pytest.approx(0.0027)
        assert version.chart_result["interval"] == "day"
        assert version.chart_result["eligibility_evidence"][0]["eligible"] is True
        assert version.chart_result["exclusion_evidence"] == []
        assert version.stability_result["residual_method"] == "binomial_pearson"
        assert version.analysis_options == {
            "interval": "day", "chart_type": "p", "alpha": 0.0027,
        }
        assert version.samples[0].values == [3.0, 100.0]


def test_attribute_ongoing_uses_approved_limits_without_recentering(
    app, db_session, monkeypatch
):
    with app.app_context():
        actor = _viewer(db_session)
        monkeypatch.setitem(ATTRIBUTE_ADAPTERS, "shipping", _attribute_adapter)
        baseline = SpcStudyService.analyze(
            "shipping", {"field": "外徑"}, actor.id,
            analysis_family="attribute",
            options={"interval": "day", "chart_type": "p"},
        )
        limit = SpcLimitVersion(
            study_version_id=baseline.id,
            analysis_family="attribute",
            process_stream_key=baseline.study.process_stream_key,
            characteristic=baseline.study.characteristic,
            revision=1,
            chart_type="p",
            limits={
                "center": baseline.chart_result["center"],
                "alpha": baseline.chart_result["exact_alpha"],
                "interval": "day",
                "baseline_n": 100,
                "rules_used": ["beyond_limits", "run_9_same_side", "trend_6"],
            },
            status="active", created_by=actor.id, approved_by=actor.id,
        )
        db_session.add(limit)
        db_session.commit()
        monkeypatch.setitem(
            ATTRIBUTE_ADAPTERS, "shipping",
            lambda _source, _filters, *, options, input_contract_version:
                _attribute_input(counts=(20, 7, 9, 8, 10), options=options, data_hash="b" * 64),
        )

        ongoing = SpcStudyService.analyze(
            "shipping", {"field": "外徑"}, actor.id, study_type="ongoing",
            analysis_family="attribute",
            options={"interval": "day", "chart_type": "p"},
        )

        assert ongoing.chart_result["center"] == baseline.chart_result["center"]
        assert ongoing.chart_result["x"] == [20, 7, 9, 8, 10]
        assert ongoing.stability_result["residual_method"] == "binomial_pearson"
        assert ongoing.time_model_result["limit_version_id"] == limit.id
        workbook = load_workbook(SpcReportService.generate_version_report(ongoing.id))
        audit = {
            workbook["版本稽核"].cell(row=row, column=1).value:
            workbook["版本稽核"].cell(row=row, column=2).value
            for row in range(1, workbook["版本稽核"].max_row + 1)
        }
        assert audit["監控所用界限版本 ID"] == limit.id
        assert audit["核准者ID"] == actor.id
        assert audit["界限已凍結"] == "是"
        event = SpcEvent.query.filter_by(study_version_id=ongoing.id).first()
        assert event.sample_id is not None
        assert float(event.observed_value) == pytest.approx(0.2)
        sample = next(sample for sample in ongoing.samples if sample.id == event.sample_id)
        assert sample.values == [20.0, 100.0]
        assert ongoing.chart_result["pearson_residuals"][event.point_index] > 0


def test_attribute_ongoing_inherits_frozen_options_and_rejects_explicit_drift(
    app, db_session, monkeypatch
):
    with app.app_context():
        actor = _viewer(db_session)
        monkeypatch.setitem(ATTRIBUTE_ADAPTERS, "shipping", _attribute_adapter)
        frozen_options = {"interval": "week", "chart_type": "p", "alpha": 0.01}
        baseline = SpcStudyService.analyze(
            "shipping", {"field": "外徑"}, actor.id,
            analysis_family="attribute", options=frozen_options,
        )
        limit = SpcLimitVersion(
            study_version_id=baseline.id, analysis_family="attribute",
            process_stream_key=baseline.study.process_stream_key,
            characteristic=baseline.study.characteristic, revision=1, chart_type="p",
            limits={
                "center": baseline.chart_result["center"], "alpha": 0.01,
                "interval": "week", "options": frozen_options,
                "rules_used": ["beyond_limits"],
            },
            status="active", created_by=actor.id, approved_by=actor.id,
        )
        db_session.add(limit)
        db_session.commit()

        inherited = SpcStudyService.analyze(
            "shipping", {"field": "外徑"}, actor.id,
            study_type="ongoing", analysis_family="attribute",
        )
        assert inherited.analysis_options == frozen_options
        assert inherited.chart_result["exact_alpha"] == 0.01

        with pytest.raises(SpcValidationError) as error:
            SpcStudyService.analyze(
                "shipping", {"field": "外徑"}, actor.id,
                study_type="ongoing", analysis_family="attribute",
                options={"interval": "day", "chart_type": "p", "alpha": 0.01},
            )
        assert error.value.code == "SPC_ATTRIBUTE_OPTIONS_MISMATCH"


def test_attribute_ongoing_np_rejects_subgroup_size_mismatch(
    app, db_session, monkeypatch
):
    with app.app_context():
        actor = _viewer(db_session)
        monkeypatch.setitem(ATTRIBUTE_ADAPTERS, "shipping", _attribute_adapter)
        options = {"interval": "day", "chart_type": "np", "alpha": 0.0027}
        baseline = SpcStudyService.analyze(
            "shipping", {"field": "外徑"}, actor.id,
            analysis_family="attribute", options=options,
        )
        limit = SpcLimitVersion(
            study_version_id=baseline.id, analysis_family="attribute",
            process_stream_key=baseline.study.process_stream_key,
            characteristic=baseline.study.characteristic, revision=1, chart_type="np",
            limits={"center": baseline.chart_result["center"], "alpha": 0.0027,
                    "interval": "day", "baseline_n": 100, "options": options,
                    "rules_used": ["beyond_limits"]},
            status="active", created_by=actor.id, approved_by=actor.id,
        )
        db_session.add(limit)
        db_session.commit()
        monkeypatch.setitem(
            ATTRIBUTE_ADAPTERS, "shipping",
            lambda _source, _filters, *, options, input_contract_version:
                _attribute_input(inspected=101, options=options, data_hash="n" * 64),
        )

        with pytest.raises(SpcValidationError) as error:
            SpcStudyService.analyze(
                "shipping", {"field": "外徑"}, actor.id,
                study_type="ongoing", analysis_family="attribute", options=options,
            )
        assert error.value.code == "NP_SUBGROUP_SIZE_MISMATCH"


def test_attribute_approval_freezes_canonical_options(app, db_session, monkeypatch):
    with app.app_context():
        actor = _admin(db_session)
        monkeypatch.setitem(ATTRIBUTE_ADAPTERS, "shipping", _attribute_adapter)
        options = {"interval": "month", "chart_type": "p", "alpha": 0.01}
        version = SpcStudyService.analyze(
            "shipping", {"field": "外徑"}, actor.id,
            analysis_family="attribute", options=options,
        )
        submitted = SpcStudyService.submit(version.id, actor.id, reason="送審屬性基準")
        limit = SpcStudyService.approve_and_activate(
            submitted.id, actor.id, reason="核准屬性基準")

        assert limit.analysis_family == "attribute"
        assert limit.limits["options"] == options
        assert limit.limits["alpha"] == 0.01


def test_attribute_report_includes_immutable_counts_and_method_metadata(
    app, db_session, monkeypatch
):
    with app.app_context():
        actor = _viewer(db_session)
        monkeypatch.setitem(ATTRIBUTE_ADAPTERS, "shipping", _attribute_adapter)
        version = SpcStudyService.analyze(
            "shipping", {"field": "外徑"}, actor.id,
            analysis_family="attribute",
            options={"interval": "week", "chart_type": "p"},
        )

        workbook = load_workbook(SpcReportService.generate_version_report(version.id))
        audit = {
            workbook["版本稽核"].cell(row=row, column=1).value:
            workbook["版本稽核"].cell(row=row, column=2).value
            for row in range(1, workbook["版本稽核"].max_row + 1)
        }
        headers = [cell.value for cell in workbook["屬性管制圖數據"][1]]
        content = "\n".join(
            str(cell.value) for sheet in workbook.worksheets
            for row in sheet.iter_rows() for cell in row if cell.value is not None
        )

        assert audit["研究版本ID"] == version.id
        assert audit["資料雜湊"] == version.data_hash
        assert audit["方法版本"] == "2026.2"
        assert audit["分析族別"] == "attribute"
        assert audit["時間區間"] == "week"
        assert "受檢數(n)" in headers
        assert "不符合數(x)" in headers
        assert "X̄" not in content
        assert "Pp" not in content
        assert "G法" not in content
