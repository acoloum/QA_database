"""機械性質 Excel 匯入解析工具測試。"""

import openpyxl
import pytest

from backend.services.mechanical_import import (
    build_payloads_from_workbook,
    normalize_product_size,
)


@pytest.mark.parametrize(("raw", "expected"), [
    ("80X70", "80*70"),
    ("66.7X59", "66.7*59"),
    ("44.5x36.7", "44.5*36.7"),
    ("90X80.2", "90*80.2"),
    ("36x25.2", "36*25.2"),
    ("54.5×46.2", "54.5*46.2"),
    # 已是 * 或無分隔者維持原樣
    ("82.5*70.15", "82.5*70.15"),
    ("65.5", "65.5"),
    ("", ""),
])
def test_normalize_product_size_unifies_separator(raw, expected):
    assert normalize_product_size(raw) == expected


def _workbook(*sheet_names: str):
    """建立最小可解析的轉置表：A 欄列標籤，B 欄一筆有資料的檢驗。"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name in sheet_names:
        ws = wb.create_sheet(name)
        ws.cell(row=1, column=1, value="擠製日期/批號")
        ws.cell(row=1, column=2, value="A123")
        ws.cell(row=2, column=1, value="T4溫度/時間")
        ws.cell(row=3, column=1, value="測試日期")
        ws.cell(row=3, column=2, value="2026-01-05")
        ws.cell(row=4, column=1, value="抗拉強度1-1(爐門)")
        ws.cell(row=4, column=2, value=385)
    return wb


def _resolver(mapping):
    """以 {產品尺寸: 反查結果} 模擬 resolve_material_by_spec。"""
    def resolve(product_size):
        return mapping.get(
            product_size,
            {"status": "not_found", "material": None, "candidates": [], "match": None},
        )
    return resolve


def test_build_payloads_fills_material_per_sheet_from_tolerance():
    """不同工作表（產品尺寸）各自套用公差檔反查到的材質，而非整批同一材質。"""
    wb = _workbook("80X70", "66.7X59")
    plan = build_payloads_from_workbook(wb, None, vendor_id=2, resolve_material=_resolver({
        "80*70": {"status": "resolved", "material": "6061-T651", "candidates": ["6061-T651"], "match": "exact"},
        "66.7*59": {"status": "resolved", "material": "7050-T76", "candidates": ["7050-T76"], "match": "exact"},
    }))

    materials = {sheet: payload["材質"] for sheet, _col, payload in plan.payloads}
    assert materials == {"80X70": "6061-T651", "66.7X59": "7050-T76"}
    assert plan.errors == []
    assert {r["產品尺寸"]: r["來源"] for r in plan.resolutions} == {
        "80*70": "公差檔", "66.7*59": "公差檔",
    }


def test_build_payloads_falls_back_to_default_material_when_not_found():
    """公差檔查無規格時退回使用者指定的預設材質，並標示來源為預設值。"""
    wb = _workbook("99X99")
    plan = build_payloads_from_workbook(wb, "6061-T651", vendor_id=2, resolve_material=_resolver({}))

    assert [payload["材質"] for _s, _c, payload in plan.payloads] == ["6061-T651"]
    assert plan.errors == []
    assert plan.resolutions[0]["來源"] == "預設值"


def test_build_payloads_errors_when_not_found_and_no_default():
    """查無規格又沒有預設材質時，該工作表不得匯入（材質為必填）。"""
    wb = _workbook("99X99")
    plan = build_payloads_from_workbook(wb, None, vendor_id=2, resolve_material=_resolver({}))

    assert plan.payloads == []
    assert len(plan.errors) == 1
    assert plan.errors[0]["工作表"] == "99X99"
    assert plan.errors[0]["欄位"] is None
    assert "預設材質" in plan.errors[0]["錯誤"]


def test_build_payloads_errors_and_lists_candidates_when_ambiguous():
    """同規格對應多材質時整表不匯入，錯誤訊息須列出候選材質。"""
    wb = _workbook("28X3.4")
    plan = build_payloads_from_workbook(wb, "6061-T651", vendor_id=1, resolve_material=_resolver({
        "28*3.4": {
            "status": "ambiguous", "material": None,
            "candidates": ["2014-T6", "6066-T6"], "match": "exact",
        },
    }))

    assert plan.payloads == []
    assert len(plan.errors) == 1
    assert "2014-T6" in plan.errors[0]["錯誤"]
    assert "6066-T6" in plan.errors[0]["錯誤"]


def test_build_payloads_ambiguous_does_not_silently_use_default_material():
    """預設材質只用於「查無」，不得掩蓋多材質的歧義。"""
    wb = _workbook("28X3.4")
    plan = build_payloads_from_workbook(wb, "2014-T6", vendor_id=1, resolve_material=_resolver({
        "28*3.4": {
            "status": "ambiguous", "material": None,
            "candidates": ["2014-T6", "6066-T6"], "match": "exact",
        },
    }))

    assert plan.payloads == []
    assert plan.errors


def test_build_payloads_without_resolver_uses_default_material():
    """未提供 resolver（例如未指定廠商）時，沿用舊行為整批套用預設材質。"""
    wb = _workbook("80X70")
    plan = build_payloads_from_workbook(wb, "6061-T651", vendor_id=None)

    assert [payload["材質"] for _s, _c, payload in plan.payloads] == ["6061-T651"]
    assert plan.errors == []


def test_build_payloads_skips_resolution_for_empty_sheets():
    """完全沒有資料的工作表不應產生材質判定紀錄或錯誤。"""
    wb = _workbook("80X70")
    empty = wb.create_sheet("空表")
    empty.cell(row=1, column=1, value="擠製日期/批號")
    empty.cell(row=2, column=1, value="測試日期")
    empty.cell(row=1, column=2, value=None)

    plan = build_payloads_from_workbook(wb, None, vendor_id=2, resolve_material=_resolver({
        "80*70": {"status": "resolved", "material": "6061-T651", "candidates": ["6061-T651"], "match": "exact"},
    }))

    assert [r["工作表"] for r in plan.resolutions] == ["80X70"]
    assert plan.errors == []


def _trace_sheet(*rows: tuple[str, object]):
    """依 (A欄標籤, B欄值) 順序建立單一工作表工作簿；標籤為 None 表示未標籤列。"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("80X70")
    for index, (label, value) in enumerate(rows, start=1):
        if label is not None:
            ws.cell(row=index, column=1, value=label)
        if value is not None:
            ws.cell(row=index, column=2, value=value)
    return wb


def _only_payload(wb):
    plan = build_payloads_from_workbook(wb, "6061-T651", vendor_id=None)
    assert len(plan.payloads) == 1
    return plan.payloads[0][2]


def test_unlabeled_row_above_t4_furnace_label_is_extrusion_not_furnace():
    """「T4爐具編號」標籤之前的未標籤列屬第二個擠製編號，不得落入 T4 爐號。"""
    payload = _only_payload(_trace_sheet(
        ("擠製日期/批號", "042451 D11"),
        (None, "052571 D50"),
        ("T4爐具編號", "051103T42"),
        ("T4溫度/時間", "530/6H"),
        ("測試日期", "2026-01-05"),
    ))

    assert payload["extrusion_numbers"] == [
        {"序號": 1, "編號": "042451 D11"},
        {"序號": 2, "編號": "052571 D50"},
    ]
    assert payload["t4_furnace_numbers"] == [{"序號": 1, "編號": "051103T42"}]


def test_unlabeled_row_below_t4_furnace_label_is_still_furnace():
    """「T4爐具編號」標籤之後的未標籤列仍是 T4 爐號（同欄可有多個爐號）。"""
    payload = _only_payload(_trace_sheet(
        ("擠製日期/批號", "042451 D11"),
        ("T4爐具編號", "051103T42"),
        (None, "051104T42"),
        ("T4溫度/時間", "530/6H"),
        ("測試日期", "2026-01-05"),
    ))

    assert payload["extrusion_numbers"] == [{"序號": 1, "編號": "042451 D11"}]
    assert payload["t4_furnace_numbers"] == [
        {"序號": 1, "編號": "051103T42"},
        {"序號": 2, "編號": "051104T42"},
    ]


def test_extrusion_label_row_empty_still_reads_unlabeled_extrusion_row():
    """擠製標籤列該欄空白、值落在下一列未標籤列時，序號仍從 1 起算。"""
    payload = _only_payload(_trace_sheet(
        ("擠製日期/批號", None),
        (None, "052571 D50"),
        ("T4爐具編號", "051103T42"),
        ("T4溫度/時間", None),
        ("測試日期", "2026-01-05"),
    ))

    assert payload["extrusion_numbers"] == [{"序號": 1, "編號": "052571 D50"}]
    assert payload["t4_furnace_numbers"] == [{"序號": 1, "編號": "051103T42"}]


def test_duplicate_trace_values_across_rows_are_deduplicated():
    """同欄重複填寫的追溯編號只保留一筆，序號才能維持從 1 連續。"""
    payload = _only_payload(_trace_sheet(
        ("擠製日期/批號", "042451 D11"),
        (None, "042451 D11"),
        ("T4爐具編號", "051103T42"),
        (None, "051103T42"),
        ("T4溫度/時間", None),
        ("測試日期", "2026-01-05"),
    ))

    assert payload["extrusion_numbers"] == [{"序號": 1, "編號": "042451 D11"}]
    assert payload["t4_furnace_numbers"] == [{"序號": 1, "編號": "051103T42"}]


def test_missing_t4_furnace_label_keeps_rows_between_as_furnace():
    """沒有「T4爐具編號」標籤時無從切分，維持舊行為避免漏收爐號。"""
    payload = _only_payload(_trace_sheet(
        ("擠製日期/批號", "042451 D11"),
        (None, "051103T42"),
        ("T4溫度/時間", None),
        ("測試日期", "2026-01-05"),
    ))

    assert payload["extrusion_numbers"] == [{"序號": 1, "編號": "042451 D11"}]
    assert payload["t4_furnace_numbers"] == [{"序號": 1, "編號": "051103T42"}]
