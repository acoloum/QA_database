from backend.services.shipping_export import (
    build_shipping_export_columns,
    build_shipping_export_row,
)


def test_build_shipping_export_columns_keeps_measurement_order():
    columns = build_shipping_export_columns(max_groups=1)

    assert columns == [
        "識別碼", "檢驗日期", "材質", "檢驗規格", "訂單號碼", "檢驗人員", "廠商名稱", "組數",
        "外徑1-最小", "外徑1-最大", "內徑1-最小", "內徑1-最大", "真圓度1",
        "厚度1-最小", "厚度1-最大", "同心度1", "長度1", "硬度1", "真直度1",
    ]


def test_build_shipping_export_row_maps_nested_measurements_and_empty_values():
    row = {
        "識別碼": 7,
        "檢驗日期": "2026-06-27",
        "材質": "6061",
        "檢驗規格": "10*2",
        "訂單號碼": "SO-1",
        "檢驗人員": "檢驗員A",
        "廠商中文名稱": "廠商A",
        "組數": 1,
        "measurements": {
            "1": {
                "外徑": {"value_min": 9.8, "value_max": 10.2},
                "硬度": {"value_single": None},
            },
        },
    }

    export_row = build_shipping_export_row(row, max_groups=1)

    assert export_row["外徑1-最小"] == 9.8
    assert export_row["外徑1-最大"] == 10.2
    assert export_row["硬度1"] == ""
    assert export_row["廠商名稱"] == "廠商A"


def test_build_shipping_export_row_segmented_od_uses_extremes():
    """分段紀錄的外徑欄位取三段整體極值,欄位格式不變"""
    row = {
        "識別碼": 8,
        "檢驗日期": "2026-07-09",
        "材質": "6061",
        "檢驗規格": "10*2",
        "訂單號碼": "SO-2",
        "檢驗人員": "檢驗員A",
        "廠商中文名稱": "廠商A",
        "組數": 1,
        "measurements": {
            "1": {
                "外徑@前段": {"value_min": 9.8, "value_max": 10.1},
                "外徑@中段": {"value_min": 9.9, "value_max": 10.2},
                "外徑@後段": {"value_min": 9.7, "value_max": 10.0},
            },
        },
    }

    export_row = build_shipping_export_row(row, max_groups=1)

    assert export_row["外徑1-最小"] == 9.7
    assert export_row["外徑1-最大"] == 10.2
    # 無資料的欄位維持空字串
    assert export_row["內徑1-最小"] == ""


def test_export_excel_streams_rows_without_pandas(app, db_session):
    """匯出改為逐批串流寫入後，表頭與資料列仍需與欄位定義一一對應。"""
    from datetime import date

    from openpyxl import load_workbook

    from backend.models import ShippingData, ShippingMeasurement
    from backend.services.shipping_service import ShippingService

    for index in range(3):
        record = ShippingData(
            date=date(2026, 7, index + 1), material="6061", spec="10*1*100",
            order_num=f"SO-{index + 1}", group_count=1,
        )
        db_session.add(record)
        db_session.flush()
        db_session.add(ShippingMeasurement(
            shipping_id=record.id, group_num=1, item="外徑", position="",
            value_min=9.9 + index * 0.01, value_max=10.1 + index * 0.01,
            lower_limit=9.5, upper_limit=10.5,
        ))
    db_session.commit()

    output = ShippingService.export_excel({})

    workbook = load_workbook(output)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    columns = build_shipping_export_columns()

    assert list(rows[0]) == columns
    assert len(rows) == 4                       # 表頭 + 3 筆
    order_index = columns.index("訂單號碼")
    assert [row[order_index] for row in rows[1:]] == ["SO-1", "SO-2", "SO-3"]
    od_min_index = columns.index("外徑1-最小")
    assert rows[1][od_min_index] == 9.9


def test_export_excel_with_no_data_still_writes_header(app, db_session):
    from openpyxl import load_workbook

    from backend.services.shipping_service import ShippingService

    output = ShippingService.export_excel({})

    sheet = load_workbook(output).active
    rows = list(sheet.iter_rows(values_only=True))
    assert list(rows[0]) == build_shipping_export_columns()
    assert len(rows) == 1


def test_spc_source_size_guard_rejects_oversized_range(app, db_session, monkeypatch):
    """SPC 來源筆數超過上限時要明確擋下並回可讀原因，而不是默默把整張表載進記憶體。"""
    from datetime import date

    import pytest

    from backend.models import ShippingData
    from backend.services.spc_adapters import common as spc_common
    from backend.services.spc_errors import SpcValidationError
    from backend.services.shipping_service import ShippingService

    monkeypatch.setattr(spc_common, "MAX_SOURCE_RECORDS", 2)
    for index in range(3):
        db_session.add(ShippingData(
            date=date(2026, 7, index + 1), material="6061", spec="10*1*100",
            order_num=f"SO-{index + 1}", group_count=1,
        ))
    db_session.commit()

    with pytest.raises(SpcValidationError) as excinfo:
        ShippingService.get_stats({"field": "外徑", "material": "6061"})

    assert excinfo.value.code == "SPC_SOURCE_TOO_LARGE"
    assert excinfo.value.status_code == 422
    assert excinfo.value.details["record_count"] == 3


def test_spc_service_error_maps_to_contract_response(app, client, db_session):
    """SPC 錯誤經由全域處理器要回 422 與代碼，而不是 500 內部錯誤。"""
    from datetime import date

    from backend.models import Role, ShippingData, User
    from backend.services.spc_adapters import common as spc_common
    from backend.utils import generate_token

    original_max = spc_common.MAX_SOURCE_RECORDS
    spc_common.MAX_SOURCE_RECORDS = 1
    try:
        db_session.add(Role(code='qa', name='品保', permissions={'shipping.view': True}))
        user = User(username='spc_guard_user', password='pw', role='qa')
        db_session.add(user)
        for index in range(2):
            db_session.add(ShippingData(
                date=date(2026, 7, index + 1), material="6061", spec="10*1*100",
                order_num=f"SO-G{index}", group_count=1,
            ))
        db_session.commit()
        token = generate_token(user.id, user.username, user.role, user.token_version)

        response = client.get(
            '/api/stats?field=外徑&material=6061',
            headers={'Authorization': f'Bearer {token}'},
        )

        assert response.status_code == 422
        body = response.get_json()
        assert body['success'] is False
        assert body['error']['code'] == 'SPC_SOURCE_TOO_LARGE'
        assert '縮小' in body['error']['message']
        assert body['error']['details']['max_records'] == 1
    finally:
        spc_common.MAX_SOURCE_RECORDS = original_max
