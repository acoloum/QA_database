"""SPC 報表由不可變研究版本重建的回歸測試。"""

from datetime import date

from openpyxl import load_workbook

from backend.models import (
    Role, ShippingData, ShippingMeasurement, SpcStudy, SpcStudySample,
    SpcStudyVersion, User,
)
from backend.services.spc_report import SpcReportService
from backend.services.spc_study_service import SpcStudyService


def _sheet_values(workbook, sheet_name):
    sheet = workbook[sheet_name]
    return [tuple(cell.value for cell in row) for row in sheet.iter_rows()]


def test_report_rebuilds_from_saved_version_after_source_changes(app, db_session):
    with app.app_context():
        db_session.add(Role(
            code="spc_reporter", name="SPC報表", permissions={"spc.view": True}
        ))
        actor = User(
            username="spc-reporter", password="hashed", role="spc_reporter"
        )
        db_session.add(actor)
        for index in range(5):
            record = ShippingData(
                date=date(2026, 7, index + 1), material="6061", spec="10*1*100",
                order_num=f"SO-R-{index + 1}", group_count=3,
            )
            db_session.add(record)
            db_session.flush()
            for group in range(1, 4):
                center = 10 + index * 0.01 + group * 0.01
                db_session.add(ShippingMeasurement(
                    shipping_id=record.id, group_num=group, item="外徑", position="",
                    value_min=center - 0.08, value_max=center + 0.08,
                    lower_limit=9.5, upper_limit=10.5,
                ))
        db_session.commit()

        version = SpcStudyService.analyze(
            "shipping",
            {"field": "外徑", "material": "6061", "spec": "10*1*100"},
            actor.id,
        )
        before = load_workbook(SpcReportService.generate_version_report(version.id))

        changed = ShippingMeasurement.query.order_by(ShippingMeasurement.id).first()
        changed.value_min = 99
        changed.value_max = 100
        db_session.commit()
        after = load_workbook(SpcReportService.generate_version_report(version.id))

        assert _sheet_values(before, "管制圖數據") == _sheet_values(after, "管制圖數據")
        assert _sheet_values(before, "研究樣本") == _sheet_values(after, "研究樣本")
        assert _sheet_values(before, "版本稽核") == _sheet_values(after, "版本稽核")

        audit = dict(
            (before["版本稽核"].cell(row=row, column=1).value,
             before["版本稽核"].cell(row=row, column=2).value)
            for row in range(1, before["版本稽核"].max_row + 1)
        )
        assert audit["研究版本ID"] == version.id
        assert audit["資料雜湊"] == version.data_hash
        assert audit["方法版本"] == "2026.2"
        assert audit["程式版本"] == "2026.1"
        assert audit["圖表選型"] == "xbar_s"
        assert audit["完整篩選條件"]
        assert audit["規格快照"]
        assert audit["每組樣本數"] == "3, 3, 3, 3, 3"

        chart_sheet = before["管制圖數據"]
        headers = [cell.value for cell in chart_sheet[1]]
        assert "S_LCL" in headers
        lcl_column = headers.index("S_LCL") + 1
        expected_lcls = version.chart_result["variation"]["lcl"]
        assert [
            chart_sheet.cell(row=index + 2, column=lcl_column).value
            for index in range(len(expected_lcls))
        ] == [round(value, 4) for value in expected_lcls]


def test_report_never_recalculates_missing_saved_stability():
    output = SpcReportService.generate_report({
        "labels": ["A", "B", "C", "D", "E"],
        "dates": ["2026-07-01"] * 5,
        "avgs": [10.0, 10.1, 9.9, 10.0, 10.1],
        "ranges": [0.2, 0.3, 0.2, 0.25, 0.2],
        "all_values": [],
        "x_cl": 10.0, "x_ucl": 10.5, "x_lcl": 9.5,
        "r_cl": 0.2, "r_ucl": 0.5, "r_lcl": 0.05,
        "stability": {},
        "process_capability": {"available": False},
    }, "外徑", {})

    assert load_workbook(output)["WECO 異常清單"]["A2"].value == "無異常檢出"


def test_machine_report_uses_research_semantics_without_control_limits(app, db_session):
    with app.app_context():
        study = SpcStudy(
            source="patrol", study_type="retrospective", analysis_family="machine",
            process_stream_key="machine-report", characteristic="外徑",
            filters={"m_id": 1, "mat": "6061", "spec": "10*1*100", "item": "外徑", "pos": "前段"},
            status="approved",
        )
        db_session.add(study)
        db_session.flush()
        version = SpcStudyVersion(
            study_id=study.id, version_no=1, method_version="2026.2", data_hash="m" * 64,
            analysis_options={"conditions_confirmed": True, "condition_reason": "量具與機台設定受控"},
            specification_snapshot={"found": True, "consistent": True, "LSL": 9.0, "USL": 11.0},
            distribution_result={"accepted": True, "model": "normal"},
            capability_result={
                "index_family": "machine", "n": 50, "pm": 2.0, "pmk": 1.67,
                "pmu": 1.7, "pml": 1.67, "targets": {"pm": 2.0, "pmk": 1.67},
                "quantiles": {"q_0_135": 9.5, "q_50": 10.0, "q_99_865": 10.5},
                "evidence": {"source_semantics": "patrol_min_max_observations"},
            }, status="approved",
        )
        db_session.add(version)
        db_session.flush()
        db_session.add(SpcStudySample(
            version_id=version.id, source_record_type="PatrolDetail", source_record_id=1,
            source_record_ids=[1], source_measurement_ids=[101], subgroup_key="patrol:1",
            subgroup_order=0, values=[9.9, 10.1], distribution_values=[9.9, 10.1],
        ))
        db_session.commit()

        workbook = load_workbook(SpcReportService.generate_version_report(version.id))

        assert "機器績效摘要" in workbook.sheetnames
        assert "版本稽核" in workbook.sheetnames
        assert "管制圖數據" not in workbook.sheetnames
        rows = _sheet_values(workbook, "機器績效摘要")
        assert ("資料來源語意", "patrol_min_max_observations") in rows
        assert ("Pm", 2.0) in rows
        audit = dict(_sheet_values(workbook, "版本稽核"))
        assert audit["分析族別"] == "machine"
        assert audit["方法版本"] == "2026.2"
        assert audit["分布證據"]


def test_machine_report_marks_missing_source_semantics_as_not_auditable(app, db_session):
    with app.app_context():
        study = SpcStudy(
            source="patrol", study_type="retrospective", analysis_family="machine",
            process_stream_key="machine-missing-evidence", characteristic="外徑", filters={}, status="draft",
        )
        db_session.add(study)
        db_session.flush()
        version = SpcStudyVersion(
            study_id=study.id, version_no=1, method_version="2026.2", data_hash="z" * 64,
            analysis_options={"conditions_confirmed": True, "condition_reason": "條件"},
            capability_result={"index_family": "machine", "n": 50}, status="draft",
        )
        db_session.add(version)
        db_session.commit()

        workbook = load_workbook(SpcReportService.generate_version_report(version.id))
        rows = _sheet_values(workbook, "機器績效摘要")

        assert ("資料來源語意", "未保存／不可稽核") in rows
        assert any(name == "限制說明" and "不可稽核" in value for name, value in rows)
