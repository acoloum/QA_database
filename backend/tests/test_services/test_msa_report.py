"""MSA 報告只讀不可變結果版本的測試。"""

from datetime import date, timedelta

import pytest
from openpyxl import load_workbook

from backend.extensions import db
from backend.models import (
    EquipmentCalibrationRecord,
    MeasurementEquipment,
    MsaCriteriaProfile,
    MsaCriteriaVersion,
    MsaResultVersion,
    User,
)
from backend.services.msa_analysis_service import MsaAnalysisService
from backend.services.msa_design_service import MsaDesignService
from backend.services.msa_errors import MsaNotFound, MsaValidationError
from backend.services.msa_observation_service import MsaObservationService
from backend.services.msa_report import SHEET_NAMES, MsaReportService
from backend.services.msa_study_service import MsaStudyService
from backend.services.msa_workflow_service import MsaWorkflowService


_READINGS = {
    ("P01", "A01"): [10.001, 10.002, 10.001],
    ("P01", "A02"): [10.003, 10.002, 10.003],
    ("P02", "A01"): [10.301, 10.302, 10.301],
    ("P02", "A02"): [10.303, 10.302, 10.303],
    ("P03", "A01"): [10.601, 10.602, 10.601],
    ("P03", "A02"): [10.603, 10.602, 10.603],
}


@pytest.fixture
def people(db_session):
    users = {
        name: User(username=f"msa_report_{name}", password="x")
        for name in ("creator", "executor", "approver")
    }
    db_session.add_all(users.values())
    db_session.commit()
    return users


def _analyzed(db_session, people):
    profile = MsaCriteriaProfile(name="報告準則")
    db_session.add(profile)
    db_session.flush()
    version = MsaCriteriaVersion(
        profile_id=profile.id, version_no=1, method_version="MSA4-1.0",
        effective_date=date(2026, 1, 1),
        thresholds={
            "grr_accept_max": 10.0, "grr_conditional_max": 30.0,
            "ndc_min": 5, "alpha": 0.05,
        },
        conditional_actions=["加嚴抽樣"],
        basis="AIAG MSA 第四版",
        status="approved",
    )
    db_session.add(version)
    db_session.flush()
    profile.current_version_id = version.id
    db_session.commit()

    equipment = MeasurementEquipment(
        equipment_no="EQ-REPORT-1", name="報告用量具", status="active",
        calibration_type="external", resolution="0.001", unit="mm",
    )
    db_session.add(equipment)
    db_session.flush()
    db_session.add(EquipmentCalibrationRecord(
        equipment_id=equipment.id, calibration_type="external",
        calibration_date=date.today() - timedelta(days=10),
        next_due_date=date.today() + timedelta(days=300),
        result="pass", status="approved",
    ))
    db_session.commit()

    study = MsaStudyService.create(
        {
            "study_type": "grr_anova",
            "measurement_purpose": "product_control",
            "characteristic": "外徑", "unit": "mm",
            "lsl": 9.5, "usl": 11.5,
            "criteria_profile_id": profile.id,
            "primary_executor_id": people["executor"].id,
        },
        actor_id=people["creator"].id,
    )
    plan = MsaDesignService.create_plan(
        study.id,
        {
            "method_code": "MSA4_GRR_ANOVA_1_0",
            "part_count": 3, "appraiser_count": 2, "trial_count": 3,
            "random_seed": 11, "percent_basis": "tolerance",
            "equipment": [
                {"equipment_id": equipment.id, "role": "primary_gauge"}
            ],
            "parts": [{"part_identifier": f"件-{i}"} for i in range(1, 4)],
            "appraisers": [{"name": "甲"}, {"name": "乙"}],
            "sampling_notes": "=SUM(A1)",
        },
        actor_id=people["creator"].id,
    )
    frozen = MsaDesignService.freeze(plan.id, actor_id=people["creator"].id)

    part_codes = {part.id: part.blind_code for part in frozen.parts}
    appraiser_codes = {r.id: r.blind_code for r in frozen.appraisers}
    for item in frozen.randomized_order:
        key = (
            part_codes[item["part_id"]], appraiser_codes[item["appraiser_id"]],
        )
        MsaObservationService.record(
            frozen.id,
            {
                "task_order": item["requested_order"],
                "numeric_value": str(_READINGS[key][item["trial_no"] - 1]),
            },
            actor_id=people["executor"].id, can_manage=True,
        )
    MsaObservationService.validate(frozen.id, actor_id=people["executor"].id)
    return MsaAnalysisService.analyze(
        frozen.id, actor_id=people["executor"].id,
    )


@pytest.fixture
def analyzed_msa_result(db_session, people):
    return _analyzed(db_session, people)


@pytest.fixture
def approved_msa_result(db_session, people, analyzed_msa_result):
    submitted = MsaWorkflowService.submit(
        analyzed_msa_result.id,
        {"expected_status": "analyzed", "reason": "送審"},
        actor_id=people["executor"].id,
    )
    return MsaWorkflowService.approve(
        submitted.id,
        {"expected_status": "submitted", "reason": "證據充分"},
        actor_id=people["approver"].id,
    )


def _sheet_values(workbook, sheet_name):
    return [
        tuple(cell.value for cell in row)
        for row in workbook[sheet_name].iter_rows()
    ]


# ---------------------------------------------------------------------------
# 只讀保存版本
# ---------------------------------------------------------------------------


def test_excel_rebuilds_only_from_saved_result_version(
    db_session, approved_msa_result,
):
    """報告產生後改動來源主檔，重建的報告必須逐格相同。"""
    before = load_workbook(
        MsaReportService.generate_excel(approved_msa_result.id)
    )

    equipment = MeasurementEquipment.query.filter_by(
        equipment_no="EQ-REPORT-1"
    ).one()
    equipment.name = "報告建立後更名的設備"
    db_session.commit()

    after = load_workbook(
        MsaReportService.generate_excel(approved_msa_result.id)
    )
    for sheet_name in before.sheetnames:
        assert _sheet_values(before, sheet_name) == _sheet_values(
            after, sheet_name
        ), sheet_name


def test_observations_come_from_the_snapshot_not_current_rows(
    db_session, approved_msa_result,
):
    """事後修正觀測不會改變已產生的報告讀值。"""
    before = _sheet_values(
        load_workbook(
            MsaReportService.generate_excel(approved_msa_result.id)
        ),
        "原始讀值",
    )

    from backend.models import MsaObservation

    original = MsaObservation.query.filter_by(
        plan_version_id=approved_msa_result.plan_version_id,
        is_effective=True,
    ).order_by(MsaObservation.id).first()
    MsaObservationService.correct(
        original.id,
        {"numeric_value": "99.999", "reason": "報告後才發現的輸入錯誤"},
        actor_id=approved_msa_result.created_by_id,
    )

    after = _sheet_values(
        load_workbook(
            MsaReportService.generate_excel(approved_msa_result.id)
        ),
        "原始讀值",
    )
    assert before == after
    assert not any("99.999" in str(cell) for row in after for cell in row)


# ---------------------------------------------------------------------------
# 工作表與稽核欄位
# ---------------------------------------------------------------------------


def test_excel_contains_auditable_sheets(db_session, approved_msa_result):
    workbook = load_workbook(
        MsaReportService.generate_excel(approved_msa_result.id),
        data_only=False,
    )

    assert workbook.sheetnames == list(SHEET_NAMES)
    audit = dict(
        row[:2] for row in _sheet_values(workbook, "版本稽核") if row[0]
    )
    assert audit["結果版本ID"] == approved_msa_result.id
    assert audit["資料雜湊"] == approved_msa_result.data_hash
    assert audit["方法代碼"] == approved_msa_result.method_code
    assert audit["方法版本"] == approved_msa_result.method_version
    assert audit["程式版本"] == approved_msa_result.code_version


def test_statistics_sheet_covers_the_saved_statistics(
    db_session, approved_msa_result,
):
    workbook = load_workbook(
        MsaReportService.generate_excel(approved_msa_result.id)
    )
    paths = {row[0] for row in _sheet_values(workbook, "統計結果") if row[0]}

    assert "grr" in paths
    assert "ndc" in paths
    assert "anova.selected_model" in paths
    assert "percent_tolerance.grr" in paths


def test_chart_sheet_carries_the_exact_series_used_on_screen(
    db_session, approved_msa_result,
):
    workbook = load_workbook(
        MsaReportService.generate_excel(approved_msa_result.id)
    )
    rows = {
        row[0]: row[1]
        for row in _sheet_values(workbook, "圖表資料") if row[0]
    }

    assert any(path.startswith("part_means") for path in rows)
    saved = approved_msa_result.chart_data["part_means"]
    for key, value in saved.items():
        assert rows[f"part_means.{key}"] == pytest.approx(value)


def test_equipment_sheet_reports_calibration_evidence(
    db_session, approved_msa_result,
):
    workbook = load_workbook(
        MsaReportService.generate_excel(approved_msa_result.id)
    )
    values = _sheet_values(workbook, "設備與校驗")
    flat = [str(cell) for row in values for cell in row if cell is not None]

    assert "EQ-REPORT-1" in flat
    assert "primary_gauge" in flat
    assert any("pass" == item for item in flat)


def test_criteria_sheet_reports_thresholds_and_percent_basis(
    db_session, approved_msa_result,
):
    workbook = load_workbook(
        MsaReportService.generate_excel(approved_msa_result.id)
    )
    values = dict(
        row[:2] for row in _sheet_values(workbook, "準則與判定") if row[0]
    )

    assert values["百分比口徑"] == "tolerance"
    assert values["口徑來源"] == "explicit"
    assert values["grr_accept_max"] == 10.0
    assert values["ndc_min"] == 5


def test_workflow_sheet_records_actor_time_reason_and_separation(
    db_session, approved_msa_result, people,
):
    workbook = load_workbook(
        MsaReportService.generate_excel(approved_msa_result.id)
    )
    rows = _sheet_values(workbook, "核准歷程")
    header = rows[0]
    approve_row = next(row for row in rows[1:] if row[0] == "approve")

    assert header[:3] == ("動作", "原狀態", "新狀態")
    assert approve_row[3] == people["approver"].username
    assert approve_row[5]
    assert approve_row[6] == "證據充分"
    assert "checked_actor_id" in str(approve_row[7])


# ---------------------------------------------------------------------------
# 未核准標示與注入防護
# ---------------------------------------------------------------------------


@pytest.mark.parametrize("status", ["analyzed", "submitted", "rejected"])
def test_unapproved_workbook_is_clearly_marked(
    db_session, people, analyzed_msa_result, status,
):
    if status != "analyzed":
        MsaWorkflowService.submit(
            analyzed_msa_result.id,
            {"expected_status": "analyzed", "reason": "送審"},
            actor_id=people["executor"].id,
        )
    if status == "rejected":
        MsaWorkflowService.reject(
            analyzed_msa_result.id,
            {"expected_status": "submitted", "reason": "退回"},
            actor_id=people["approver"].id,
        )

    workbook = load_workbook(
        MsaReportService.generate_excel(analyzed_msa_result.id)
    )
    flat = [
        str(cell) for row in _sheet_values(workbook, "研究摘要")
        for cell in row if cell is not None
    ]

    assert any("草稿／未核准" in item for item in flat)


def test_approved_workbook_has_no_unapproved_notice(
    db_session, approved_msa_result,
):
    workbook = load_workbook(
        MsaReportService.generate_excel(approved_msa_result.id)
    )
    flat = [
        str(cell) for row in _sheet_values(workbook, "研究摘要")
        for cell in row if cell is not None
    ]

    assert not any("草稿／未核准" in item for item in flat)


def test_formula_like_text_is_neutralised(db_session, approved_msa_result):
    """來源文字以 = 開頭時必須加上前導單引號，避免被當成公式。"""
    workbook = load_workbook(
        MsaReportService.generate_excel(approved_msa_result.id)
    )
    flat = [
        str(cell)
        for sheet in workbook.sheetnames
        for row in _sheet_values(workbook, sheet)
        for cell in row if cell is not None
    ]

    assert not any(item.startswith("=SUM(") for item in flat)


# ---------------------------------------------------------------------------
# 錯誤情形
# ---------------------------------------------------------------------------


def test_missing_result_is_reported_as_not_found(db_session):
    with pytest.raises(MsaNotFound) as error:
        MsaReportService.generate_excel(999999)

    assert error.value.code == "MSA_RESULT_NOT_FOUND"


def test_incomplete_evidence_is_refused_instead_of_recomputed(
    db_session, approved_msa_result,
):
    """快照不完整時必須拒絕出報告，不得回頭重算。"""
    # 直接以 core update 製造殘缺快照，繞過 ORM 的不可變保護
    table = MsaResultVersion.__table__
    db.session.execute(
        table.update()
        .where(table.c["識別碼"] == approved_msa_result.id)
        .values({"原始資料摘要": {"study": {}, "plan": {}}})
    )
    db.session.commit()
    db.session.expire_all()

    with pytest.raises(MsaValidationError) as error:
        MsaReportService.generate_excel(approved_msa_result.id)

    assert error.value.code == "MSA_REPORT_EVIDENCE_INCOMPLETE"
    assert "effective_observations" in error.value.details["missing"]


# ---------------------------------------------------------------------------
# PDF 報告
# ---------------------------------------------------------------------------


def _pdf_text(output):
    from pypdf import PdfReader

    return "\n".join(page.extract_text() or "" for page in PdfReader(output).pages)


def test_approved_pdf_has_no_draft_watermark(db_session, approved_msa_result):
    from pypdf import PdfReader

    output = MsaReportService.generate_pdf(approved_msa_result.id)
    reader = PdfReader(output)
    text = "\n".join(page.extract_text() or "" for page in reader.pages)

    assert approved_msa_result.study.study_no in text
    assert "未核准" not in text
    assert reader.metadata.title.startswith("MSA ")


@pytest.mark.parametrize("status", ["analyzed", "submitted"])
def test_unapproved_pdf_carries_the_watermark(
    db_session, people, analyzed_msa_result, status,
):
    if status == "submitted":
        MsaWorkflowService.submit(
            analyzed_msa_result.id,
            {"expected_status": "analyzed", "reason": "送審"},
            actor_id=people["executor"].id,
        )

    text = _pdf_text(MsaReportService.generate_pdf(analyzed_msa_result.id))

    assert "未核准" in text


def test_pdf_repeats_traceable_identifiers_on_every_page(
    db_session, approved_msa_result,
):
    from pypdf import PdfReader

    reader = PdfReader(MsaReportService.generate_pdf(approved_msa_result.id))
    short_hash = approved_msa_result.data_hash[:12]

    assert len(reader.pages) > 1
    for index, page in enumerate(reader.pages):
        text = page.extract_text() or ""
        assert approved_msa_result.study.study_no in text, index
        assert str(approved_msa_result.id) in text, index
        assert short_hash in text, index


def test_pdf_renders_traditional_chinese_extractably(
    db_session, approved_msa_result,
):
    text = _pdf_text(MsaReportService.generate_pdf(approved_msa_result.id))

    assert "量測系統分析報告" in text
    assert "核准與稽核歷程" in text
    assert "外徑" in text


def test_pdf_chart_values_come_from_the_saved_chart_data(
    db_session, approved_msa_result,
):
    text = _pdf_text(MsaReportService.generate_pdf(approved_msa_result.id))

    assert "圖表資料" in text
    saved = approved_msa_result.chart_data["part_means"]
    first_key = sorted(saved)[0]
    assert f"part_means.{first_key}" in text.replace("\n", "")


def test_pdf_contains_the_required_evidence_sections(
    db_session, approved_msa_result,
):
    text = _pdf_text(MsaReportService.generate_pdf(approved_msa_result.id))

    for section in (
        "研究設計與適用性", "設備、校驗與解析度", "原始資料摘要",
        "統計證據", "圖表資料", "準則與三層判定", "核准與稽核歷程",
    ):
        assert section in text, section


def test_pdf_reports_three_conclusion_layers(
    db_session, approved_msa_result,
):
    text = _pdf_text(MsaReportService.generate_pdf(approved_msa_result.id))

    assert "客觀統計結果" in text
    assert "系統處置" in text
    assert "人工工程判斷" in text


def test_missing_cjk_font_refuses_to_produce_a_garbled_pdf(
    db_session, approved_msa_result, monkeypatch,
):
    from reportlab.pdfbase import pdfmetrics

    import backend.services.msa_report as report_module

    monkeypatch.setattr(report_module, "CJK_FONT_CANDIDATES", ())
    monkeypatch.setattr(
        pdfmetrics, "getRegisteredFontNames", lambda: ["Helvetica"],
    )

    with pytest.raises(MsaValidationError) as error:
        MsaReportService.generate_pdf(approved_msa_result.id)

    assert error.value.code == "MSA_REPORT_FONT_MISSING"


def test_pdf_is_also_rebuilt_only_from_the_saved_result(
    db_session, approved_msa_result,
):
    before = _pdf_text(MsaReportService.generate_pdf(approved_msa_result.id))

    equipment = MeasurementEquipment.query.filter_by(
        equipment_no="EQ-REPORT-1"
    ).one()
    equipment.name = "PDF 產生後更名的設備"
    db_session.commit()

    after = _pdf_text(MsaReportService.generate_pdf(approved_msa_result.id))

    assert before == after
    assert "PDF 產生後更名的設備" not in after
