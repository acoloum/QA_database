"""CQI-9 爐溫測試模組測試"""
import pytest
from datetime import date
from backend.models import Furnace
from backend.services.pyrometry_service import PyrometryService


def test_furnace_add_and_get(app, db_session):
    with app.app_context():
        fid = PyrometryService.add_furnace({
            "爐號": "F-01", "名稱": "1號時效爐", "製程類型": "T6時效",
            "TUS點數": 12, "SAT點數": 2, "TUS允許公差": 10, "SAT允許誤差": 5,
        })
        assert fid is not None
        detail = PyrometryService.get_furnace(fid)
        assert detail["爐號"] == "F-01"
        assert detail["TUS點數"] == 12


def test_furnace_list_only_active_by_default(app, db_session):
    with app.app_context():
        a = PyrometryService.add_furnace({"爐號": "F-A", "名稱": "啟用爐"})
        b = PyrometryService.add_furnace({"爐號": "F-B", "名稱": "停用爐"})
        PyrometryService.update_furnace(b, {"爐號": "F-B", "名稱": "停用爐", "啟用狀態": False})
        rows = PyrometryService.list_furnaces(active_only=True)
        codes = [r["爐號"] for r in rows]
        assert "F-A" in codes and "F-B" not in codes


def _auth_header(client, db_session):
    """建立測試使用者並回傳 Authorization header"""
    from backend.models import User
    from backend.utils import hash_password, generate_token
    u = User(username="pyro_tester", password=hash_password("pw"), role="admin")
    db_session.add(u)
    db_session.commit()
    token = generate_token(u.id, u.username, u.role)
    return {"Authorization": f"Bearer {token}"}


def test_furnace_api_crud(client, db_session):
    headers = _auth_header(client, db_session)
    r = client.post("/api/pyrometry/furnaces", json={"爐號": "F-09", "名稱": "退火爐"}, headers=headers)
    assert r.status_code == 200
    fid = r.get_json()["id"]
    r = client.get("/api/pyrometry/furnaces", headers=headers)
    assert any(x["爐號"] == "F-09" for x in r.get_json()["data"])


def test_recorder_api_crud(client, db_session):
    headers = _auth_header(client, db_session)
    r = client.post("/api/pyrometry/recorders", json={
        "編號": "REC-1", "熱電偶補正值": -1.15,
        "校正點": [{"頻道": 1, "標準溫度": 500, "器差值": 0.1}],
    }, headers=headers)
    assert r.status_code == 200
    rid = r.get_json()["id"]
    r = client.get("/api/pyrometry/recorders", headers=headers)
    assert any(x["編號"] == "REC-1" for x in r.get_json()["data"])
    r = client.get(f"/api/pyrometry/recorders/{rid}", headers=headers)
    assert r.get_json()["data"]["校正點"][0]["頻道"] == 1


def test_thermocouple_api_crud(client, db_session):
    headers = _auth_header(client, db_session)
    r = client.post("/api/pyrometry/thermocouples", json={
        "編號": "220716", "型式": "TYPE K",
        "校正點": [{"標準溫度": 100, "器差值": 0.42}, {"標準溫度": 200, "器差值": 0.65}],
    }, headers=headers)
    assert r.status_code == 200
    tcid = r.get_json()["id"]
    r = client.get("/api/pyrometry/thermocouples", headers=headers)
    assert any(x["編號"] == "220716" for x in r.get_json()["data"])
    r = client.get(f"/api/pyrometry/thermocouples/{tcid}", headers=headers)
    assert len(r.get_json()["data"]["校正點"]) == 2


def test_corrections_endpoint(client, db_session):
    headers = _auth_header(client, db_session)
    client.post("/api/pyrometry/recorders", json={
        "編號": "REC-C", "熱電偶補正值": -1.15,
        "校正點": [{"頻道": 1, "標準溫度": 500, "器差值": 0.1},
                   {"頻道": 1, "標準溫度": 600, "器差值": 0.2}],
    }, headers=headers)
    r = client.get("/api/pyrometry/corrections?setpoint=520&type=TUS&count=1", headers=headers)
    assert r.status_code == 200
    assert r.get_json()["data"][0] == -1.27


def test_evaluate_tus_pass():
    """設定溫度180、公差±10：各點偏差皆在內 → 合格，並算均勻度極差"""
    points = [
        {"最高溫": 186, "最低溫": 178},
        {"最高溫": 183, "最低溫": 179},
    ]
    result = PyrometryService.evaluate_tus(setpoint=180, tolerance=10, points=points)
    assert result["是否合格"] is True
    assert result["TUS均勻度極差"] == 8        # 186 - 178
    assert result["TUS最大正偏差"] == 6         # 186 - 180
    assert result["TUS最大負偏差"] == -2        # 178 - 180
    assert result["points"][0]["最大偏差"] == 6
    assert result["points"][0]["是否合格"] is True


def test_evaluate_tus_fail():
    """某點最高溫191 → 偏差+11 超過±10 → 不合格"""
    points = [{"最高溫": 191, "最低溫": 175}]
    result = PyrometryService.evaluate_tus(setpoint=180, tolerance=10, points=points)
    assert result["是否合格"] is False
    assert result["points"][0]["是否合格"] is False


def test_evaluate_tus_applies_correction():
    """修正值補正：校正後溫度=量測值+修正值，再對設定溫度算偏差與均勻度"""
    points = [{"最高溫": 186, "最低溫": 178, "修正值": -2}]
    result = PyrometryService.evaluate_tus(setpoint=180, tolerance=10, points=points)
    # 校正後 184/176 → 偏差 +4/-4，取絕對值最大 → +4
    assert result["points"][0]["最大偏差"] == 4
    assert result["TUS最大正偏差"] == 4         # 184 - 180
    assert result["TUS最大負偏差"] == -4        # 176 - 180
    assert result["TUS均勻度極差"] == 8         # 184 - 176


def _make_recorder(tc=-1.15):
    return PyrometryService.add_recorder({
        "編號": "23B230004", "熱電偶補正值": tc,
        "校正點": [
            {"頻道": 1, "標準溫度": 500, "器差值": 0.1},
            {"頻道": 1, "標準溫度": 600, "器差值": 0.2},
            {"頻道": 13, "標準溫度": 500, "器差值": -0.2},
            {"頻道": 13, "標準溫度": 600, "器差值": -0.2},
        ],
    })


def test_compute_corrections_interpolates_tus(app, db_session):
    """TUS-1→CH1：設定520在500/600間內插器差值=0.12，修正值=-1.15-0.12=-1.27"""
    with app.app_context():
        _make_recorder()
        corr = PyrometryService.compute_corrections(setpoint=520, test_type="TUS", count=1)
        assert corr[0] == -1.27


def test_compute_corrections_uses_thermocouple_curve(app, db_session):
    """熱電偶補正改為內插曲線：修正值 = −(熱電偶器差值 + 記錄器器差值)"""
    with app.app_context():
        PyrometryService.add_recorder({
            "編號": "R-TC", "熱電偶補正值": -1.15,  # 舊固定值應被熱電偶曲線取代
            "校正點": [{"頻道": 1, "標準溫度": 200, "器差值": 0.1}],
        })
        PyrometryService.add_thermocouple({
            "編號": "220716", "型式": "TYPE K",
            "校正點": [{"標準溫度": 100, "器差值": 0.42}, {"標準溫度": 200, "器差值": 0.65}],
        })
        corr = PyrometryService.compute_corrections(setpoint=200, test_type="TUS", count=1)
        # −(0.65 + 0.1) = −0.75（不再用固定 −1.15）
        assert corr[0] == -0.75


def test_compute_corrections_sat_channel_offset_and_clamp(app, db_session):
    """SAT-1→CH13；設定溫度700超出校正範圍 → 夾擠取600端點器差值-0.2，修正值=-1.15+0.2=-0.95"""
    with app.app_context():
        _make_recorder()
        corr = PyrometryService.compute_corrections(setpoint=700, test_type="SAT", count=1)
        assert corr[0] == -0.95


def test_evaluate_sat_pass_and_fail():
    """SAT：多讀值格式；所有讀值均需在公差內；公差±5"""
    points = [
        {"控溫區": "Z1", "修正值": 0, "readings": [
            {"控制儀表讀值": 180, "校正測試讀值": 183},  # 差+3 OK
        ]},
        {"控溫區": "Z2", "修正值": 0, "readings": [
            {"控制儀表讀值": 180, "校正測試讀值": 183},  # OK
            {"控制儀表讀值": 180, "校正測試讀值": 187},  # 差+7 NG → zone NG
        ]},
    ]
    result = PyrometryService.evaluate_sat(tolerance=5, points=points)
    assert result["points"][0]["是否合格"] is True
    assert result["points"][0]["readings"][0]["差值"] == 3
    assert result["points"][1]["是否合格"] is False   # 有一筆超出
    assert result["是否合格"] is False


def _make_furnace(tol=10):
    return PyrometryService.add_furnace({"爐號": "F-T", "名稱": "測試爐", "TUS允許公差": tol})


def test_create_tus_test_auto_judges(app, db_session):
    with app.app_context():
        fid = _make_furnace(tol=10)
        tid = PyrometryService.create_test({
            "爐子ID": fid, "測試類型": "TUS", "測試日期": "2026-04-15",
            "設定溫度": 180, "允許公差": 10,
            "points": [{"點位": "P1", "最高溫": 186, "最低溫": 178},
                       {"點位": "P2", "最高溫": 183, "最低溫": 179}],
        })
        detail = PyrometryService.get_test(tid)
        assert detail["main"]["是否合格"] is True
        assert detail["main"]["季別"] == "2026Q2"          # 由日期自動帶
        assert detail["main"]["TUS均勻度極差"] == 8
        assert len(detail["tus_points"]) == 2


def test_create_test_persists_chart_data(app, db_session):
    """上傳的時間序列（曲線資料）存檔後可重現"""
    with app.app_context():
        fid = _make_furnace(tol=10)
        tid = PyrometryService.create_test({
            "爐子ID": fid, "測試類型": "TUS", "測試日期": "2026-04-15",
            "設定溫度": 180, "允許公差": 10,
            "points": [{"點位": "P1", "最高溫": 186, "最低溫": 178}],
            "曲線資料": {"時間": ["11:36", "11:38"], "數值": {"TUS-1": [178, 186]},
                       "穩定開始": 0, "穩定結束": 1},
        })
        detail = PyrometryService.get_test(tid)
        assert detail["曲線資料"]["時間"] == ["11:36", "11:38"]
        assert detail["曲線資料"]["數值"]["TUS-1"] == [178, 186]
        assert detail["曲線資料"]["穩定開始"] == 0


def test_create_test_persists_report_meta(app, db_session):
    """報告表頭欄位（客戶/料號/核准等）存檔後可重現"""
    with app.app_context():
        fid = _make_furnace(tol=10)
        tid = PyrometryService.create_test({
            "爐子ID": fid, "測試類型": "TUS", "測試日期": "2026-04-15",
            "設定溫度": 180, "允許公差": 10,
            "points": [{"點位": "P1", "最高溫": 186, "最低溫": 178}],
            "報告欄位": {"客戶名稱": "台積電", "工件料號": "ABC-123", "核准": "阮俊銓", "測試條件": "滿爐"},
        })
        detail = PyrometryService.get_test(tid)
        assert detail["報告欄位"]["客戶名稱"] == "台積電"
        assert detail["報告欄位"]["核准"] == "阮俊銓"


def test_create_sat_test_auto_judges(app, db_session):
    with app.app_context():
        fid = PyrometryService.add_furnace({"爐號": "F-S", "名稱": "退火爐", "SAT允許誤差": 5})
        tid = PyrometryService.create_test({
            "爐子ID": fid, "測試類型": "SAT", "測試日期": "2026-04-15",
            "設定溫度": 180, "允許公差": 5,
            "points": [{"控溫區": "Z1", "修正值": 0, "readings": [
                {"控制儀表讀值": 180, "校正測試讀值": 187},
            ]}],
        })
        detail = PyrometryService.get_test(tid)
        assert detail["main"]["是否合格"] is False
        assert len(detail["sat_points"]) == 1


def test_due_status_overdue(app, db_session):
    with app.app_context():
        fid = PyrometryService.add_furnace({"爐號": "F-D", "名稱": "到期測試爐", "TUS頻率_月": 3})
        # 最近一次 TUS 在 2025-01-01，每季 → 早已逾期（相對 today）
        PyrometryService.create_test({
            "爐子ID": fid, "測試類型": "TUS", "測試日期": "2025-01-01",
            "設定溫度": 180, "允許公差": 10, "points": [{"最高溫": 181, "最低溫": 179}]})
        status = PyrometryService.furnace_due_status(fid)
        assert status["TUS"]["下次應測日"] == "2025-04-01"
        assert status["TUS"]["狀態"] == "逾期"


def test_attachment_accepts_pyrometry(app, db_session):
    from backend.services.attachment_service import VALID_ENTITY_TYPES
    assert 'pyrometry' in VALID_ENTITY_TYPES


def test_parse_timeseries_csv_summary(tmp_path):
    from backend.services.pyrometry_parser import parse_temperature_file
    csv = tmp_path / "tus.csv"
    csv.write_text(
        "時間,TC-01,TC-02\n"
        "0,178,179\n"
        "1,186,183\n"
        "2,182,180\n", encoding="utf-8")
    with open(csv, "rb") as f:
        result = parse_temperature_file(f, filename="tus.csv")
    assert result["時間"] == ["0", "1", "2"]
    channels = {c["名稱"]: c for c in result["通道"]}
    assert channels["TC-01"]["最高溫"] == 186
    assert channels["TC-01"]["最低溫"] == 178
    assert channels["TC-02"]["最高溫"] == 183
    assert result["數值"]["TC-01"] == [178, 186, 182]


def test_parse_minute_labels_from_time_strings(tmp_path):
    """單時間欄：時間標籤輸出為分鐘級 HH:MM（非整段日期）"""
    from backend.services.pyrometry_parser import parse_temperature_file
    csv = tmp_path / "t.csv"
    csv.write_text("時間,TC-01\n11:36:00,100\n11:38:00,101\n", encoding="utf-8")
    with open(csv, "rb") as f:
        r = parse_temperature_file(f, "t.csv")
    assert r["時間"] == ["11:36", "11:38"]


def test_parse_dual_time_detected_by_content(tmp_path):
    """雙時間欄（日期+時刻）即使欄名不含關鍵字，仍依內容判定並輸出分鐘級標籤"""
    from backend.services.pyrometry_parser import parse_temperature_file
    csv = tmp_path / "d.csv"
    csv.write_text("D,Clock,TC-01\n2026/06/15,11:22:00,100\n2026/06/15,11:24:00,101\n", encoding="utf-8")
    with open(csv, "rb") as f:
        r = parse_temperature_file(f, "d.csv")
    assert r["時間"] == ["11:22", "11:24"]
    assert [c["名稱"] for c in r["通道"]] == ["TC-01"]


def test_parse_time_with_colon_milliseconds(tmp_path):
    """真實記錄器時間格式 HH:MM:SS:fff（毫秒以冒號分隔）→ 仍輸出分鐘級 HH:MM"""
    from backend.services.pyrometry_parser import parse_temperature_file
    csv = tmp_path / "pen.csv"
    csv.write_text(
        "日期,時間,TUS-1\n"
        "06/15/2026,09:57:39:521,29.3\n"
        "06/15/2026,09:59:39:521,32.2\n", encoding="utf-8")
    with open(csv, "rb") as f:
        r = parse_temperature_file(f, "pen.csv")
    assert r["時間"] == ["09:57", "09:59"]
    assert [c["名稱"] for c in r["通道"]] == ["TUS-1"]


import io

def test_parse_upload_route(client, db_session):
    headers = _auth_header(client, db_session)
    data = {
        "file": (io.BytesIO("時間,TC-01,TC-02\n0,178,179\n1,186,183\n".encode("utf-8")), "tus.csv"),
    }
    r = client.post("/api/pyrometry/parse-data", data=data,
                    headers=headers, content_type="multipart/form-data")
    assert r.status_code == 200
    body = r.get_json()
    assert body["success"] is True
    ch = {c["名稱"]: c for c in body["data"]["通道"]}
    assert ch["TC-01"]["最高溫"] == 186


def test_dashboard_lists_all_furnaces_with_due(app, db_session):
    with app.app_context():
        fid = PyrometryService.add_furnace({"爐號": "F-DB", "名稱": "看板爐", "TUS頻率_月": 3})
        PyrometryService.create_test({
            "爐子ID": fid, "測試類型": "TUS", "測試日期": "2025-01-01",
            "設定溫度": 180, "允許公差": 10, "points": [{"最高溫": 181, "最低溫": 179}]})
        board = PyrometryService.dashboard()
        row = next(r for r in board if r["爐號"] == "F-DB")
        assert row["TUS"]["狀態"] == "逾期"
        assert "最近結果" in row


def test_trend_returns_tus_history(app, db_session):
    with app.app_context():
        fid = PyrometryService.add_furnace({"爐號": "F-TR", "名稱": "趨勢爐"})
        for d, hi in [("2026-01-10", 185), ("2026-04-10", 188)]:
            PyrometryService.create_test({
                "爐子ID": fid, "測試類型": "TUS", "測試日期": d,
                "設定溫度": 180, "允許公差": 10,
                "points": [{"最高溫": hi, "最低溫": 178}]})
        trend = PyrometryService.tus_trend(fid)
        assert len(trend) == 2
        assert trend[0]["測試日期"] <= trend[1]["測試日期"]
        assert "均勻度極差" in trend[0]


def test_export_test_xlsx(app, db_session):
    """TUS 匯出：QRA073 表 + 原始數據曲線表，含報告欄位"""
    import io as _io
    from openpyxl import load_workbook
    with app.app_context():
        fid = PyrometryService.add_furnace({"爐號": "F-EX", "名稱": "匯出爐", "TUS允許公差": 10})
        tid = PyrometryService.create_test({
            "爐子ID": fid, "測試類型": "TUS", "測試日期": "2026-04-15",
            "設定溫度": 180, "允許公差": 10,
            "points": [{"點位": "P1", "最高溫": 186, "最低溫": 178}],
            "報告欄位": {"客戶名稱": "台積電", "核准": "阮俊銓"},
            "曲線資料": {"時間": ["11:36", "11:38"], "數值": {"TUS-1": [178, 186]},
                       "穩定開始": 0, "穩定結束": 1},
        })
        content = PyrometryService.export_test_xlsx(tid)
        assert content[:2] == b"PK"
        wb = load_workbook(_io.BytesIO(content))
        assert "QRA073-TUS均勻性" in wb.sheetnames
        assert "原始數據-記錄器" in wb.sheetnames
        ws = wb["QRA073-TUS均勻性"]
        assert ws["A1"].value == "必榮實業股份有限公司"
        # 報告欄位有帶入（客戶名稱）
        assert any(c.value == "台積電" for row in ws.iter_rows() for c in row)


def test_export_sat_xlsx_uses_qra074(app, db_session):
    import io as _io
    from openpyxl import load_workbook
    with app.app_context():
        fid = PyrometryService.add_furnace({"爐號": "F-ES", "名稱": "SAT匯出爐", "SAT允許誤差": 5})
        tid = PyrometryService.create_test({
            "爐子ID": fid, "測試類型": "SAT", "測試日期": "2026-04-15",
            "設定溫度": 180, "允許公差": 5,
            "points": [{"控溫區": "Z1", "修正值": 0, "readings": [
                {"控制儀表讀值": 180, "校正測試讀值": 183},
            ]}]})
        content = PyrometryService.export_test_xlsx(tid)
        wb = load_workbook(_io.BytesIO(content))
        assert "QRA074-SAT準確度" in wb.sheetnames
