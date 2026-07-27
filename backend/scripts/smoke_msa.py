"""對執行中的服務跑一次完整 MSA 流程，驗證 API、職責分離與報告。

刻意走 HTTP 而不是直接呼叫 service：要驗的正是「重啟後的服務真的
載入了新的 ORM 與 Blueprint」，直接呼叫 service 證明不了這件事。

用法：
    venv\\Scripts\\python.exe -m backend.scripts.smoke_msa \\
        --base-url http://localhost:5001 \\
        --manager-token <token> --executor-token <token> \\
        --approver-token <token>

帳號 token 一律由環境提供，不寫進腳本也不進版控。
"""

import argparse
import json
import os
import sys
from datetime import date, datetime, timedelta
from io import BytesIO
from urllib import error, request

# 報告檔案的魔術位元組；用來確認下載到的是真檔案而不是錯誤頁
PDF_MAGIC = b"%PDF-"
XLSX_MAGIC = b"PK\x03\x04"


class SmokeError(RuntimeError):
    """smoke 步驟失敗；訊息會直接印出供排查。"""


class Client:
    """最小 HTTP 客戶端；不依賴 requests，避免增加部署相依。"""

    def __init__(self, base_url: str, token: str):
        self.base_url = base_url.rstrip("/")
        self.token = token

    def _call(self, method: str, path: str, payload=None, raw=False):
        url = f"{self.base_url}{path}"
        data = None
        headers = {"Authorization": f"Bearer {self.token}"}
        if payload is not None:
            data = json.dumps(payload).encode("utf-8")
            headers["Content-Type"] = "application/json"

        req = request.Request(url, data=data, headers=headers, method=method)
        try:
            with request.urlopen(req, timeout=60) as response:
                body = response.read()
                if raw:
                    return response.status, body
                return response.status, json.loads(body or b"{}")
        except error.HTTPError as http_error:
            body = http_error.read()
            if raw:
                return http_error.code, body
            try:
                return http_error.code, json.loads(body or b"{}")
            except json.JSONDecodeError:
                return http_error.code, {"raw": body[:200].decode("utf-8", "replace")}

    def get(self, path, raw=False):
        return self._call("GET", path, raw=raw)

    def post(self, path, payload=None):
        return self._call("POST", path, payload)

    def patch(self, path, payload=None):
        return self._call("PATCH", path, payload)


def expect(status: int, wanted: int, step: str, body=None) -> None:
    if status != wanted:
        raise SmokeError(
            f"{step}：預期 HTTP {wanted}，實得 {status}\n"
            f"回應：{json.dumps(body, ensure_ascii=False)[:600]}"
        )


def run(base_url: str, tokens: dict, keep: bool) -> dict:
    stamp = datetime.now().strftime("%Y%m%dT%H%M%S")
    prefix = f"SMOKE-MSA-{stamp}"
    manager = Client(base_url, tokens["manager"])
    executor = Client(base_url, tokens["executor"])
    approver = Client(base_url, tokens["approver"])
    # 具核准權限但會參與資料輸入；用來驗證職責分離真的擋得住
    participant = Client(base_url, tokens["participant_approver"])
    evidence = {"prefix": prefix, "base_url": base_url, "steps": []}

    def record(step: str, detail):
        evidence["steps"].append({"step": step, "detail": detail})
        print(f"  ✓ {step}：{detail}")

    # 1. 設備與已核准校驗 --------------------------------------------
    status, body = manager.post("/api/measurement-equipment", {
        "equipment_no": f"{prefix}-EQ",
        "name": "smoke 量具",
        "status": "pending_review",
        "calibration_type": "external",
        "resolution": "0.001",
        "unit": "mm",
    })
    expect(status, 201, "建立設備", body)
    equipment_id = body["data"]["id"]
    record("建立設備", f"id={equipment_id}")

    status, body = manager.post(
        f"/api/measurement-equipment/{equipment_id}/calibrations",
        {
            "calibration_type": "external",
            "calibration_date": (date.today() - timedelta(days=5)).isoformat(),
            "next_due_date": (date.today() + timedelta(days=300)).isoformat(),
            "result": "pass",
            "correction_points": [],
        },
    )
    expect(status, 201, "建立校驗紀錄", body)
    calibration_id = body["data"]["id"]

    status, body = approver.post(
        f"/api/measurement-equipment/calibrations/{calibration_id}/approve",
        {"expected_status": "draft", "reason": "smoke 驗證"},
    )
    expect(status, 200, "核准校驗", body)

    status, body = manager.post(
        f"/api/measurement-equipment/{equipment_id}/status-events",
        {
            "event_type": "review_completed",
            "expected_status": "pending_review",
            "target_status": "active",
            "reason": "smoke 啟用",
        },
    )
    expect(status, 201, "啟用設備", body)
    record("設備校驗與啟用", f"calibration={calibration_id}")

    # 2. 判定準則 -----------------------------------------------------
    status, body = manager.post("/api/msa/criteria", {
        "name": f"{prefix}-準則",
        "applicable_study_types": ["grr_anova"],
    })
    expect(status, 201, "建立準則設定", body)
    profile_id = body["data"]["id"]

    status, body = manager.post(f"/api/msa/criteria/{profile_id}/versions", {
        "method_version": "MSA4-1.0",
        "effective_date": date.today().isoformat(),
        "thresholds": {
            "grr_accept_max": 10, "grr_conditional_max": 30, "ndc_min": 5,
        },
    })
    expect(status, 201, "建立準則版本", body)
    version_id = body["data"]["id"]

    status, body = approver.post(
        f"/api/msa/criteria/versions/{version_id}/approve",
        {"expected_status": "draft", "reason": "smoke 核准"},
    )
    expect(status, 200, "核准準則版本", body)
    record("判定準則", f"profile={profile_id} version={version_id}")

    # 3. 研究與計畫 ---------------------------------------------------
    status, body = executor.post("/api/msa/studies", {
        "study_type": "grr_anova",
        "measurement_purpose": "product_control",
        "characteristic": f"{prefix}-外徑",
        "unit": "mm",
        "lsl": 9.5,
        "usl": 11.5,
        "criteria_profile_id": profile_id,
    })
    expect(status, 201, "建立研究", body)
    study = body["data"]
    record("建立研究", study["study_no"])

    status, body = manager.post(f"/api/msa/studies/{study['id']}/plans", {
        "method_code": "MSA4_GRR_ANOVA_1_0",
        "part_count": 3, "appraiser_count": 2, "trial_count": 3,
        "random_seed": 20260728, "percent_basis": "tolerance",
        "equipment": [
            {"equipment_id": equipment_id, "role": "primary_gauge"},
        ],
        "parts": [{"part_identifier": f"件-{i}"} for i in range(1, 4)],
        "appraisers": [{"name": "甲"}, {"name": "乙"}],
    })
    expect(status, 201, "建立計畫", body)
    plan_id = body["data"]["id"]

    status, body = manager.post(f"/api/msa/plans/{plan_id}/freeze", {})
    expect(status, 200, "凍結計畫", body)
    plan_hash = body["data"]["plan_hash"]
    record("凍結計畫", f"plan={plan_id} hash={plan_hash[:12]}")

    # 4. 盲測資料 -----------------------------------------------------
    status, body = executor.get(f"/api/msa/plans/{plan_id}/tasks")
    expect(status, 200, "取得盲測任務", body)
    tasks = body["data"]["items"]
    if len(tasks) != 18:
        raise SmokeError(f"盲測任務數應為 18，實得 {len(tasks)}")

    for index, task in enumerate(tasks):
        value = _smoke_reading(
            task["part_blind_code"],
            task["appraiser_blind_code"],
            task["trial_no"],
        )
        # 讓具核准權限者也輸入一筆，才驗得到職責分離而不是權限不足
        writer = participant if index == 0 else manager
        status, body = writer.post(
            f"/api/msa/plans/{plan_id}/observations",
            {
                "task_order": task["requested_order"],
                "numeric_value": f"{value:.5f}",
            },
        )
        expect(status, 201, f"輸入讀值 {task['requested_order']}", body)
    record("輸入盲測讀值", f"{len(tasks)} 筆")

    status, body = executor.post(f"/api/msa/plans/{plan_id}/validate", {})
    expect(status, 200, "完整性驗證", body)
    if not body["data"]["complete"]:
        raise SmokeError(f"資料不完整：{body['data']['missing'][:5]}")
    record("完整性驗證", "complete=True")

    # 5. 分析 ---------------------------------------------------------
    status, body = executor.post(
        f"/api/msa/plans/{plan_id}/analyze",
        {"expected_plan_hash": plan_hash},
    )
    expect(status, 201, "分析", body)
    result = body["data"]
    record(
        "分析",
        f"result={result['id']} 判定={result['conclusion']['system_disposition']}"
        f" hash={result['data_hash'][:12]}",
    )

    # 6. 送審與職責分離 -----------------------------------------------
    submit_payload = {"expected_status": "analyzed", "reason": "smoke 送審"}
    if result["conclusion"]["system_disposition"] == "conditionally_acceptable":
        submit_payload["actions"] = ["smoke 條件接受措施"]
        submit_payload["due_date"] = (
            date.today() + timedelta(days=30)
        ).isoformat()
    status, body = manager.post(
        f"/api/msa/results/{result['id']}/submit", submit_payload,
    )
    expect(status, 200, "送審", body)
    record("送審", "analyzed → submitted")

    # 沒有核准權限者會先被權限層擋下，這條驗的是「權限不足」
    status, body = executor.post(
        f"/api/msa/results/{result['id']}/approve",
        {"expected_status": "submitted", "reason": "無權限核准"},
    )
    expect(status, 403, "權限阻擋", body)
    code = (body.get("error") or {}).get("code")
    if code != "MSA_PERMISSION_DENIED":
        raise SmokeError(f"預期 MSA_PERMISSION_DENIED，實得 {code}")
    record("權限阻擋", "無 msa.approve 者回 403 MSA_PERMISSION_DENIED")

    # 有核准權限但輸入過資料者，必須被職責分離擋下
    status, body = participant.post(
        f"/api/msa/results/{result['id']}/approve",
        {"expected_status": "submitted", "reason": "自己核准"},
    )
    expect(status, 403, "職責分離阻擋", body)
    code = (body.get("error") or {}).get("code")
    if code != "MSA_SELF_APPROVAL_FORBIDDEN":
        raise SmokeError(
            f"預期 MSA_SELF_APPROVAL_FORBIDDEN，實得 {code}\n"
            f"詳細：{json.dumps(body, ensure_ascii=False)[:400]}"
        )
    roles = (body.get("error") or {}).get("details", {}).get(
        "conflicting_roles", []
    )
    if "data_entry" not in roles:
        raise SmokeError(f"衝突角色應含 data_entry，實得 {roles}")
    record(
        "職責分離",
        "有核准權限但輸入過資料者回 403 MSA_SELF_APPROVAL_FORBIDDEN"
        f"（衝突角色 {roles}）",
    )

    status, body = approver.post(
        f"/api/msa/results/{result['id']}/approve",
        {"expected_status": "submitted", "reason": "smoke 核准"},
    )
    expect(status, 200, "獨立核准", body)
    record("獨立核准", "submitted → approved")

    # 7. 報告 ---------------------------------------------------------
    for suffix, magic, label in (
        ("pdf", PDF_MAGIC, "PDF"), ("xlsx", XLSX_MAGIC, "Excel"),
    ):
        status, blob = approver.get(
            f"/api/msa/results/{result['id']}/report.{suffix}", raw=True,
        )
        expect(status, 200, f"下載 {label}", None)
        if not blob.startswith(magic):
            raise SmokeError(
                f"{label} 檔頭不符：{blob[:8]!r}（可能下載到錯誤頁）"
            )
        if len(blob) < 2048:
            raise SmokeError(f"{label} 檔案過小：{len(blob)} bytes")
        _verify_report_content(suffix, blob, result)
        record(f"下載 {label}", f"{len(blob)} bytes，內容含稽核欄位")

    evidence["study_id"] = study["id"]
    evidence["study_no"] = study["study_no"]
    evidence["result_id"] = result["id"]
    evidence["data_hash"] = result["data_hash"]
    evidence["equipment_id"] = equipment_id

    if not keep:
        _void_evidence(approver, result, evidence, record)
    return evidence


def _smoke_reading(part_code: str, appraiser_code: str, trial: int) -> float:
    """產生零件差異大、但仍具真實變異結構的讀值。

    每個變異來源都必須非零，否則 ANOVA 的均方會是 0、F 值無定義，
    引擎會（正確地）判定模型退化而拒絕產生結果。
    """
    part_index = int(part_code[1:]) - 1
    appraiser_index = int(appraiser_code[1:]) - 1

    part_effect = 0.30 * part_index          # 零件間差異大 → ndc 足夠
    appraiser_effect = 0.0020 * appraiser_index
    # 非可加的零件×評價人項，讓交互作用均方不為 0
    interaction = 0.00040 * (((part_index * 3 + appraiser_index * 5) % 4) - 1.5)
    # 組內重複差異，讓重複性均方不為 0
    repeat = 0.00060 * (((trial + part_index + appraiser_index) % 3) - 1)
    return 10.0 + part_effect + appraiser_effect + interaction + repeat


def _verify_report_content(suffix: str, blob: bytes, result: dict) -> None:
    """確認報告帶有可追溯欄位，而不只是「有檔案」。"""
    if suffix == "xlsx":
        from openpyxl import load_workbook

        workbook = load_workbook(BytesIO(blob))
        if "版本稽核" not in workbook.sheetnames:
            raise SmokeError("Excel 缺少版本稽核工作表")
        values = {
            row[0].value: row[1].value
            for row in workbook["版本稽核"].iter_rows(max_col=2) if row[0].value
        }
        for field, wanted in (
            ("資料雜湊", result["data_hash"]),
            ("方法代碼", result["method_code"]),
            ("程式版本", result["code_version"]),
        ):
            if values.get(field) != wanted:
                raise SmokeError(
                    f"Excel {field} 不符：預期 {wanted}，實得 {values.get(field)}"
                )
        return

    try:
        from pypdf import PdfReader
    except ImportError:
        print("    （未安裝 pypdf，略過 PDF 內容抽取）")
        return

    reader = PdfReader(BytesIO(blob))
    text = "\n".join(page.extract_text() or "" for page in reader.pages)
    if result["data_hash"][:12] not in text:
        raise SmokeError("PDF 未帶資料雜湊短碼")
    if "量測系統分析報告" not in text:
        raise SmokeError("PDF 中文內容無法抽取，可能缺少中文字型")
    if "未核准" in text:
        raise SmokeError("已核准的報告不應出現未核准浮水印")


def _void_evidence(approver, result, evidence, record) -> None:
    """不可變證據不能刪除，依 runbook 標記作廢。"""
    status, body = approver.post(
        f"/api/msa/results/{result['id']}/void",
        {"expected_status": "approved", "reason": "smoke 測試資料，標記作廢"},
    )
    if status == 200:
        evidence["cleanup"] = "result voided"
        record("清理", "結果版本已標記作廢（不可變證據不刪除）")
    else:
        evidence["cleanup"] = f"void failed: {status}"
        record("清理", f"作廢失敗 HTTP {status}，請依 runbook 人工處理")


def main() -> int:
    parser = argparse.ArgumentParser(description="MSA 正式服務 smoke")
    parser.add_argument("--base-url", required=True)
    parser.add_argument(
        "--keep", action="store_true",
        help="保留 smoke 資料不標記作廢（僅供除錯）",
    )
    for role in ("manager", "executor", "approver", "participant-approver"):
        parser.add_argument(f"--{role}-token", default=None)
    args = parser.parse_args()

    tokens = {}
    for role in ("manager", "executor", "approver", "participant_approver"):
        token = (
            getattr(args, f"{role}_token")
            or os.environ.get(f"MSA_SMOKE_{role.upper()}_TOKEN")
        )
        if not token:
            parser.error(
                f"缺少 {role} token：請用 --{role}-token 或環境變數 "
                f"MSA_SMOKE_{role.upper()}_TOKEN 提供"
            )
        tokens[role] = token

    print(f"MSA smoke 開始：{args.base_url}")
    try:
        evidence = run(args.base_url, tokens, args.keep)
    except SmokeError as failure:
        print(f"\n✗ smoke 失敗：{failure}", file=sys.stderr)
        return 1
    except Exception as failure:  # noqa: BLE001 - smoke 需要完整錯誤內容
        print(f"\n✗ smoke 異常：{failure!r}", file=sys.stderr)
        return 2

    print("\n✓ MSA smoke 全部通過")
    print(json.dumps(evidence, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
